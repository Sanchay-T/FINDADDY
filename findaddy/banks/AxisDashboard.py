import pandas as pd
import numpy as np
from openpyxl import Workbook,load_workbook 
from openpyxl.styles import PatternFill,Border,Side
import xlwings as xw

w = pd.read_excel('Excel_Files/Dashboard/BankStatement.xlsx')

df = pd.DataFrame(w)
df1=df.rename(columns= {df.columns[0]:'Tran Date',df.columns[1]:'Chq No',df.columns[2]:'Particulars',df.columns[3]:'Debit',df.columns[4]:'Credit',df.columns[5]:'Balance',df.columns[6]:'Init. Br'})
df1=df1.iloc[2:]
df1['Debit'] = df1['Debit'].astype(float).round(2)
df1['Credit'] = df1['Credit'].astype(float).round(2)
df1['Balance'] = df1['Balance'].astype(float).round(2)
df1['Init. Br'] = df1['Init. Br'].astype(float).round(2)
df1['Tran Date'] = pd.to_datetime(df1['Tran Date'], format = '%d-%m-%Y')
df1['Categories'] = "suspense"

def general():
    #  row 1 => 
    Sum_of_credits = df1.groupby(df1['Tran Date'].dt.strftime('%b-%Y'))['Credit'].sum().sort_values()
    # print(Sum_of_credits)

    #  row =>
    Sum_of_debits = df1.groupby(df1['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
    # print(Sum_of_debits)

    #     row4 =>Total Amount of Cash Deposits
    Total_amount_withdrawals = df1[df1["Particulars"].str.contains("ATM-CASH")]
    Total_no_withdrawals = Total_amount_withdrawals.groupby(Total_amount_withdrawals['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
    Total_amount_withdrawals['Categories']='Withdrawal'
    df1.update(Total_amount_withdrawals)

    
    #     row 3 =>Total Amount of Cash Withdrawals
    Total_amount_cash_depo = df1[df1["Particulars"].str.contains("BY CASH DEPOSIT")|df1["Particulars"].str.contains("SAK")]
    Total_no_cash_depo = Total_amount_cash_depo.groupby(Total_amount_cash_depo['Tran Date'].dt.strftime('%b-%Y'))['Credit'].sum().sort_values()
    Total_amount_cash_depo['Categories']='Cash Deposits'
    df1.update(Total_amount_cash_depo)

    # pos_df = df1[df1["Particulars"].str.contains("POS/", na=False)]
    # POS_cr = pos_df.groupby(pos_df['Tran Date'].dt.strftime('%b-%Y'))['Credit'].sum().sort_values()
    # # print(POS_cr)
    # POS_dr = pos_df.groupby(pos_df['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
    # pos_dr = (df1[df1["Particulars"].str.contains("POS/")]
    #           .groupby('Debit')
    #           .apply(lambda x: x)
    #           .reset_index(drop=True)
    #           .assign(Categories = 'POS-dr'))
    # pos_cr = (df1[df1["Particulars"].str.contains("POS/")]
    #           .groupby('Credit')
    #           .apply(lambda x: x)
    #           .reset_index(drop=True)
    #           .assign(Categories = 'POS-cr'))
    # df1.update(pos_cr)
    # df1.update(pos_dr)

    Inv_details = df1[df1["Particulars"].str.contains("Zerodha -Dr", na=False)]
    Inv_dr = Inv_details.groupby(Inv_details['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
    # print(Inv_dr)

    wb = load_workbook('Excel_Files/Dashboard/BankStatement.xlsx')
    wb.create_sheet(index=0,title='summary')
    sumr = wb['summary']
    eod = wb['EOD Balance']
    
    #   Border:
    #   Medium border   
    medium_border = Border(left=Side(style='medium'),
                           right=Side(style='medium'),
                           top=Side(style='medium'),
                           bottom=Side(style='medium'))
    sumr.cell(row=2, column=2).border = medium_border
    sumr.cell(row=3, column=2).border = medium_border
    sumr.cell(row=4, column=2).border = medium_border
    sumr.cell(row=5, column=2).border = medium_border
    sumr.cell(row=6, column=2).border = medium_border
    sumr.cell(row=7, column=2).border = medium_border
    sumr.cell(row=8, column=2).border = medium_border
    sumr.cell(row=9, column=2).border = medium_border
    sumr.cell(row=10, column=2).border = medium_border
    sumr.cell(row=2, column=3).border = medium_border
    sumr.cell(row=3, column=3).border = medium_border
    sumr.cell(row=4, column=3).border = medium_border
    sumr.cell(row=5, column=3).border = medium_border
    sumr.cell(row=6, column=3).border = medium_border
    sumr.cell(row=7, column=3).border = medium_border
    sumr.cell(row=8, column=3).border = medium_border
    sumr.cell(row=9, column=3).border = medium_border
    sumr.cell(row=10, column=3).border = medium_border

    sumr.cell(row=13, column=2).border = medium_border
    sumr.cell(row=13, column=3).border = medium_border
    sumr.cell(row=13, column=4).border = medium_border
    sumr.cell(row=13, column=5).border = medium_border
    sumr.cell(row=13, column=6).border = medium_border
    sumr.cell(row=13, column=7).border = medium_border
    sumr.cell(row=13, column=8).border = medium_border
    sumr.cell(row=13, column=9).border = medium_border
    sumr.cell(row=13, column=10).border = medium_border
    sumr.cell(row=13, column=11).border = medium_border
    sumr.cell(row=13, column=12).border = medium_border
    sumr.cell(row=13, column=13).border = medium_border
    sumr.cell(row=13, column=14).border = medium_border

    sumr.cell(row=14, column=2).border = medium_border
    sumr.cell(row=15, column=2).border = medium_border
    sumr.cell(row=16, column=2).border = medium_border
    sumr.cell(row=17, column=2).border = medium_border
    sumr.cell(row=18, column=2).border = medium_border
    sumr.cell(row=19, column=2).border = medium_border
    sumr.cell(row=20, column=2).border = medium_border
    sumr.cell(row=21, column=2).border = medium_border
    sumr.cell(row=22, column=2).border = medium_border

    sumr.cell(row=25, column=2).border = medium_border
    sumr.cell(row=25, column=3).border = medium_border
    sumr.cell(row=25, column=4).border = medium_border
    sumr.cell(row=25, column=5).border = medium_border
    sumr.cell(row=25, column=6).border = medium_border
    sumr.cell(row=25, column=7).border = medium_border
    sumr.cell(row=25, column=8).border = medium_border
    sumr.cell(row=25, column=9).border = medium_border
    sumr.cell(row=25, column=10).border = medium_border
    sumr.cell(row=25, column=11).border = medium_border
    sumr.cell(row=25, column=12).border = medium_border
    sumr.cell(row=25, column=13).border = medium_border
    sumr.cell(row=25, column=14).border = medium_border

    sumr.cell(row=26, column=2).border = medium_border
    sumr.cell(row=27, column=2).border = medium_border
    sumr.cell(row=28, column=2).border = medium_border
    sumr.cell(row=29, column=2).border = medium_border
    sumr.cell(row=30, column=2).border = medium_border
    sumr.cell(row=31, column=2).border = medium_border
    sumr.cell(row=32, column=2).border = medium_border
    sumr.cell(row=33, column=2).border = medium_border
    sumr.cell(row=34, column=2).border = medium_border

    sumr.cell(row=36, column=2).border = medium_border
    sumr.cell(row=36, column=3).border = medium_border
    sumr.cell(row=36, column=4).border = medium_border
    sumr.cell(row=36, column=5).border = medium_border
    sumr.cell(row=36, column=6).border = medium_border
    sumr.cell(row=36, column=7).border = medium_border
    sumr.cell(row=36, column=8).border = medium_border
    sumr.cell(row=36, column=9).border = medium_border
    sumr.cell(row=36, column=10).border = medium_border
    sumr.cell(row=36, column=11).border = medium_border
    sumr.cell(row=36, column=12).border = medium_border
    sumr.cell(row=36, column=13).border = medium_border
    sumr.cell(row=36, column=14).border = medium_border

    sumr.cell(row=37, column=2).border = medium_border
    sumr.cell(row=38, column=2).border = medium_border
    sumr.cell(row=39, column=2).border = medium_border
    sumr.cell(row=40, column=2).border = medium_border
    sumr.cell(row=41, column=2).border = medium_border
    sumr.cell(row=42, column=2).border = medium_border
    sumr.cell(row=43, column=2).border = medium_border
    sumr.cell(row=44, column=2).border = medium_border
    sumr.cell(row=45, column=2).border = medium_border
    sumr.cell(row=46, column=2).border = medium_border
    sumr.cell(row=47, column=2).border = medium_border
    sumr.cell(row=48, column=2).border = medium_border

    sumr.cell(row=37, column=2).border = medium_border
    sumr.cell(row=38, column=2).border = medium_border
    sumr.cell(row=39, column=2).border = medium_border
    sumr.cell(row=40, column=2).border = medium_border
    sumr.cell(row=41, column=2).border = medium_border
    sumr.cell(row=42, column=2).border = medium_border
    sumr.cell(row=43, column=2).border = medium_border
    sumr.cell(row=44, column=2).border = medium_border
    sumr.cell(row=45, column=2).border = medium_border
    sumr.cell(row=46, column=2).border = medium_border
    sumr.cell(row=47, column=2).border = medium_border
    sumr.cell(row=48, column=2).border = medium_border

    sumr.cell(row=50, column=2).border = medium_border
    sumr.cell(row=51, column=2).border = medium_border
    sumr.cell(row=52, column=2).border = medium_border
    sumr.cell(row=53, column=2).border = medium_border
    sumr.cell(row=54, column=2).border = medium_border
    sumr.cell(row=55, column=2).border = medium_border
    sumr.cell(row=56, column=2).border = medium_border
    sumr.cell(row=57, column=2).border = medium_border
    sumr.cell(row=58, column=2).border = medium_border
    sumr.cell(row=59, column=2).border = medium_border
    sumr.cell(row=60, column=2).border = medium_border

    sumr.cell(row=50, column=2).border = medium_border
    sumr.cell(row=50, column=3).border = medium_border
    sumr.cell(row=50, column=4).border = medium_border
    sumr.cell(row=50, column=5).border = medium_border
    sumr.cell(row=50, column=6).border = medium_border
    sumr.cell(row=50, column=7).border = medium_border
    sumr.cell(row=50, column=8).border = medium_border
    sumr.cell(row=50, column=9).border = medium_border
    sumr.cell(row=50, column=10).border = medium_border
    sumr.cell(row=50, column=11).border = medium_border
    sumr.cell(row=50, column=12).border = medium_border
    sumr.cell(row=50, column=13).border = medium_border
    sumr.cell(row=50, column=14).border = medium_border

    sumr.cell(row=62, column=2).border = medium_border
    sumr.cell(row=62, column=3).border = medium_border
    sumr.cell(row=62, column=4).border = medium_border
    sumr.cell(row=62, column=5).border = medium_border
    sumr.cell(row=62, column=6).border = medium_border
    sumr.cell(row=62, column=7).border = medium_border
    sumr.cell(row=62, column=8).border = medium_border
    sumr.cell(row=62, column=9).border = medium_border
    sumr.cell(row=62, column=10).border = medium_border
    sumr.cell(row=62, column=11).border = medium_border
    sumr.cell(row=62, column=12).border = medium_border
    sumr.cell(row=62, column=13).border = medium_border
    sumr.cell(row=62, column=14).border = medium_border

    sumr.cell(row=62, column=2).border = medium_border
    sumr.cell(row=63, column=2).border = medium_border
    sumr.cell(row=64, column=2).border = medium_border
    sumr.cell(row=65, column=2).border = medium_border
    sumr.cell(row=66, column=2).border = medium_border
    sumr.cell(row=67, column=2).border = medium_border
    sumr.cell(row=68, column=2).border = medium_border
    sumr.cell(row=69, column=2).border = medium_border
    sumr.cell(row=70, column=2).border = medium_border
    sumr.cell(row=71, column=2).border = medium_border
    sumr.cell(row=72, column=2).border = medium_border
    sumr.cell(row=73, column=2).border = medium_border
    sumr.cell(row=74, column=2).border = medium_border
    sumr.cell(row=75, column=2).border = medium_border
    sumr.cell(row=76, column=2).border = medium_border
    sumr.cell(row=77, column=2).border = medium_border
    sumr.cell(row=78, column=2).border = medium_border
    sumr.cell(row=79, column=2).border = medium_border
    sumr.cell(row=80, column=2).border = medium_border
    sumr.cell(row=81, column=2).border = medium_border
    sumr.cell(row=82, column=2).border = medium_border
    sumr.cell(row=86, column=2).border = medium_border
    sumr.cell(row=87, column=2).border = medium_border
    sumr.cell(row=88, column=2).border = medium_border
    sumr.cell(row=89, column=2).border = medium_border
    sumr.cell(row=91, column=2).border = medium_border
    sumr.cell(row=92, column=2).border = medium_border
    sumr.cell(row=93, column=2).border = medium_border
    sumr.cell(row=94, column=2).border = medium_border
    

    sumr.cell(row=86, column=3).border = medium_border
    sumr.cell(row=86, column=4).border = medium_border
    sumr.cell(row=86, column=5).border = medium_border
    sumr.cell(row=86, column=6).border = medium_border
    sumr.cell(row=86, column=7).border = medium_border
    sumr.cell(row=86, column=8).border = medium_border
    sumr.cell(row=86, column=9).border = medium_border
    sumr.cell(row=86, column=10).border = medium_border
    sumr.cell(row=86, column=11).border = medium_border
    sumr.cell(row=86, column=12).border = medium_border
    sumr.cell(row=86, column=13).border = medium_border
    sumr.cell(row=86, column=14).border = medium_border

    sumr.cell(row=91, column=3).border = medium_border
    sumr.cell(row=91, column=4).border = medium_border
    sumr.cell(row=91, column=5).border = medium_border
    sumr.cell(row=91, column=6).border = medium_border
    sumr.cell(row=91, column=7).border = medium_border
    sumr.cell(row=91, column=8).border = medium_border
    sumr.cell(row=91, column=9).border = medium_border
    sumr.cell(row=91, column=10).border = medium_border
    sumr.cell(row=91, column=11).border = medium_border
    sumr.cell(row=91, column=12).border = medium_border
    sumr.cell(row=91, column=13).border = medium_border
    sumr.cell(row=91, column=14).border = medium_border

    #   Right medium 

    Rmedium_border = Border(left=Side(style='thin'),
                            right=Side(style='medium'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

    sumr.cell(row=14, column=14).border = Rmedium_border
    sumr.cell(row=15, column=14).border = Rmedium_border
    sumr.cell(row=16, column=14).border = Rmedium_border
    sumr.cell(row=17, column=14).border = Rmedium_border
    sumr.cell(row=18, column=14).border = Rmedium_border
    sumr.cell(row=19, column=14).border = Rmedium_border
    sumr.cell(row=20, column=14).border = Rmedium_border
    sumr.cell(row=21, column=14).border = Rmedium_border

    sumr.cell(row=26, column=14).border = Rmedium_border
    sumr.cell(row=27, column=14).border = Rmedium_border
    sumr.cell(row=28, column=14).border = Rmedium_border
    sumr.cell(row=29, column=14).border = Rmedium_border
    sumr.cell(row=30, column=14).border = Rmedium_border
    sumr.cell(row=31, column=14).border = Rmedium_border
    sumr.cell(row=32, column=14).border = Rmedium_border
    sumr.cell(row=33, column=14).border = Rmedium_border

    sumr.cell(row=37, column=14).border = Rmedium_border
    sumr.cell(row=38, column=14).border = Rmedium_border
    sumr.cell(row=39, column=14).border = Rmedium_border
    sumr.cell(row=40, column=14).border = Rmedium_border
    sumr.cell(row=41, column=14).border = Rmedium_border
    sumr.cell(row=42, column=14).border = Rmedium_border
    sumr.cell(row=43, column=14).border = Rmedium_border
    sumr.cell(row=44, column=14).border = Rmedium_border
    sumr.cell(row=45, column=14).border = Rmedium_border
    sumr.cell(row=46, column=14).border = Rmedium_border
    sumr.cell(row=47, column=14).border = Rmedium_border

    sumr.cell(row=37, column=14).border = Rmedium_border
    sumr.cell(row=38, column=14).border = Rmedium_border
    sumr.cell(row=39, column=14).border = Rmedium_border
    sumr.cell(row=40, column=14).border = Rmedium_border
    sumr.cell(row=41, column=14).border = Rmedium_border
    sumr.cell(row=42, column=14).border = Rmedium_border
    sumr.cell(row=43, column=14).border = Rmedium_border
    sumr.cell(row=44, column=14).border = Rmedium_border
    sumr.cell(row=45, column=14).border = Rmedium_border
    sumr.cell(row=46, column=14).border = Rmedium_border
    sumr.cell(row=47, column=14).border = Rmedium_border

    sumr.cell(row=37, column=14).border = Rmedium_border
    sumr.cell(row=38, column=14).border = Rmedium_border
    sumr.cell(row=39, column=14).border = Rmedium_border
    sumr.cell(row=40, column=14).border = Rmedium_border
    sumr.cell(row=41, column=14).border = Rmedium_border
    sumr.cell(row=42, column=14).border = Rmedium_border
    sumr.cell(row=43, column=14).border = Rmedium_border
    sumr.cell(row=44, column=14).border = Rmedium_border
    sumr.cell(row=45, column=14).border = Rmedium_border
    sumr.cell(row=46, column=14).border = Rmedium_border
    sumr.cell(row=47, column=14).border = Rmedium_border

    sumr.cell(row=51, column=14).border = Rmedium_border
    sumr.cell(row=52, column=14).border = Rmedium_border
    sumr.cell(row=53, column=14).border = Rmedium_border
    sumr.cell(row=54, column=14).border = Rmedium_border
    sumr.cell(row=55, column=14).border = Rmedium_border
    sumr.cell(row=56, column=14).border = Rmedium_border
    sumr.cell(row=57, column=14).border = Rmedium_border
    sumr.cell(row=58, column=14).border = Rmedium_border
    sumr.cell(row=59, column=14).border = Rmedium_border

    sumr.cell(row=63, column=14).border = Rmedium_border
    sumr.cell(row=64, column=14).border = Rmedium_border
    sumr.cell(row=65, column=14).border = Rmedium_border
    sumr.cell(row=66, column=14).border = Rmedium_border
    sumr.cell(row=67, column=14).border = Rmedium_border
    sumr.cell(row=68, column=14).border = Rmedium_border
    sumr.cell(row=69, column=14).border = Rmedium_border
    sumr.cell(row=70, column=14).border = Rmedium_border
    sumr.cell(row=71, column=14).border = Rmedium_border
    sumr.cell(row=72, column=14).border = Rmedium_border
    sumr.cell(row=73, column=14).border = Rmedium_border
    sumr.cell(row=74, column=14).border = Rmedium_border
    sumr.cell(row=75, column=14).border = Rmedium_border
    sumr.cell(row=76, column=14).border = Rmedium_border
    sumr.cell(row=77, column=14).border = Rmedium_border
    sumr.cell(row=78, column=14).border = Rmedium_border
    sumr.cell(row=79, column=14).border = Rmedium_border
    sumr.cell(row=80, column=14).border = Rmedium_border
    sumr.cell(row=81, column=14).border = Rmedium_border
    sumr.cell(row=87, column=14).border = Rmedium_border
    sumr.cell(row=88, column=14).border = Rmedium_border
    sumr.cell(row=92, column=14).border = Rmedium_border
    sumr.cell(row=93, column=14).border = Rmedium_border

    # double sided medium 
    Bmedium_border = Border(left=Side(style='thin'),
                            right=Side(style='medium'),
                            top=Side(style='thin'),
                            bottom=Side(style='medium'))

    sumr.cell(row=48, column=14).border = Bmedium_border
    sumr.cell(row=34, column=14).border = Bmedium_border
    sumr.cell(row=22, column=14).border = Bmedium_border
    sumr.cell(row=60, column=14).border = Bmedium_border
    sumr.cell(row=82, column=14).border = Bmedium_border
    sumr.cell(row=89, column=14).border = Bmedium_border
    sumr.cell(row=94, column=14).border = Bmedium_border

    # down side medium 
    Dmedium_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='medium'))

    sumr.cell(row=22, column=3).border = Dmedium_border
    sumr.cell(row=22, column=4).border = Dmedium_border
    sumr.cell(row=22, column=5).border = Dmedium_border
    sumr.cell(row=22, column=6).border = Dmedium_border
    sumr.cell(row=22, column=7).border = Dmedium_border
    sumr.cell(row=22, column=8).border = Dmedium_border
    sumr.cell(row=22, column=9).border = Dmedium_border
    sumr.cell(row=22, column=10).border = Dmedium_border
    sumr.cell(row=22, column=11).border = Dmedium_border
    sumr.cell(row=22, column=12).border = Dmedium_border
    sumr.cell(row=22, column=13).border = Dmedium_border

    sumr.cell(row=34, column=3).border = Dmedium_border
    sumr.cell(row=34, column=4).border = Dmedium_border
    sumr.cell(row=34, column=5).border = Dmedium_border
    sumr.cell(row=34, column=6).border = Dmedium_border
    sumr.cell(row=34, column=7).border = Dmedium_border
    sumr.cell(row=34, column=8).border = Dmedium_border
    sumr.cell(row=34, column=9).border = Dmedium_border
    sumr.cell(row=34, column=10).border = Dmedium_border
    sumr.cell(row=34, column=11).border = Dmedium_border
    sumr.cell(row=34, column=12).border = Dmedium_border
    sumr.cell(row=34, column=13).border = Dmedium_border

    sumr.cell(row=48, column=3).border = Dmedium_border
    sumr.cell(row=48, column=4).border = Dmedium_border
    sumr.cell(row=48, column=5).border = Dmedium_border
    sumr.cell(row=48, column=6).border = Dmedium_border
    sumr.cell(row=48, column=7).border = Dmedium_border
    sumr.cell(row=48, column=8).border = Dmedium_border
    sumr.cell(row=48, column=9).border = Dmedium_border
    sumr.cell(row=48, column=10).border = Dmedium_border
    sumr.cell(row=48, column=11).border = Dmedium_border
    sumr.cell(row=48, column=12).border = Dmedium_border
    sumr.cell(row=48, column=13).border = Dmedium_border

    sumr.cell(row=60, column=3).border = Dmedium_border
    sumr.cell(row=60, column=4).border = Dmedium_border
    sumr.cell(row=60, column=5).border = Dmedium_border
    sumr.cell(row=60, column=6).border = Dmedium_border
    sumr.cell(row=60, column=7).border = Dmedium_border
    sumr.cell(row=60, column=8).border = Dmedium_border
    sumr.cell(row=60, column=9).border = Dmedium_border
    sumr.cell(row=60, column=10).border = Dmedium_border
    sumr.cell(row=60, column=11).border = Dmedium_border
    sumr.cell(row=60, column=12).border = Dmedium_border
    sumr.cell(row=60, column=13).border = Dmedium_border

    sumr.cell(row=82, column=3).border = Dmedium_border
    sumr.cell(row=82, column=4).border = Dmedium_border
    sumr.cell(row=82, column=5).border = Dmedium_border
    sumr.cell(row=82, column=6).border = Dmedium_border
    sumr.cell(row=82, column=7).border = Dmedium_border
    sumr.cell(row=82, column=8).border = Dmedium_border
    sumr.cell(row=82, column=9).border = Dmedium_border
    sumr.cell(row=82, column=10).border = Dmedium_border
    sumr.cell(row=82, column=11).border = Dmedium_border
    sumr.cell(row=82, column=12).border = Dmedium_border
    sumr.cell(row=82, column=13).border = Dmedium_border


    sumr.cell(row=89, column=3).border = Dmedium_border
    sumr.cell(row=89, column=4).border = Dmedium_border
    sumr.cell(row=89, column=5).border = Dmedium_border
    sumr.cell(row=89, column=6).border = Dmedium_border
    sumr.cell(row=89, column=7).border = Dmedium_border
    sumr.cell(row=89, column=8).border = Dmedium_border
    sumr.cell(row=89, column=9).border = Dmedium_border
    sumr.cell(row=89, column=10).border = Dmedium_border
    sumr.cell(row=89, column=11).border = Dmedium_border
    sumr.cell(row=89, column=12).border = Dmedium_border
    sumr.cell(row=89, column=13).border = Dmedium_border

    sumr.cell(row=94, column=3).border = Dmedium_border
    sumr.cell(row=94, column=4).border = Dmedium_border
    sumr.cell(row=94, column=5).border = Dmedium_border
    sumr.cell(row=94, column=6).border = Dmedium_border
    sumr.cell(row=94, column=7).border = Dmedium_border
    sumr.cell(row=94, column=8).border = Dmedium_border
    sumr.cell(row=94, column=9).border = Dmedium_border
    sumr.cell(row=94, column=10).border = Dmedium_border
    sumr.cell(row=94, column=11).border = Dmedium_border
    sumr.cell(row=94, column=12).border = Dmedium_border
    sumr.cell(row=94, column=13).border = Dmedium_border

    # all side thin
    Athin_border = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))

    sumr.cell(row=14, column=3).border = Athin_border
    sumr.cell(row=15, column=3).border = Athin_border
    sumr.cell(row=16, column=3).border = Athin_border
    sumr.cell(row=17, column=3).border = Athin_border
    sumr.cell(row=18, column=3).border = Athin_border
    sumr.cell(row=19, column=3).border = Athin_border
    sumr.cell(row=20, column=3).border = Athin_border
    sumr.cell(row=21, column=3).border = Athin_border

    sumr.cell(row=14, column=4).border = Athin_border
    sumr.cell(row=15, column=4).border = Athin_border
    sumr.cell(row=16, column=4).border = Athin_border
    sumr.cell(row=17, column=4).border = Athin_border
    sumr.cell(row=18, column=4).border = Athin_border
    sumr.cell(row=19, column=4).border = Athin_border
    sumr.cell(row=20, column=4).border = Athin_border
    sumr.cell(row=21, column=4).border = Athin_border

    sumr.cell(row=14, column=5).border = Athin_border
    sumr.cell(row=15, column=5).border = Athin_border
    sumr.cell(row=16, column=5).border = Athin_border
    sumr.cell(row=17, column=5).border = Athin_border
    sumr.cell(row=18, column=5).border = Athin_border
    sumr.cell(row=19, column=5).border = Athin_border
    sumr.cell(row=20, column=5).border = Athin_border
    sumr.cell(row=21, column=5).border = Athin_border

    sumr.cell(row=14, column=6).border = Athin_border
    sumr.cell(row=15, column=6).border = Athin_border
    sumr.cell(row=16, column=6).border = Athin_border
    sumr.cell(row=17, column=6).border = Athin_border
    sumr.cell(row=18, column=6).border = Athin_border
    sumr.cell(row=19, column=6).border = Athin_border
    sumr.cell(row=20, column=6).border = Athin_border
    sumr.cell(row=21, column=6).border = Athin_border

    sumr.cell(row=14, column=7).border = Athin_border
    sumr.cell(row=15, column=7).border = Athin_border
    sumr.cell(row=16, column=7).border = Athin_border
    sumr.cell(row=17, column=7).border = Athin_border
    sumr.cell(row=18, column=7).border = Athin_border
    sumr.cell(row=19, column=7).border = Athin_border
    sumr.cell(row=20, column=7).border = Athin_border
    sumr.cell(row=21, column=7).border = Athin_border


    sumr.cell(row=14, column=8).border = Athin_border
    sumr.cell(row=15, column=8).border = Athin_border
    sumr.cell(row=16, column=8).border = Athin_border
    sumr.cell(row=17, column=8).border = Athin_border
    sumr.cell(row=18, column=8).border = Athin_border
    sumr.cell(row=19, column=8).border = Athin_border
    sumr.cell(row=20, column=8).border = Athin_border
    sumr.cell(row=21, column=8).border = Athin_border

    sumr.cell(row=14, column=9).border = Athin_border
    sumr.cell(row=15, column=9).border = Athin_border
    sumr.cell(row=16, column=9).border = Athin_border
    sumr.cell(row=17, column=9).border = Athin_border
    sumr.cell(row=18, column=9).border = Athin_border
    sumr.cell(row=19, column=9).border = Athin_border
    sumr.cell(row=20, column=9).border = Athin_border
    sumr.cell(row=21, column=9).border = Athin_border

    sumr.cell(row=14, column=10).border = Athin_border
    sumr.cell(row=15, column=10).border = Athin_border
    sumr.cell(row=16, column=10).border = Athin_border
    sumr.cell(row=17, column=10).border = Athin_border
    sumr.cell(row=18, column=10).border = Athin_border
    sumr.cell(row=19, column=10).border = Athin_border
    sumr.cell(row=20, column=10).border = Athin_border
    sumr.cell(row=21, column=10).border = Athin_border

    sumr.cell(row=14, column=11).border = Athin_border
    sumr.cell(row=15, column=11).border = Athin_border
    sumr.cell(row=16, column=11).border = Athin_border
    sumr.cell(row=17, column=11).border = Athin_border
    sumr.cell(row=18, column=11).border = Athin_border
    sumr.cell(row=19, column=11).border = Athin_border
    sumr.cell(row=20, column=11).border = Athin_border
    sumr.cell(row=21, column=11).border = Athin_border
    
    sumr.cell(row=14, column=12).border = Athin_border
    sumr.cell(row=15, column=12).border = Athin_border
    sumr.cell(row=16, column=12).border = Athin_border
    sumr.cell(row=17, column=12).border = Athin_border
    sumr.cell(row=18, column=12).border = Athin_border
    sumr.cell(row=19, column=12).border = Athin_border
    sumr.cell(row=20, column=12).border = Athin_border
    sumr.cell(row=21, column=12).border = Athin_border
    
    sumr.cell(row=14, column=13).border = Athin_border
    sumr.cell(row=15, column=13).border = Athin_border
    sumr.cell(row=16, column=13).border = Athin_border
    sumr.cell(row=17, column=13).border = Athin_border
    sumr.cell(row=18, column=13).border = Athin_border
    sumr.cell(row=19, column=13).border = Athin_border
    sumr.cell(row=20, column=13).border = Athin_border
    sumr.cell(row=21, column=13).border = Athin_border
    
    sumr.cell(row=26, column=3).border = Athin_border
    sumr.cell(row=27, column=3).border = Athin_border
    sumr.cell(row=28, column=3).border = Athin_border
    sumr.cell(row=29, column=3).border = Athin_border
    sumr.cell(row=30, column=3).border = Athin_border
    sumr.cell(row=31, column=3).border = Athin_border
    sumr.cell(row=32, column=3).border = Athin_border
    sumr.cell(row=33, column=3).border = Athin_border
    
    sumr.cell(row=26, column=4).border = Athin_border
    sumr.cell(row=27, column=4).border = Athin_border
    sumr.cell(row=28, column=4).border = Athin_border
    sumr.cell(row=29, column=4).border = Athin_border
    sumr.cell(row=30, column=4).border = Athin_border
    sumr.cell(row=31, column=4).border = Athin_border
    sumr.cell(row=32, column=4).border = Athin_border
    sumr.cell(row=33, column=4).border = Athin_border
    
    sumr.cell(row=26, column=5).border = Athin_border
    sumr.cell(row=27, column=5).border = Athin_border
    sumr.cell(row=28, column=5).border = Athin_border
    sumr.cell(row=29, column=5).border = Athin_border
    sumr.cell(row=30, column=5).border = Athin_border
    sumr.cell(row=31, column=5).border = Athin_border
    sumr.cell(row=32, column=5).border = Athin_border
    sumr.cell(row=33, column=5).border = Athin_border
    
    sumr.cell(row=26, column=6).border = Athin_border
    sumr.cell(row=27, column=6).border = Athin_border
    sumr.cell(row=28, column=6).border = Athin_border
    sumr.cell(row=29, column=6).border = Athin_border
    sumr.cell(row=30, column=6).border = Athin_border
    sumr.cell(row=31, column=6).border = Athin_border
    sumr.cell(row=32, column=6).border = Athin_border
    sumr.cell(row=33, column=6).border = Athin_border
    
    sumr.cell(row=26, column=7).border = Athin_border
    sumr.cell(row=27, column=7).border = Athin_border
    sumr.cell(row=28, column=7).border = Athin_border
    sumr.cell(row=29, column=7).border = Athin_border
    sumr.cell(row=30, column=7).border = Athin_border
    sumr.cell(row=31, column=7).border = Athin_border
    sumr.cell(row=32, column=7).border = Athin_border
    sumr.cell(row=33, column=7).border = Athin_border
    
    sumr.cell(row=26, column=8).border = Athin_border
    sumr.cell(row=27, column=8).border = Athin_border
    sumr.cell(row=28, column=8).border = Athin_border
    sumr.cell(row=29, column=8).border = Athin_border
    sumr.cell(row=30, column=8).border = Athin_border
    sumr.cell(row=31, column=8).border = Athin_border
    sumr.cell(row=32, column=8).border = Athin_border
    sumr.cell(row=33, column=8).border = Athin_border
    
    sumr.cell(row=26, column=9).border = Athin_border
    sumr.cell(row=27, column=9).border = Athin_border
    sumr.cell(row=28, column=9).border = Athin_border
    sumr.cell(row=29, column=9).border = Athin_border
    sumr.cell(row=30, column=9).border = Athin_border
    sumr.cell(row=31, column=9).border = Athin_border
    sumr.cell(row=32, column=9).border = Athin_border
    sumr.cell(row=33, column=9).border = Athin_border
    
    sumr.cell(row=26, column=10).border = Athin_border
    sumr.cell(row=27, column=10).border = Athin_border
    sumr.cell(row=28, column=10).border = Athin_border
    sumr.cell(row=29, column=10).border = Athin_border
    sumr.cell(row=30, column=10).border = Athin_border
    sumr.cell(row=31, column=10).border = Athin_border
    sumr.cell(row=32, column=10).border = Athin_border
    sumr.cell(row=33, column=10).border = Athin_border
    
    sumr.cell(row=26, column=11).border = Athin_border
    sumr.cell(row=27, column=11).border = Athin_border
    sumr.cell(row=28, column=11).border = Athin_border
    sumr.cell(row=29, column=11).border = Athin_border
    sumr.cell(row=30, column=11).border = Athin_border
    sumr.cell(row=31, column=11).border = Athin_border
    sumr.cell(row=32, column=11).border = Athin_border
    sumr.cell(row=33, column=11).border = Athin_border
    
    sumr.cell(row=26, column=12).border = Athin_border
    sumr.cell(row=27, column=12).border = Athin_border
    sumr.cell(row=28, column=12).border = Athin_border
    sumr.cell(row=29, column=12).border = Athin_border
    sumr.cell(row=30, column=12).border = Athin_border
    sumr.cell(row=31, column=12).border = Athin_border
    sumr.cell(row=32, column=12).border = Athin_border
    sumr.cell(row=33, column=12).border = Athin_border
    
    sumr.cell(row=26, column=13).border = Athin_border
    sumr.cell(row=27, column=13).border = Athin_border
    sumr.cell(row=28, column=13).border = Athin_border
    sumr.cell(row=29, column=13).border = Athin_border
    sumr.cell(row=30, column=13).border = Athin_border
    sumr.cell(row=31, column=13).border = Athin_border
    sumr.cell(row=32, column=13).border = Athin_border
    sumr.cell(row=33, column=13).border = Athin_border
    
    sumr.cell(row=37, column=13).border = Athin_border
    sumr.cell(row=38, column=13).border = Athin_border
    sumr.cell(row=39, column=13).border = Athin_border
    sumr.cell(row=40, column=13).border = Athin_border
    sumr.cell(row=41, column=13).border = Athin_border
    sumr.cell(row=42, column=13).border = Athin_border
    sumr.cell(row=43, column=13).border = Athin_border
    sumr.cell(row=44, column=13).border = Athin_border
    sumr.cell(row=45, column=13).border = Athin_border
    sumr.cell(row=46, column=13).border = Athin_border
    sumr.cell(row=47, column=13).border = Athin_border
    
    sumr.cell(row=37, column=3).border = Athin_border
    sumr.cell(row=38, column=3).border = Athin_border
    sumr.cell(row=39, column=3).border = Athin_border
    sumr.cell(row=40, column=3).border = Athin_border
    sumr.cell(row=41, column=3).border = Athin_border
    sumr.cell(row=42, column=3).border = Athin_border
    sumr.cell(row=43, column=3).border = Athin_border
    sumr.cell(row=44, column=3).border = Athin_border
    sumr.cell(row=45, column=3).border = Athin_border
    sumr.cell(row=46, column=3).border = Athin_border
    sumr.cell(row=47, column=3).border = Athin_border

    sumr.cell(row=37, column=4).border = Athin_border
    sumr.cell(row=38, column=4).border = Athin_border
    sumr.cell(row=39, column=4).border = Athin_border
    sumr.cell(row=40, column=4).border = Athin_border
    sumr.cell(row=41, column=4).border = Athin_border
    sumr.cell(row=42, column=4).border = Athin_border
    sumr.cell(row=43, column=4).border = Athin_border
    sumr.cell(row=44, column=4).border = Athin_border
    sumr.cell(row=45, column=4).border = Athin_border
    sumr.cell(row=46, column=4).border = Athin_border
    sumr.cell(row=47, column=4).border = Athin_border

    sumr.cell(row=37, column=5).border = Athin_border
    sumr.cell(row=38, column=5).border = Athin_border
    sumr.cell(row=39, column=5).border = Athin_border
    sumr.cell(row=40, column=5).border = Athin_border
    sumr.cell(row=41, column=5).border = Athin_border
    sumr.cell(row=42, column=5).border = Athin_border
    sumr.cell(row=43, column=5).border = Athin_border
    sumr.cell(row=44, column=5).border = Athin_border
    sumr.cell(row=45, column=5).border = Athin_border
    sumr.cell(row=46, column=5).border = Athin_border
    sumr.cell(row=47, column=5).border = Athin_border

    sumr.cell(row=37, column=6).border = Athin_border
    sumr.cell(row=38, column=6).border = Athin_border
    sumr.cell(row=39, column=6).border = Athin_border
    sumr.cell(row=40, column=6).border = Athin_border
    sumr.cell(row=41, column=6).border = Athin_border
    sumr.cell(row=42, column=6).border = Athin_border
    sumr.cell(row=43, column=6).border = Athin_border
    sumr.cell(row=44, column=6).border = Athin_border
    sumr.cell(row=45, column=6).border = Athin_border
    sumr.cell(row=46, column=6).border = Athin_border
    sumr.cell(row=47, column=6).border = Athin_border

    sumr.cell(row=37, column=7).border = Athin_border
    sumr.cell(row=38, column=7).border = Athin_border
    sumr.cell(row=39, column=7).border = Athin_border
    sumr.cell(row=40, column=7).border = Athin_border
    sumr.cell(row=41, column=7).border = Athin_border
    sumr.cell(row=42, column=7).border = Athin_border
    sumr.cell(row=43, column=7).border = Athin_border
    sumr.cell(row=44, column=7).border = Athin_border
    sumr.cell(row=45, column=7).border = Athin_border
    sumr.cell(row=46, column=7).border = Athin_border
    sumr.cell(row=47, column=7).border = Athin_border

    sumr.cell(row=37, column=8).border = Athin_border
    sumr.cell(row=38, column=8).border = Athin_border
    sumr.cell(row=39, column=8).border = Athin_border
    sumr.cell(row=40, column=8).border = Athin_border
    sumr.cell(row=41, column=8).border = Athin_border
    sumr.cell(row=42, column=8).border = Athin_border
    sumr.cell(row=43, column=8).border = Athin_border
    sumr.cell(row=44, column=8).border = Athin_border
    sumr.cell(row=45, column=8).border = Athin_border
    sumr.cell(row=46, column=8).border = Athin_border
    sumr.cell(row=47, column=8).border = Athin_border

    sumr.cell(row=37, column=9).border = Athin_border
    sumr.cell(row=38, column=9).border = Athin_border
    sumr.cell(row=39, column=9).border = Athin_border
    sumr.cell(row=40, column=9).border = Athin_border
    sumr.cell(row=41, column=9).border = Athin_border
    sumr.cell(row=42, column=9).border = Athin_border
    sumr.cell(row=43, column=9).border = Athin_border
    sumr.cell(row=44, column=9).border = Athin_border
    sumr.cell(row=45, column=9).border = Athin_border
    sumr.cell(row=46, column=9).border = Athin_border
    sumr.cell(row=47, column=9).border = Athin_border

    sumr.cell(row=37, column=10).border = Athin_border
    sumr.cell(row=38, column=10).border = Athin_border
    sumr.cell(row=39, column=10).border = Athin_border
    sumr.cell(row=40, column=10).border = Athin_border
    sumr.cell(row=41, column=10).border = Athin_border
    sumr.cell(row=42, column=10).border = Athin_border
    sumr.cell(row=43, column=10).border = Athin_border
    sumr.cell(row=44, column=10).border = Athin_border
    sumr.cell(row=45, column=10).border = Athin_border
    sumr.cell(row=46, column=10).border = Athin_border
    sumr.cell(row=47, column=10).border = Athin_border

    sumr.cell(row=37, column=11).border = Athin_border
    sumr.cell(row=38, column=11).border = Athin_border
    sumr.cell(row=39, column=11).border = Athin_border
    sumr.cell(row=40, column=11).border = Athin_border
    sumr.cell(row=41, column=11).border = Athin_border
    sumr.cell(row=42, column=11).border = Athin_border
    sumr.cell(row=43, column=11).border = Athin_border
    sumr.cell(row=44, column=11).border = Athin_border
    sumr.cell(row=45, column=11).border = Athin_border
    sumr.cell(row=46, column=11).border = Athin_border
    sumr.cell(row=47, column=11).border = Athin_border

    sumr.cell(row=37, column=12).border = Athin_border
    sumr.cell(row=38, column=12).border = Athin_border
    sumr.cell(row=39, column=12).border = Athin_border
    sumr.cell(row=40, column=12).border = Athin_border
    sumr.cell(row=41, column=12).border = Athin_border
    sumr.cell(row=42, column=12).border = Athin_border
    sumr.cell(row=43, column=12).border = Athin_border
    sumr.cell(row=44, column=12).border = Athin_border
    sumr.cell(row=45, column=12).border = Athin_border
    sumr.cell(row=46, column=12).border = Athin_border
    sumr.cell(row=47, column=12).border = Athin_border

    sumr.cell(row=51, column=3).border = Athin_border
    sumr.cell(row=52, column=3).border = Athin_border
    sumr.cell(row=53, column=3).border = Athin_border
    sumr.cell(row=54, column=3).border = Athin_border
    sumr.cell(row=55, column=3).border = Athin_border
    sumr.cell(row=56, column=3).border = Athin_border
    sumr.cell(row=57, column=3).border = Athin_border
    sumr.cell(row=58, column=3).border = Athin_border
    sumr.cell(row=59, column=3).border = Athin_border
    sumr.cell(row=51, column=4).border = Athin_border
    sumr.cell(row=52, column=4).border = Athin_border
    sumr.cell(row=53, column=4).border = Athin_border
    sumr.cell(row=54, column=4).border = Athin_border
    sumr.cell(row=55, column=4).border = Athin_border
    sumr.cell(row=56, column=4).border = Athin_border
    sumr.cell(row=57, column=4).border = Athin_border
    sumr.cell(row=58, column=4).border = Athin_border
    sumr.cell(row=59, column=4).border = Athin_border
    sumr.cell(row=51, column=5).border = Athin_border
    sumr.cell(row=52, column=5).border = Athin_border
    sumr.cell(row=53, column=5).border = Athin_border
    sumr.cell(row=54, column=5).border = Athin_border
    sumr.cell(row=55, column=5).border = Athin_border
    sumr.cell(row=56, column=5).border = Athin_border
    sumr.cell(row=57, column=5).border = Athin_border
    sumr.cell(row=58, column=5).border = Athin_border
    sumr.cell(row=59, column=5).border = Athin_border
    sumr.cell(row=51, column=6).border = Athin_border
    sumr.cell(row=52, column=6).border = Athin_border
    sumr.cell(row=53, column=6).border = Athin_border
    sumr.cell(row=54, column=6).border = Athin_border
    sumr.cell(row=55, column=6).border = Athin_border
    sumr.cell(row=56, column=6).border = Athin_border
    sumr.cell(row=57, column=6).border = Athin_border
    sumr.cell(row=58, column=6).border = Athin_border
    sumr.cell(row=59, column=6).border = Athin_border
    sumr.cell(row=51, column=7).border = Athin_border
    sumr.cell(row=52, column=7).border = Athin_border
    sumr.cell(row=53, column=7).border = Athin_border
    sumr.cell(row=54, column=7).border = Athin_border
    sumr.cell(row=55, column=7).border = Athin_border
    sumr.cell(row=56, column=7).border = Athin_border
    sumr.cell(row=57, column=7).border = Athin_border
    sumr.cell(row=58, column=7).border = Athin_border
    sumr.cell(row=59, column=7).border = Athin_border
    sumr.cell(row=51, column=8).border = Athin_border
    sumr.cell(row=52, column=8).border = Athin_border
    sumr.cell(row=53, column=8).border = Athin_border
    sumr.cell(row=54, column=8).border = Athin_border
    sumr.cell(row=55, column=8).border = Athin_border
    sumr.cell(row=56, column=8).border = Athin_border
    sumr.cell(row=57, column=8).border = Athin_border
    sumr.cell(row=58, column=8).border = Athin_border
    sumr.cell(row=59, column=8).border = Athin_border
    sumr.cell(row=51, column=9).border = Athin_border
    sumr.cell(row=52, column=9).border = Athin_border
    sumr.cell(row=53, column=9).border = Athin_border
    sumr.cell(row=54, column=9).border = Athin_border
    sumr.cell(row=55, column=9).border = Athin_border
    sumr.cell(row=56, column=9).border = Athin_border
    sumr.cell(row=57, column=9).border = Athin_border
    sumr.cell(row=58, column=9).border = Athin_border
    sumr.cell(row=59, column=9).border = Athin_border
    sumr.cell(row=51, column=10).border = Athin_border
    sumr.cell(row=52, column=10).border = Athin_border
    sumr.cell(row=53, column=10).border = Athin_border
    sumr.cell(row=54, column=10).border = Athin_border
    sumr.cell(row=55, column=10).border = Athin_border
    sumr.cell(row=56, column=10).border = Athin_border
    sumr.cell(row=57, column=10).border = Athin_border
    sumr.cell(row=58, column=10).border = Athin_border
    sumr.cell(row=59, column=10).border = Athin_border
    sumr.cell(row=51, column=11).border = Athin_border
    sumr.cell(row=52, column=11).border = Athin_border
    sumr.cell(row=53, column=11).border = Athin_border
    sumr.cell(row=54, column=11).border = Athin_border
    sumr.cell(row=55, column=11).border = Athin_border
    sumr.cell(row=56, column=11).border = Athin_border
    sumr.cell(row=57, column=11).border = Athin_border
    sumr.cell(row=58, column=11).border = Athin_border
    sumr.cell(row=59, column=11).border = Athin_border
    sumr.cell(row=51, column=12).border = Athin_border
    sumr.cell(row=52, column=12).border = Athin_border
    sumr.cell(row=53, column=12).border = Athin_border
    sumr.cell(row=54, column=12).border = Athin_border
    sumr.cell(row=55, column=12).border = Athin_border
    sumr.cell(row=56, column=12).border = Athin_border
    sumr.cell(row=57, column=12).border = Athin_border
    sumr.cell(row=58, column=12).border = Athin_border
    sumr.cell(row=59, column=12).border = Athin_border
    sumr.cell(row=51, column=13).border = Athin_border
    sumr.cell(row=52, column=13).border = Athin_border
    sumr.cell(row=53, column=13).border = Athin_border
    sumr.cell(row=54, column=13).border = Athin_border
    sumr.cell(row=55, column=13).border = Athin_border
    sumr.cell(row=56, column=13).border = Athin_border
    sumr.cell(row=57, column=13).border = Athin_border
    sumr.cell(row=58, column=13).border = Athin_border
    sumr.cell(row=59, column=13).border = Athin_border

    sumr.cell(row=63, column=3).border = Athin_border
    sumr.cell(row=64, column=3).border = Athin_border
    sumr.cell(row=65, column=3).border = Athin_border
    sumr.cell(row=66, column=3).border = Athin_border
    sumr.cell(row=67, column=3).border = Athin_border
    sumr.cell(row=68, column=3).border = Athin_border
    sumr.cell(row=69, column=3).border = Athin_border
    sumr.cell(row=70, column=3).border = Athin_border
    sumr.cell(row=71, column=3).border = Athin_border
    sumr.cell(row=72, column=3).border = Athin_border
    sumr.cell(row=73, column=3).border = Athin_border
    sumr.cell(row=74, column=3).border = Athin_border
    sumr.cell(row=75, column=3).border = Athin_border
    sumr.cell(row=76, column=3).border = Athin_border
    sumr.cell(row=77, column=3).border = Athin_border
    sumr.cell(row=78, column=3).border = Athin_border
    sumr.cell(row=79, column=3).border = Athin_border
    sumr.cell(row=81, column=3).border = Athin_border

    sumr.cell(row=63, column=4).border = Athin_border
    sumr.cell(row=64, column=4).border = Athin_border
    sumr.cell(row=65, column=4).border = Athin_border
    sumr.cell(row=66, column=4).border = Athin_border
    sumr.cell(row=67, column=4).border = Athin_border
    sumr.cell(row=68, column=4).border = Athin_border
    sumr.cell(row=69, column=4).border = Athin_border
    sumr.cell(row=70, column=4).border = Athin_border
    sumr.cell(row=71, column=4).border = Athin_border
    sumr.cell(row=72, column=4).border = Athin_border
    sumr.cell(row=73, column=4).border = Athin_border
    sumr.cell(row=74, column=4).border = Athin_border
    sumr.cell(row=75, column=4).border = Athin_border
    sumr.cell(row=76, column=4).border = Athin_border
    sumr.cell(row=77, column=4).border = Athin_border
    sumr.cell(row=78, column=4).border = Athin_border
    sumr.cell(row=79, column=4).border = Athin_border
    sumr.cell(row=81, column=4).border = Athin_border

    sumr.cell(row=63, column=5).border = Athin_border
    sumr.cell(row=64, column=5).border = Athin_border
    sumr.cell(row=65, column=5).border = Athin_border
    sumr.cell(row=66, column=5).border = Athin_border
    sumr.cell(row=67, column=5).border = Athin_border
    sumr.cell(row=68, column=5).border = Athin_border
    sumr.cell(row=69, column=5).border = Athin_border
    sumr.cell(row=70, column=5).border = Athin_border
    sumr.cell(row=71, column=5).border = Athin_border
    sumr.cell(row=72, column=5).border = Athin_border
    sumr.cell(row=73, column=5).border = Athin_border
    sumr.cell(row=74, column=5).border = Athin_border
    sumr.cell(row=75, column=5).border = Athin_border
    sumr.cell(row=76, column=5).border = Athin_border
    sumr.cell(row=77, column=5).border = Athin_border
    sumr.cell(row=78, column=5).border = Athin_border
    sumr.cell(row=79, column=5).border = Athin_border
    sumr.cell(row=81, column=5).border = Athin_border

    sumr.cell(row=63, column=6).border = Athin_border
    sumr.cell(row=64, column=6).border = Athin_border
    sumr.cell(row=65, column=6).border = Athin_border
    sumr.cell(row=66, column=6).border = Athin_border
    sumr.cell(row=67, column=6).border = Athin_border
    sumr.cell(row=68, column=6).border = Athin_border
    sumr.cell(row=69, column=6).border = Athin_border
    sumr.cell(row=70, column=6).border = Athin_border
    sumr.cell(row=71, column=6).border = Athin_border
    sumr.cell(row=72, column=6).border = Athin_border
    sumr.cell(row=73, column=6).border = Athin_border
    sumr.cell(row=74, column=6).border = Athin_border
    sumr.cell(row=75, column=6).border = Athin_border
    sumr.cell(row=76, column=6).border = Athin_border
    sumr.cell(row=77, column=6).border = Athin_border
    sumr.cell(row=78, column=6).border = Athin_border
    sumr.cell(row=79, column=6).border = Athin_border
    sumr.cell(row=81, column=6).border = Athin_border

    sumr.cell(row=63, column=7).border = Athin_border
    sumr.cell(row=64, column=7).border = Athin_border
    sumr.cell(row=65, column=7).border = Athin_border
    sumr.cell(row=66, column=7).border = Athin_border
    sumr.cell(row=67, column=7).border = Athin_border
    sumr.cell(row=68, column=7).border = Athin_border
    sumr.cell(row=69, column=7).border = Athin_border
    sumr.cell(row=70, column=7).border = Athin_border
    sumr.cell(row=71, column=7).border = Athin_border
    sumr.cell(row=72, column=7).border = Athin_border
    sumr.cell(row=73, column=7).border = Athin_border
    sumr.cell(row=74, column=7).border = Athin_border
    sumr.cell(row=75, column=7).border = Athin_border
    sumr.cell(row=76, column=7).border = Athin_border
    sumr.cell(row=77, column=7).border = Athin_border
    sumr.cell(row=78, column=7).border = Athin_border
    sumr.cell(row=79, column=7).border = Athin_border
    sumr.cell(row=81, column=7).border = Athin_border

    sumr.cell(row=63, column=8).border = Athin_border
    sumr.cell(row=64, column=8).border = Athin_border
    sumr.cell(row=65, column=8).border = Athin_border
    sumr.cell(row=66, column=8).border = Athin_border
    sumr.cell(row=67, column=8).border = Athin_border
    sumr.cell(row=68, column=8).border = Athin_border
    sumr.cell(row=69, column=8).border = Athin_border
    sumr.cell(row=70, column=8).border = Athin_border
    sumr.cell(row=71, column=8).border = Athin_border
    sumr.cell(row=72, column=8).border = Athin_border
    sumr.cell(row=73, column=8).border = Athin_border
    sumr.cell(row=74, column=8).border = Athin_border
    sumr.cell(row=75, column=8).border = Athin_border
    sumr.cell(row=76, column=8).border = Athin_border
    sumr.cell(row=77, column=8).border = Athin_border
    sumr.cell(row=78, column=8).border = Athin_border
    sumr.cell(row=79, column=8).border = Athin_border
    sumr.cell(row=81, column=8).border = Athin_border

    sumr.cell(row=63, column=9).border = Athin_border
    sumr.cell(row=64, column=9).border = Athin_border
    sumr.cell(row=65, column=9).border = Athin_border
    sumr.cell(row=66, column=9).border = Athin_border
    sumr.cell(row=67, column=9).border = Athin_border
    sumr.cell(row=68, column=9).border = Athin_border
    sumr.cell(row=69, column=9).border = Athin_border
    sumr.cell(row=70, column=9).border = Athin_border
    sumr.cell(row=71, column=9).border = Athin_border
    sumr.cell(row=72, column=9).border = Athin_border
    sumr.cell(row=73, column=9).border = Athin_border
    sumr.cell(row=74, column=9).border = Athin_border
    sumr.cell(row=75, column=9).border = Athin_border
    sumr.cell(row=76, column=9).border = Athin_border
    sumr.cell(row=77, column=9).border = Athin_border
    sumr.cell(row=78, column=9).border = Athin_border
    sumr.cell(row=79, column=9).border = Athin_border
    sumr.cell(row=81, column=9).border = Athin_border

    sumr.cell(row=63, column=10).border = Athin_border
    sumr.cell(row=64, column=10).border = Athin_border
    sumr.cell(row=65, column=10).border = Athin_border
    sumr.cell(row=66, column=10).border = Athin_border
    sumr.cell(row=67, column=10).border = Athin_border
    sumr.cell(row=68, column=10).border = Athin_border
    sumr.cell(row=69, column=10).border = Athin_border
    sumr.cell(row=70, column=10).border = Athin_border
    sumr.cell(row=71, column=10).border = Athin_border
    sumr.cell(row=72, column=10).border = Athin_border
    sumr.cell(row=73, column=10).border = Athin_border
    sumr.cell(row=74, column=10).border = Athin_border
    sumr.cell(row=75, column=10).border = Athin_border
    sumr.cell(row=76, column=10).border = Athin_border
    sumr.cell(row=77, column=10).border = Athin_border
    sumr.cell(row=78, column=10).border = Athin_border
    sumr.cell(row=79, column=10).border = Athin_border
    sumr.cell(row=81, column=10).border = Athin_border

    sumr.cell(row=63, column=11).border = Athin_border
    sumr.cell(row=64, column=11).border = Athin_border
    sumr.cell(row=65, column=11).border = Athin_border
    sumr.cell(row=66, column=11).border = Athin_border
    sumr.cell(row=67, column=11).border = Athin_border
    sumr.cell(row=68, column=11).border = Athin_border
    sumr.cell(row=69, column=11).border = Athin_border
    sumr.cell(row=70, column=11).border = Athin_border
    sumr.cell(row=71, column=11).border = Athin_border
    sumr.cell(row=72, column=11).border = Athin_border
    sumr.cell(row=73, column=11).border = Athin_border
    sumr.cell(row=74, column=11).border = Athin_border
    sumr.cell(row=75, column=11).border = Athin_border
    sumr.cell(row=76, column=11).border = Athin_border
    sumr.cell(row=77, column=11).border = Athin_border
    sumr.cell(row=78, column=11).border = Athin_border
    sumr.cell(row=79, column=11).border = Athin_border
    sumr.cell(row=81, column=11).border = Athin_border

    sumr.cell(row=63, column=12).border = Athin_border
    sumr.cell(row=64, column=12).border = Athin_border
    sumr.cell(row=65, column=12).border = Athin_border
    sumr.cell(row=66, column=12).border = Athin_border
    sumr.cell(row=67, column=12).border = Athin_border
    sumr.cell(row=68, column=12).border = Athin_border
    sumr.cell(row=69, column=12).border = Athin_border
    sumr.cell(row=70, column=12).border = Athin_border
    sumr.cell(row=71, column=12).border = Athin_border
    sumr.cell(row=72, column=12).border = Athin_border
    sumr.cell(row=73, column=12).border = Athin_border
    sumr.cell(row=74, column=12).border = Athin_border
    sumr.cell(row=75, column=12).border = Athin_border
    sumr.cell(row=76, column=12).border = Athin_border
    sumr.cell(row=77, column=12).border = Athin_border
    sumr.cell(row=78, column=12).border = Athin_border
    sumr.cell(row=79, column=12).border = Athin_border
    sumr.cell(row=81, column=12).border = Athin_border

    sumr.cell(row=63, column=13).border = Athin_border
    sumr.cell(row=64, column=13).border = Athin_border
    sumr.cell(row=65, column=13).border = Athin_border
    sumr.cell(row=66, column=13).border = Athin_border
    sumr.cell(row=67, column=13).border = Athin_border
    sumr.cell(row=68, column=13).border = Athin_border
    sumr.cell(row=69, column=13).border = Athin_border
    sumr.cell(row=70, column=13).border = Athin_border
    sumr.cell(row=71, column=13).border = Athin_border
    sumr.cell(row=72, column=13).border = Athin_border
    sumr.cell(row=73, column=13).border = Athin_border
    sumr.cell(row=74, column=13).border = Athin_border
    sumr.cell(row=75, column=13).border = Athin_border
    sumr.cell(row=76, column=13).border = Athin_border
    sumr.cell(row=77, column=13).border = Athin_border
    sumr.cell(row=78, column=13).border = Athin_border
    sumr.cell(row=79, column=13).border = Athin_border
    sumr.cell(row=81, column=13).border = Athin_border

    sumr.cell(row=87, column=3).border = Athin_border
    sumr.cell(row=87, column=4).border = Athin_border
    sumr.cell(row=87, column=5).border = Athin_border
    sumr.cell(row=87, column=6).border = Athin_border
    sumr.cell(row=87, column=7).border = Athin_border
    sumr.cell(row=87, column=8).border = Athin_border
    sumr.cell(row=87, column=9).border = Athin_border
    sumr.cell(row=87, column=10).border = Athin_border
    sumr.cell(row=87, column=11).border = Athin_border
    sumr.cell(row=87, column=12).border = Athin_border
    sumr.cell(row=87, column=13).border = Athin_border

    sumr.cell(row=88, column=3).border = Athin_border
    sumr.cell(row=88, column=4).border = Athin_border
    sumr.cell(row=88, column=5).border = Athin_border
    sumr.cell(row=88, column=6).border = Athin_border
    sumr.cell(row=88, column=7).border = Athin_border
    sumr.cell(row=88, column=8).border = Athin_border
    sumr.cell(row=88, column=9).border = Athin_border
    sumr.cell(row=88, column=10).border = Athin_border
    sumr.cell(row=88, column=11).border = Athin_border
    sumr.cell(row=88, column=12).border = Athin_border
    sumr.cell(row=88, column=13).border = Athin_border

    sumr.cell(row=92, column=3).border = Athin_border
    sumr.cell(row=92, column=4).border = Athin_border
    sumr.cell(row=92, column=5).border = Athin_border
    sumr.cell(row=92, column=6).border = Athin_border
    sumr.cell(row=92, column=7).border = Athin_border
    sumr.cell(row=92, column=8).border = Athin_border
    sumr.cell(row=92, column=9).border = Athin_border
    sumr.cell(row=92, column=10).border = Athin_border
    sumr.cell(row=92, column=11).border = Athin_border
    sumr.cell(row=92, column=12).border = Athin_border
    sumr.cell(row=92, column=13).border = Athin_border

    sumr.cell(row=93, column=3).border = Athin_border
    sumr.cell(row=93, column=4).border = Athin_border
    sumr.cell(row=93, column=5).border = Athin_border
    sumr.cell(row=93, column=6).border = Athin_border
    sumr.cell(row=93, column=7).border = Athin_border
    sumr.cell(row=93, column=8).border = Athin_border
    sumr.cell(row=93, column=9).border = Athin_border
    sumr.cell(row=93, column=10).border = Athin_border
    sumr.cell(row=93, column=11).border = Athin_border
    sumr.cell(row=93, column=12).border = Athin_border
    sumr.cell(row=93, column=13).border = Athin_border


    #   Fill Colour:
    fill_pattern_yellow = PatternFill(patternType='solid', fgColor='FDFDBD')
    sumr['B2'].fill = fill_pattern_yellow
    sumr['B3'].fill = fill_pattern_yellow
    sumr['B4'].fill = fill_pattern_yellow
    sumr['B5'].fill = fill_pattern_yellow
    sumr['B6'].fill = fill_pattern_yellow
    sumr['B7'].fill = fill_pattern_yellow
    sumr['B8'].fill = fill_pattern_yellow
    sumr['B9'].fill = fill_pattern_yellow
    sumr['B10'].fill = fill_pattern_yellow
    sumr['B14'].fill = fill_pattern_yellow
    sumr['B15'].fill = fill_pattern_yellow
    sumr['B16'].fill = fill_pattern_yellow
    sumr['B17'].fill = fill_pattern_yellow
    sumr['B18'].fill = fill_pattern_yellow
    sumr['B19'].fill = fill_pattern_yellow
    sumr['B20'].fill = fill_pattern_yellow
    sumr['B21'].fill = fill_pattern_yellow
    sumr['B22'].fill = fill_pattern_yellow
    sumr['B26'].fill = fill_pattern_yellow
    sumr['B27'].fill = fill_pattern_yellow
    sumr['B28'].fill = fill_pattern_yellow
    sumr['B29'].fill = fill_pattern_yellow
    sumr['B30'].fill = fill_pattern_yellow
    sumr['B31'].fill = fill_pattern_yellow
    sumr['B32'].fill = fill_pattern_yellow
    sumr['B33'].fill = fill_pattern_yellow
    sumr['B34'].fill = fill_pattern_yellow
    sumr['B37'].fill = fill_pattern_yellow
    sumr['B38'].fill = fill_pattern_yellow
    sumr['B39'].fill = fill_pattern_yellow
    sumr['B40'].fill = fill_pattern_yellow
    sumr['B41'].fill = fill_pattern_yellow
    sumr['B42'].fill = fill_pattern_yellow
    sumr['B43'].fill = fill_pattern_yellow
    sumr['B44'].fill = fill_pattern_yellow
    sumr['B45'].fill = fill_pattern_yellow
    sumr['B46'].fill = fill_pattern_yellow
    sumr['B47'].fill = fill_pattern_yellow
    sumr['B48'].fill = fill_pattern_yellow
    sumr['B51'].fill = fill_pattern_yellow
    sumr['B52'].fill = fill_pattern_yellow
    sumr['B53'].fill = fill_pattern_yellow
    sumr['B54'].fill = fill_pattern_yellow
    sumr['B55'].fill = fill_pattern_yellow
    sumr['B56'].fill = fill_pattern_yellow
    sumr['B57'].fill = fill_pattern_yellow
    sumr['B58'].fill = fill_pattern_yellow
    sumr['B59'].fill = fill_pattern_yellow
    sumr['B60'].fill = fill_pattern_yellow
    sumr['B62'].fill = fill_pattern_yellow
    sumr['B63'].fill = fill_pattern_yellow
    sumr['B64'].fill = fill_pattern_yellow
    sumr['B65'].fill = fill_pattern_yellow
    sumr['B66'].fill = fill_pattern_yellow
    sumr['B67'].fill = fill_pattern_yellow
    sumr['B68'].fill = fill_pattern_yellow
    sumr['B69'].fill = fill_pattern_yellow
    sumr['B70'].fill = fill_pattern_yellow
    sumr['B71'].fill = fill_pattern_yellow
    sumr['B72'].fill = fill_pattern_yellow
    sumr['B73'].fill = fill_pattern_yellow
    sumr['B74'].fill = fill_pattern_yellow
    sumr['B75'].fill = fill_pattern_yellow
    sumr['B76'].fill = fill_pattern_yellow
    sumr['B77'].fill = fill_pattern_yellow
    sumr['B78'].fill = fill_pattern_yellow
    sumr['B79'].fill = fill_pattern_yellow
    sumr['B80'].fill = fill_pattern_yellow
    sumr['B81'].fill = fill_pattern_yellow
    sumr['B82'].fill = fill_pattern_yellow

    sumr['B87'].fill = fill_pattern_yellow
    sumr['B88'].fill = fill_pattern_yellow
    sumr['B89'].fill = fill_pattern_yellow

    sumr['B92'].fill = fill_pattern_yellow
    sumr['B93'].fill = fill_pattern_yellow
    sumr['B94'].fill = fill_pattern_yellow

    fill_pattern_blue = PatternFill(patternType='solid', fgColor='B8E8FC')
    sumr['C2'].fill = fill_pattern_blue
    sumr['C3'].fill = fill_pattern_blue
    sumr['C4'].fill = fill_pattern_blue
    sumr['C5'].fill = fill_pattern_blue
    sumr['C6'].fill = fill_pattern_blue
    sumr['C7'].fill = fill_pattern_blue
    sumr['C8'].fill = fill_pattern_blue
    sumr['C9'].fill = fill_pattern_blue
    sumr['C10'].fill = fill_pattern_blue

    fill_pattern_Dblue = PatternFill(patternType='solid', fgColor='B1AFFF')
    sumr['B13'].fill = fill_pattern_Dblue
    sumr['C13'].fill = fill_pattern_Dblue
    sumr['D13'].fill = fill_pattern_Dblue
    sumr['E13'].fill = fill_pattern_Dblue
    sumr['F13'].fill = fill_pattern_Dblue
    sumr['G13'].fill = fill_pattern_Dblue
    sumr['H13'].fill = fill_pattern_Dblue
    sumr['I13'].fill = fill_pattern_Dblue
    sumr['J13'].fill = fill_pattern_Dblue
    sumr['K13'].fill = fill_pattern_Dblue
    sumr['L13'].fill = fill_pattern_Dblue
    sumr['M13'].fill = fill_pattern_Dblue
    sumr['N13'].fill = fill_pattern_Dblue

    sumr['B25'].fill = fill_pattern_Dblue
    sumr['C25'].fill = fill_pattern_Dblue
    sumr['D25'].fill = fill_pattern_Dblue
    sumr['E25'].fill = fill_pattern_Dblue
    sumr['F25'].fill = fill_pattern_Dblue
    sumr['G25'].fill = fill_pattern_Dblue
    sumr['H25'].fill = fill_pattern_Dblue
    sumr['I25'].fill = fill_pattern_Dblue
    sumr['J25'].fill = fill_pattern_Dblue
    sumr['K25'].fill = fill_pattern_Dblue
    sumr['L25'].fill = fill_pattern_Dblue
    sumr['M25'].fill = fill_pattern_Dblue
    sumr['N25'].fill = fill_pattern_Dblue

    sumr['B36'].fill = fill_pattern_Dblue
    sumr['C36'].fill = fill_pattern_Dblue
    sumr['D36'].fill = fill_pattern_Dblue
    sumr['E36'].fill = fill_pattern_Dblue
    sumr['F36'].fill = fill_pattern_Dblue
    sumr['G36'].fill = fill_pattern_Dblue
    sumr['H36'].fill = fill_pattern_Dblue
    sumr['I36'].fill = fill_pattern_Dblue
    sumr['J36'].fill = fill_pattern_Dblue
    sumr['K36'].fill = fill_pattern_Dblue
    sumr['L36'].fill = fill_pattern_Dblue
    sumr['M36'].fill = fill_pattern_Dblue
    sumr['N36'].fill = fill_pattern_Dblue

    sumr['B50'].fill = fill_pattern_Dblue
    sumr['C50'].fill = fill_pattern_Dblue
    sumr['D50'].fill = fill_pattern_Dblue
    sumr['E50'].fill = fill_pattern_Dblue
    sumr['F50'].fill = fill_pattern_Dblue
    sumr['G50'].fill = fill_pattern_Dblue
    sumr['H50'].fill = fill_pattern_Dblue
    sumr['I50'].fill = fill_pattern_Dblue
    sumr['J50'].fill = fill_pattern_Dblue
    sumr['K50'].fill = fill_pattern_Dblue
    sumr['L50'].fill = fill_pattern_Dblue
    sumr['M50'].fill = fill_pattern_Dblue
    sumr['N50'].fill = fill_pattern_Dblue

    sumr['B62'].fill = fill_pattern_Dblue
    sumr['C62'].fill = fill_pattern_Dblue
    sumr['D62'].fill = fill_pattern_Dblue
    sumr['E62'].fill = fill_pattern_Dblue
    sumr['F62'].fill = fill_pattern_Dblue
    sumr['G62'].fill = fill_pattern_Dblue
    sumr['H62'].fill = fill_pattern_Dblue
    sumr['I62'].fill = fill_pattern_Dblue
    sumr['J62'].fill = fill_pattern_Dblue
    sumr['K62'].fill = fill_pattern_Dblue
    sumr['L62'].fill = fill_pattern_Dblue
    sumr['M62'].fill = fill_pattern_Dblue
    sumr['N62'].fill = fill_pattern_Dblue

    sumr['B86'].fill = fill_pattern_Dblue
    sumr['C86'].fill = fill_pattern_Dblue
    sumr['D86'].fill = fill_pattern_Dblue
    sumr['E86'].fill = fill_pattern_Dblue
    sumr['F86'].fill = fill_pattern_Dblue
    sumr['G86'].fill = fill_pattern_Dblue
    sumr['H86'].fill = fill_pattern_Dblue
    sumr['I86'].fill = fill_pattern_Dblue
    sumr['J86'].fill = fill_pattern_Dblue
    sumr['K86'].fill = fill_pattern_Dblue
    sumr['L86'].fill = fill_pattern_Dblue
    sumr['M86'].fill = fill_pattern_Dblue
    sumr['N86'].fill = fill_pattern_Dblue

    sumr['B91'].fill = fill_pattern_Dblue
    sumr['C91'].fill = fill_pattern_Dblue
    sumr['D91'].fill = fill_pattern_Dblue
    sumr['E91'].fill = fill_pattern_Dblue
    sumr['F91'].fill = fill_pattern_Dblue
    sumr['G91'].fill = fill_pattern_Dblue
    sumr['H91'].fill = fill_pattern_Dblue
    sumr['I91'].fill = fill_pattern_Dblue
    sumr['J91'].fill = fill_pattern_Dblue
    sumr['K91'].fill = fill_pattern_Dblue
    sumr['L91'].fill = fill_pattern_Dblue
    sumr['M91'].fill = fill_pattern_Dblue
    sumr['N91'].fill = fill_pattern_Dblue

    #   creating name index :-
    sumr.cell(row=1,column=2,value = "Customer Details")
    sumr.cell(row=2,column=2,value = "Name of the Account Holder")
    sumr.cell(row=3,column=2,value = "Address")
    sumr.cell(row=4,column=2,value = "Name of the Bank")
    sumr.cell(row=5,column=2,value = "Account Number")
    sumr.cell(row=6,column=2,value = "Account Type")
    sumr.cell(row=7,column=2,value = "Joint Holder")
    sumr.cell(row=8,column=2,value = "Customer No")
    sumr.cell(row=9,column=2,value = "Overdraft Limit")
    sumr.cell(row=10,column=2,value ="")

    # Creating months columns-------->
    sumr.cell(row=13, column=3).value = eod.cell(row=1, column=2).value
    sumr.cell(row=13, column=4).value = eod.cell(row=1, column=3).value
    sumr.cell(row=13, column=5).value = eod.cell(row=1, column=4).value
    sumr.cell(row=13, column=6).value = eod.cell(row=1, column=5).value
    sumr.cell(row=13, column=7).value = eod.cell(row=1, column=6).value
    sumr.cell(row=13, column=8).value = eod.cell(row=1, column=7).value
    sumr.cell(row=13, column=9).value = eod.cell(row=1, column=8).value
    sumr.cell(row=13, column=10).value = eod.cell(row=1, column=9).value
    sumr.cell(row=13, column=11).value = eod.cell(row=1, column=10).value
    sumr.cell(row=13, column=12).value = eod.cell(row=1, column=11).value
    sumr.cell(row=13, column=13).value = eod.cell(row=1, column=12).value
    sumr.cell(row=13, column=14).value = eod.cell(row=1, column=13).value

    sumr.cell(row=14, column=2, value='Total Amount of Credit Transactions')
    sumr.cell(row=15, column=2, value='Total Amount of Debit Transactions')
    sumr.cell(row=16, column=2, value='Total Amount of Cash Withdrawals')
    sumr.cell(row=17, column=2, value='Total Amount of Cash Deposits')
    sumr.cell(row=18, column=2, value='POS Txns - Cr')
    sumr.cell(row=19, column=2, value='Investment Details')
    sumr.cell(row=20, column=2, value='POS Txns - Dr')
    sumr.cell(row=21, column=2, value='Opening Balance')
    sumr.cell(row=22, column=2, value='Closing Balance')

    #    creating name index :-
    sumr.cell(row=25, column=2, value="Income")
    sumr.cell(row=26, column=2, value="Total Amount of Credit Transactions")
    sumr.cell(row=27, column=2, value="Bank Intrest Received")
    sumr.cell(row=28, column=2, value="Salary Received")
    sumr.cell(row=29, column=2, value="NACH Receipts")
    sumr.cell(row=30, column=2, value="Loans Received")
    sumr.cell(row=31, column=2, value="Income Tax Refund")
    sumr.cell(row=32, column=2, value="Dividend")
    sumr.cell(row=33, column=2, value="Rent Received")
    sumr.cell(row=34, column=2, value="Total Income")

    #   Personal index 

    # Creating months columns-------->
    sumr.cell(row=25, column=3).value = eod.cell(row=1, column=2).value
    sumr.cell(row=25, column=4).value = eod.cell(row=1, column=3).value
    sumr.cell(row=25, column=5).value = eod.cell(row=1, column=4).value
    sumr.cell(row=25, column=6).value = eod.cell(row=1, column=5).value
    sumr.cell(row=25, column=7).value = eod.cell(row=1, column=6).value
    sumr.cell(row=25, column=8).value = eod.cell(row=1, column=7).value
    sumr.cell(row=25, column=9).value = eod.cell(row=1, column=8).value
    sumr.cell(row=25, column=10).value = eod.cell(row=1, column=9).value
    sumr.cell(row=25, column=11).value = eod.cell(row=1, column=10).value
    sumr.cell(row=25, column=12).value = eod.cell(row=1, column=11).value
    sumr.cell(row=25, column=13).value = eod.cell(row=1, column=12).value
    sumr.cell(row=25, column=14).value = eod.cell(row=1, column=13).value

    sumr.cell(row=36, column=2, value="Expenditure")
    sumr.cell(row=37, column=2, value="Total Amount of Debit Transactions")
    sumr.cell(row=38, column=2, value="Bank Interest Paid (Only in OD/CC A/c)")
    sumr.cell(row=39, column=2, value="Salaries Paid")
    sumr.cell(row=40, column=2, value="Bank Charges")
    sumr.cell(row=41, column=2, value="EMI***")
    sumr.cell(row=42, column=2, value="TDS Deducted")
    sumr.cell(row=43, column=2, value="Total GST")
    sumr.cell(row=44, column=2, value="Total Income Tax Paid")
    sumr.cell(row=45, column=2, value="Utility Bills")
    sumr.cell(row=46, column=2, value="Travelling Expense")
    sumr.cell(row=47, column=2, value="Rent Paid")
    sumr.cell(row=48, column=2, value="Total Expenses")

    sumr.cell(row=50, column=2, value="Personal Expenses")
    sumr.cell(row=51, column=2, value="General Insurance")
    sumr.cell(row=52, column=2, value="Life Insurance")
    sumr.cell(row=53, column=2, value="Food Expenses")
    sumr.cell(row=54, column=2, value="Credit Card Payment")
    sumr.cell(row=55, column=2, value="Online Shopping")
    sumr.cell(row=56, column=2, value="Property Tax")
    sumr.cell(row=57, column=2, value="Gas payments")
    sumr.cell(row=58, column=2, value="Gold Loan (Only Interest)")
    sumr.cell(row=59, column=2, value="Rent Paid")
    sumr.cell(row=60, column=2, value="Total Amount")

    sumr.cell(row=62, column=2, value="Expenditure")
    sumr.cell(row=63, column=2, value="Total Amount of Debit Transactions")
    sumr.cell(row=64, column=2, value="Bank Interest Paid (Only in OD/CC A/c)")
    sumr.cell(row=65, column=2, value="Salaries Paid")
    sumr.cell(row=66, column=2, value="Bank Charges")
    sumr.cell(row=67, column=2, value="EMI***")
    sumr.cell(row=68, column=2, value="TDS Deducted")
    sumr.cell(row=69, column=2, value="Total GST")
    sumr.cell(row=70, column=2, value="Total Income Tax Paid)")
    sumr.cell(row=71, column=2, value="Utility Bills")
    sumr.cell(row=72, column=2, value="Travelling Expense")
    sumr.cell(row=73, column=2, value="General Insurance")
    sumr.cell(row=74, column=2, value="Life Insurance")
    sumr.cell(row=75, column=2, value="Food Expenses")
    sumr.cell(row=76, column=2, value="Credit Card Payment")
    sumr.cell(row=77, column=2, value="Online Shopping")
    sumr.cell(row=78, column=2, value="Property Tax")
    sumr.cell(row=79, column=2, value="Gas payments")
    sumr.cell(row=80, column=2, value="Gold Loan (Only Interest)")
    sumr.cell(row=81, column=2, value="Rent Paid")
    sumr.cell(row=82, column=2, value="Total Expenses")
    sumr.cell(row=86, column=2, value="Profit with Personal Income")
    sumr.cell(row=87, column=2, value="Total Sales (receipts)")
    sumr.cell(row=88, column=2, value="Total Purchase (payments)")
    sumr.cell(row=89, column=2, value="Total Profit as per Bank A/c")
    sumr.cell(row=91, column=2, value="Profit without Personal Income")
    sumr.cell(row=92, column=2, value="Total Sales (receipts)")
    sumr.cell(row=93, column=2, value="Total Purchase (payments)")
    sumr.cell(row=94, column=2, value="Total Profit as per Bank A/c")


    # Creating months columns-------->
    sumr.cell(row=36, column=3).value = eod.cell(row=1, column=2).value
    sumr.cell(row=36, column=4).value = eod.cell(row=1, column=3).value
    sumr.cell(row=36, column=5).value = eod.cell(row=1, column=4).value
    sumr.cell(row=36, column=6).value = eod.cell(row=1, column=5).value
    sumr.cell(row=36, column=7).value = eod.cell(row=1, column=6).value
    sumr.cell(row=36, column=8).value = eod.cell(row=1, column=7).value
    sumr.cell(row=36, column=9).value = eod.cell(row=1, column=8).value
    sumr.cell(row=36, column=10).value = eod.cell(row=1, column=9).value
    sumr.cell(row=36, column=11).value = eod.cell(row=1, column=10).value
    sumr.cell(row=36, column=12).value = eod.cell(row=1, column=11).value
    sumr.cell(row=36, column=13).value = eod.cell(row=1, column=12).value
    sumr.cell(row=36, column=14).value = eod.cell(row=1, column=13).value

    # Creating months columns-------->
    sumr.cell(row=50, column=3).value = eod.cell(row=1, column=2).value
    sumr.cell(row=50, column=4).value = eod.cell(row=1, column=3).value
    sumr.cell(row=50, column=5).value = eod.cell(row=1, column=4).value
    sumr.cell(row=50, column=6).value = eod.cell(row=1, column=5).value
    sumr.cell(row=50, column=7).value = eod.cell(row=1, column=6).value
    sumr.cell(row=50, column=8).value = eod.cell(row=1, column=7).value
    sumr.cell(row=50, column=9).value = eod.cell(row=1, column=8).value
    sumr.cell(row=50, column=10).value = eod.cell(row=1, column=9).value
    sumr.cell(row=50, column=11).value = eod.cell(row=1, column=10).value
    sumr.cell(row=50, column=12).value = eod.cell(row=1, column=11).value
    sumr.cell(row=50, column=13).value = eod.cell(row=1, column=12).value
    sumr.cell(row=50, column=14).value = eod.cell(row=1, column=13).value

    # Creating months columns-------->
    sumr.cell(row=62, column=3).value = eod.cell(row=1, column=2).value
    sumr.cell(row=62, column=4).value = eod.cell(row=1, column=3).value
    sumr.cell(row=62, column=5).value = eod.cell(row=1, column=4).value
    sumr.cell(row=62, column=6).value = eod.cell(row=1, column=5).value
    sumr.cell(row=62, column=7).value = eod.cell(row=1, column=6).value
    sumr.cell(row=62, column=8).value = eod.cell(row=1, column=7).value
    sumr.cell(row=62, column=9).value = eod.cell(row=1, column=8).value
    sumr.cell(row=62, column=10).value = eod.cell(row=1, column=9).value
    sumr.cell(row=62, column=11).value = eod.cell(row=1, column=10).value
    sumr.cell(row=62, column=12).value = eod.cell(row=1, column=11).value
    sumr.cell(row=62, column=13).value = eod.cell(row=1, column=12).value
    sumr.cell(row=62, column=14).value = eod.cell(row=1, column=13).value


    sumr.cell(row=86, column=3).value = eod.cell(row=1, column=2).value
    sumr.cell(row=86, column=4).value = eod.cell(row=1, column=3).value
    sumr.cell(row=86, column=5).value = eod.cell(row=1, column=4).value
    sumr.cell(row=86, column=6).value = eod.cell(row=1, column=5).value
    sumr.cell(row=86, column=7).value = eod.cell(row=1, column=6).value
    sumr.cell(row=86, column=8).value = eod.cell(row=1, column=7).value
    sumr.cell(row=86, column=9).value = eod.cell(row=1, column=8).value
    sumr.cell(row=86, column=10).value = eod.cell(row=1, column=9).value
    sumr.cell(row=86, column=11).value = eod.cell(row=1, column=10).value
    sumr.cell(row=86, column=12).value = eod.cell(row=1, column=11).value
    sumr.cell(row=86, column=13).value = eod.cell(row=1, column=12).value
    sumr.cell(row=86, column=14).value = eod.cell(row=1, column=13).value


    sumr.cell(row=91, column=3).value = eod.cell(row=1, column=2).value
    sumr.cell(row=91, column=4).value = eod.cell(row=1, column=3).value
    sumr.cell(row=91, column=5).value = eod.cell(row=1, column=4).value
    sumr.cell(row=91, column=6).value = eod.cell(row=1, column=5).value
    sumr.cell(row=91, column=7).value = eod.cell(row=1, column=6).value
    sumr.cell(row=91, column=8).value = eod.cell(row=1, column=7).value
    sumr.cell(row=91, column=9).value = eod.cell(row=1, column=8).value
    sumr.cell(row=91, column=10).value = eod.cell(row=1, column=9).value
    sumr.cell(row=91, column=11).value = eod.cell(row=1, column=10).value
    sumr.cell(row=91, column=12).value = eod.cell(row=1, column=11).value
    sumr.cell(row=91, column=13).value = eod.cell(row=1, column=12).value
    sumr.cell(row=91, column=14).value = eod.cell(row=1, column=13).value

    #   Sum_of_credits
    try:
        sumr.cell(row=14, column=3, value=Sum_of_credits[sumr.cell(row=13, column=3).value])
    except Exception as e:
        sumr.cell(row=14, column=3, value=0)
    try:
        sumr.cell(row=14, column=4, value=Sum_of_credits[sumr.cell(row=13, column=4).value])
    except Exception as e:
        sumr.cell(row=14, column=4, value=0)
    try:
        sumr.cell(row=14, column=5, value=Sum_of_credits[sumr.cell(row=13, column=5).value])
    except Exception as e:
        sumr.cell(row=14, column=5, value=0)
    try:
        sumr.cell(row=14, column=6, value=Sum_of_credits[sumr.cell(row=13, column=6).value])
    except Exception as e:
        sumr.cell(row=14, column=6, value=0)
    try:
        sumr.cell(row=14, column=7, value=Sum_of_credits[sumr.cell(row=13, column=7).value])
    except Exception as e:
        sumr.cell(row=14, column=7, value=0)
    try:
        sumr.cell(row=14, column=8, value=Sum_of_credits[sumr.cell(row=13, column=8).value])
    except Exception as e:
        sumr.cell(row=14, column=8, value=0)
    try:
        sumr.cell(row=14, column=9, value=Sum_of_credits[sumr.cell(row=13, column=9).value])
    except Exception as e:
        sumr.cell(row=14, column=9, value=0)
    try:
        sumr.cell(row=14, column=10, value=Sum_of_credits[sumr.cell(row=13, column=10).value])
    except Exception as e:
        sumr.cell(row=14, column=10, value=0)
    try:
        sumr.cell(row=14, column=11, value=Sum_of_credits[sumr.cell(row=13, column=11).value])
    except Exception as e:
        sumr.cell(row=14, column=11, value=0)
    try:
        sumr.cell(row=14, column=12, value=Sum_of_credits[sumr.cell(row=13, column=12).value])
    except Exception as e:
        sumr.cell(row=14, column=12, value=0)
    try:
        sumr.cell(row=14, column=13, value=Sum_of_credits[sumr.cell(row=13, column=13).value])
    except Exception as e:
        sumr.cell(row=14, column=13, value=0)
    try:
        sumr.cell(row=14, column=14, value=Sum_of_credits[sumr.cell(row=13, column=14).value])
    except Exception as e:
        sumr.cell(row=14, column=14, value=0)

    #   Sum_of_credits
    try:
        sumr.cell(row=26, column=3, value=Sum_of_credits[sumr.cell(row=13, column=3).value])
    except Exception as e:
        sumr.cell(row=26, column=3, value=0)
    try:
        sumr.cell(row=26, column=4, value=Sum_of_credits[sumr.cell(row=13, column=4).value])
    except Exception as e:
        sumr.cell(row=26, column=4, value=0)
    try:
        sumr.cell(row=26, column=5, value=Sum_of_credits[sumr.cell(row=13, column=5).value])
    except Exception as e:
        sumr.cell(row=26, column=5, value=0)
    try:
        sumr.cell(row=26, column=6, value=Sum_of_credits[sumr.cell(row=13, column=6).value])
    except Exception as e:
        sumr.cell(row=26, column=6, value=0)
    try:
        sumr.cell(row=26, column=7, value=Sum_of_credits[sumr.cell(row=13, column=7).value])
    except Exception as e:
        sumr.cell(row=26, column=7, value=0)
    try:
        sumr.cell(row=26, column=8, value=Sum_of_credits[sumr.cell(row=13, column=8).value])
    except Exception as e:
        sumr.cell(row=26, column=8, value=0)
    try:
        sumr.cell(row=26, column=9, value=Sum_of_credits[sumr.cell(row=13, column=9).value])
    except Exception as e:
        sumr.cell(row=26, column=9, value=0)
    try:
        sumr.cell(row=26, column=10, value=Sum_of_credits[sumr.cell(row=13, column=10).value])
    except Exception as e:
        sumr.cell(row=26, column=10, value=0)
    try:
        sumr.cell(row=26, column=11, value=Sum_of_credits[sumr.cell(row=13, column=11).value])
    except Exception as e:
        sumr.cell(row=26, column=11, value=0)
    try:
        sumr.cell(row=26, column=12, value=Sum_of_credits[sumr.cell(row=13, column=12).value])
    except Exception as e:
        sumr.cell(row=26, column=12, value=0)
    try:
        sumr.cell(row=26, column=13, value=Sum_of_credits[sumr.cell(row=13, column=13).value])
    except Exception as e:
        sumr.cell(row=26, column=13, value=0)
    try:
        sumr.cell(row=26, column=14, value=Sum_of_credits[sumr.cell(row=13, column=14).value])
    except Exception as e:
        sumr.cell(row=26, column=14, value=0)

    #   Sum_of_debits
    try:
        sumr.cell(row=15, column=3, value=Sum_of_debits[sumr.cell(row=13, column=3).value])
    except Exception as e:
        sumr.cell(row=15, column=3, value=0)
    try:
        sumr.cell(row=15, column=4, value=Sum_of_debits[sumr.cell(row=13, column=4).value])
    except Exception as e:
        sumr.cell(row=15, column=4, value=0)
    try:
        sumr.cell(row=15, column=5, value=Sum_of_debits[sumr.cell(row=13, column=5).value])
    except Exception as e:
        sumr.cell(row=15, column=5, value=0)
    try:
        sumr.cell(row=15, column=6, value=Sum_of_debits[sumr.cell(row=13, column=6).value])
    except Exception as e:
        sumr.cell(row=15, column=6, value=0)
    try:
        sumr.cell(row=15, column=7, value=Sum_of_debits[sumr.cell(row=13, column=7).value])
    except Exception as e:
        sumr.cell(row=15, column=7, value=0)
    try:
        sumr.cell(row=15, column=8, value=Sum_of_debits[sumr.cell(row=13, column=8).value])
    except Exception as e:
        sumr.cell(row=15, column=8, value=0)
    try:
        sumr.cell(row=15, column=9, value=Sum_of_debits[sumr.cell(row=13, column=9).value])
    except Exception as e:
        sumr.cell(row=15, column=9, value=0)
    try:
        sumr.cell(row=15, column=10, value=Sum_of_debits[sumr.cell(row=13, column=10).value])
    except Exception as e:
        sumr.cell(row=15, column=10, value=0)
    try:
        sumr.cell(row=15, column=11, value=Sum_of_debits[sumr.cell(row=13, column=11).value])
    except Exception as e:
        sumr.cell(row=15, column=11, value=0)
    try:
        sumr.cell(row=15, column=12, value=Sum_of_debits[sumr.cell(row=13, column=12).value])
    except Exception as e:
        sumr.cell(row=15, column=12, value=0)
    try:
        sumr.cell(row=15, column=13, value=Sum_of_debits[sumr.cell(row=13, column=13).value])
    except Exception as e:
        sumr.cell(row=15, column=13, value=0)
    try:
        sumr.cell(row=15, column=14, value=Sum_of_debits[sumr.cell(row=13, column=14).value])
    except Exception as e:
        sumr.cell(row=15, column=14, value=0)

        #   Sum_of_debits
    try:
        sumr.cell(row=63, column=3, value=Sum_of_debits[sumr.cell(row=13, column=3).value])
    except Exception as e:
        sumr.cell(row=63, column=3, value=0)
    try:
        sumr.cell(row=63, column=4, value=Sum_of_debits[sumr.cell(row=13, column=4).value])
    except Exception as e:
        sumr.cell(row=63, column=4, value=0)
    try:
        sumr.cell(row=63, column=5, value=Sum_of_debits[sumr.cell(row=13, column=5).value])
    except Exception as e:
        sumr.cell(row=63, column=5, value=0)
    try:
        sumr.cell(row=63, column=6, value=Sum_of_debits[sumr.cell(row=13, column=6).value])
    except Exception as e:
        sumr.cell(row=63, column=6, value=0)
    try:
        sumr.cell(row=63, column=7, value=Sum_of_debits[sumr.cell(row=13, column=7).value])
    except Exception as e:
        sumr.cell(row=63, column=7, value=0)
    try:
        sumr.cell(row=63, column=8, value=Sum_of_debits[sumr.cell(row=13, column=8).value])
    except Exception as e:
        sumr.cell(row=63, column=8, value=0)
    try:
        sumr.cell(row=63, column=9, value=Sum_of_debits[sumr.cell(row=13, column=9).value])
    except Exception as e:
        sumr.cell(row=63, column=9, value=0)
    try:
        sumr.cell(row=63, column=10, value=Sum_of_debits[sumr.cell(row=13, column=10).value])
    except Exception as e:
        sumr.cell(row=63, column=10, value=0)
    try:
        sumr.cell(row=63, column=11, value=Sum_of_debits[sumr.cell(row=13, column=11).value])
    except Exception as e:
        sumr.cell(row=63, column=11, value=0)
    try:
        sumr.cell(row=63, column=12, value=Sum_of_debits[sumr.cell(row=13, column=12).value])
    except Exception as e:
        sumr.cell(row=63, column=12, value=0)
    try:
        sumr.cell(row=63, column=13, value=Sum_of_debits[sumr.cell(row=13, column=13).value])
    except Exception as e:
        sumr.cell(row=63, column=13, value=0)
    try:
        sumr.cell(row=63, column=14, value=Sum_of_debits[sumr.cell(row=13, column=14).value])
    except Exception as e:
        sumr.cell(row=63, column=14, value=0)

    try:
        sumr.cell(row=37, column=3, value=Sum_of_debits[sumr.cell(row=13, column=3).value])
    except Exception as e:
        sumr.cell(row=37, column=3, value=0)
    try:
        sumr.cell(row=37, column=4, value=Sum_of_debits[sumr.cell(row=13, column=4).value])
    except Exception as e:
        sumr.cell(row=37, column=4, value=0)
    try:
        sumr.cell(row=37, column=5, value=Sum_of_debits[sumr.cell(row=13, column=5).value])
    except Exception as e:
        sumr.cell(row=37, column=5, value=0)
    try:
        sumr.cell(row=37, column=6, value=Sum_of_debits[sumr.cell(row=13, column=6).value])
    except Exception as e:
        sumr.cell(row=37, column=6, value=0)
    try:
        sumr.cell(row=37, column=7, value=Sum_of_debits[sumr.cell(row=13, column=7).value])
    except Exception as e:
        sumr.cell(row=37, column=7, value=0)
    try:
        sumr.cell(row=37, column=8, value=Sum_of_debits[sumr.cell(row=13, column=8).value])
    except Exception as e:
        sumr.cell(row=37, column=8, value=0)
    try:
        sumr.cell(row=37, column=9, value=Sum_of_debits[sumr.cell(row=13, column=9).value])
    except Exception as e:
        sumr.cell(row=37, column=9, value=0)
    try:
        sumr.cell(row=37, column=10, value=Sum_of_debits[sumr.cell(row=13, column=10).value])
    except Exception as e:
        sumr.cell(row=37, column=10, value=0)
    try:
        sumr.cell(row=37, column=11, value=Sum_of_debits[sumr.cell(row=13, column=11).value])
    except Exception as e:
        sumr.cell(row=37, column=11, value=0)
    try:
        sumr.cell(row=37, column=12, value=Sum_of_debits[sumr.cell(row=13, column=12).value])
    except Exception as e:
        sumr.cell(row=37, column=12, value=0)
    try:
        sumr.cell(row=37, column=13, value=Sum_of_debits[sumr.cell(row=13, column=13).value])
    except Exception as e:
        sumr.cell(row=37, column=13, value=0)
    try:
        sumr.cell(row=37, column=14, value=Sum_of_debits[sumr.cell(row=13, column=14).value])
    except Exception as e:
        sumr.cell(row=37, column=14, value=0)


    # Total_no_withdrawals
    try:
        sumr.cell(row=16, column=3, value=Total_no_withdrawals[sumr.cell(row=13, column=3).value])
    except Exception as e:
        sumr.cell(row=16, column=3, value=0)
    try:
        sumr.cell(row=16, column=4, value=Total_no_withdrawals[sumr.cell(row=13, column=4).value])
    except Exception as e:
        sumr.cell(row=16, column=4, value=0)
    try:
        sumr.cell(row=16, column=5, value=Total_no_withdrawals[sumr.cell(row=13, column=5).value])
    except Exception as e:
        sumr.cell(row=16, column=5, value=0)
    try:
        sumr.cell(row=16, column=6, value=Total_no_withdrawals[sumr.cell(row=13, column=6).value])
    except Exception as e:
        sumr.cell(row=16, column=6, value=0)
    try:
        sumr.cell(row=16, column=7, value=Total_no_withdrawals[sumr.cell(row=13, column=7).value])
    except Exception as e:
        sumr.cell(row=16, column=7, value=0)
    try:
        sumr.cell(row=16, column=8, value=Total_no_withdrawals[sumr.cell(row=13, column=8).value])
    except Exception as e:
        sumr.cell(row=16, column=8, value=0)
    try:
        sumr.cell(row=16, column=9, value=Total_no_withdrawals[sumr.cell(row=13, column=9).value])
    except Exception as e:
        sumr.cell(row=16, column=9, value=0)
    try:
        sumr.cell(row=16, column=10, value=Total_no_withdrawals[sumr.cell(row=13, column=10).value])
    except Exception as e:
        sumr.cell(row=16, column=10, value=0)
    try:
        sumr.cell(row=16, column=11, value=Total_no_withdrawals[sumr.cell(row=13, column=11).value])
    except Exception as e:
        sumr.cell(row=16, column=11, value=0)
    try:
        sumr.cell(row=16, column=12, value=Total_no_withdrawals[sumr.cell(row=13, column=12).value])
    except Exception as e:
        sumr.cell(row=16, column=12, value=0)
    try:
        sumr.cell(row=16, column=13, value=Total_no_withdrawals[sumr.cell(row=13, column=13).value])
    except Exception as e:
        sumr.cell(row=16, column=13, value=0)
    try:
        sumr.cell(row=16, column=14, value=Total_no_withdrawals[sumr.cell(row=13, column=14).value])
    except Exception as e:
        sumr.cell(row=16, column=14, value=0)

    # Total_no_cash_depo
    try:
        sumr.cell(row=17, column=3, value=Total_no_cash_depo[sumr.cell(row=13, column=3).value])
    except Exception as e:
        sumr.cell(row=17, column=3, value=0)
    try:
        sumr.cell(row=17, column=4, value=Total_no_cash_depo[sumr.cell(row=13, column=4).value])
    except Exception as e:
        sumr.cell(row=17, column=4, value=0)
    try:
        sumr.cell(row=17, column=5, value=Total_no_cash_depo[sumr.cell(row=13, column=5).value])
    except Exception as e:
        sumr.cell(row=17, column=5, value=0)
    try:
        sumr.cell(row=17, column=6, value=Total_no_cash_depo[sumr.cell(row=13, column=6).value])
    except Exception as e:
        sumr.cell(row=17, column=6, value=0)
    try:
        sumr.cell(row=17, column=7, value=Total_no_cash_depo[sumr.cell(row=13, column=7).value])
    except Exception as e:
        sumr.cell(row=17, column=7, value=0)
    try:
        sumr.cell(row=17, column=8, value=Total_no_cash_depo[sumr.cell(row=13, column=8).value])
    except Exception as e:
        sumr.cell(row=17, column=8, value=0)
    try:
        sumr.cell(row=17, column=9, value=Total_no_cash_depo[sumr.cell(row=13, column=9).value])
    except Exception as e:
        sumr.cell(row=17, column=9, value=0)
    try:
        sumr.cell(row=17, column=10, value=Total_no_cash_depo[sumr.cell(row=13, column=10).value])
    except Exception as e:
        sumr.cell(row=17, column=10, value=0)
    try:
        sumr.cell(row=17, column=11, value=Total_no_cash_depo[sumr.cell(row=13, column=11).value])
    except Exception as e:
        sumr.cell(row=17, column=11, value=0)
    try:
        sumr.cell(row=17, column=12, value=Total_no_cash_depo[sumr.cell(row=13, column=12).value])
    except Exception as e:
        sumr.cell(row=17, column=12, value=0)
    try:
        sumr.cell(row=17, column=13, value=Total_no_cash_depo[sumr.cell(row=13, column=13).value])
    except Exception as e:
        sumr.cell(row=17, column=13, value=0)
    try:
        sumr.cell(row=17, column=14, value=Total_no_cash_depo[sumr.cell(row=13, column=14).value])
    except Exception as e:
        sumr.cell(row=17, column=14, value=0)

    # POS_cr
    try:
        sumr.cell(row=18, column=3, value=POS_cr[sumr.cell(row=13, column=3).value])
    except Exception as e:
        sumr.cell(row=18, column=3, value=0)
    try:
        sumr.cell(row=18, column=4, value=POS_cr[sumr.cell(row=13, column=4).value])
    except Exception as e:
        sumr.cell(row=18, column=4, value=0)
    try:
        sumr.cell(row=18, column=5, value=POS_cr[sumr.cell(row=13, column=5).value])
    except Exception as e:
        sumr.cell(row=18, column=5, value=0)
    try:
        sumr.cell(row=18, column=6, value=POS_cr[sumr.cell(row=13, column=6).value])
    except Exception as e:
        sumr.cell(row=18, column=6, value=0)
    try:
        sumr.cell(row=18, column=7, value=POS_cr[sumr.cell(row=13, column=7).value])
    except Exception as e:
        sumr.cell(row=18, column=7, value=0)
    try:
        sumr.cell(row=18, column=8, value=POS_cr[sumr.cell(row=13, column=8).value])
    except Exception as e:
        sumr.cell(row=18, column=8, value=0)
    try:
        sumr.cell(row=18, column=9, value=POS_cr[sumr.cell(row=13, column=9).value])
    except Exception as e:
        sumr.cell(row=18, column=9, value=0)
    try:
        sumr.cell(row=18, column=10, value=POS_cr[sumr.cell(row=13, column=10).value])
    except Exception as e:
        sumr.cell(row=18, column=10, value=0)
    try:
        sumr.cell(row=18, column=11, value=POS_cr[sumr.cell(row=13, column=11).value])
    except Exception as e:
        sumr.cell(row=18, column=11, value=0)
    try:
        sumr.cell(row=18, column=12, value=POS_cr[sumr.cell(row=13, column=12).value])
    except Exception as e:
        sumr.cell(row=18, column=12, value=0)
    try:
        sumr.cell(row=18, column=13, value=POS_cr[sumr.cell(row=13, column=13).value])
    except Exception as e:
        sumr.cell(row=18, column=13, value=0)
    try:
        sumr.cell(row=18, column=14, value=POS_cr[sumr.cell(row=13, column=14).value])
    except Exception as e:
        sumr.cell(row=18, column=14, value=0)

    # POS_dr
    try:
        sumr.cell(row=20, column=3, value=POS_dr[sumr.cell(row=13, column=3).value])
    except Exception as e:
        sumr.cell(row=20, column=3, value=0)
    try:
        sumr.cell(row=20, column=4, value=POS_dr[sumr.cell(row=13, column=4).value])
    except Exception as e:
        sumr.cell(row=20, column=4, value=0)
    try:
        sumr.cell(row=20, column=5, value=POS_dr[sumr.cell(row=13, column=5).value])
    except Exception as e:
        sumr.cell(row=20, column=5, value=0)
    try:
        sumr.cell(row=20, column=6, value=POS_dr[sumr.cell(row=13, column=6).value])
    except Exception as e:
        sumr.cell(row=20, column=6, value=0)
    try:
        sumr.cell(row=20, column=7, value=POS_dr[sumr.cell(row=13, column=7).value])
    except Exception as e:
        sumr.cell(row=20, column=7, value=0)
    try:
        sumr.cell(row=20, column=8, value=POS_dr[sumr.cell(row=13, column=8).value])
    except Exception as e:
        sumr.cell(row=20, column=8, value=0)
    try:
        sumr.cell(row=20, column=9, value=POS_dr[sumr.cell(row=13, column=9).value])
    except Exception as e:
        sumr.cell(row=20, column=9, value=0)
    try:
        sumr.cell(row=20, column=10, value=POS_dr[sumr.cell(row=13, column=10).value])
    except Exception as e:
        sumr.cell(row=20, column=10, value=0)
    try:
        sumr.cell(row=20, column=11, value=POS_dr[sumr.cell(row=13, column=11).value])
    except Exception as e:
        sumr.cell(row=20, column=11, value=0)
    try:
        sumr.cell(row=20, column=12, value=POS_dr[sumr.cell(row=13, column=12).value])
    except Exception as e:
        sumr.cell(row=20, column=12, value=0)
    try:
        sumr.cell(row=20, column=13, value=POS_dr[sumr.cell(row=13, column=13).value])
    except Exception as e:
        sumr.cell(row=20, column=13, value=0)
    try:
        sumr.cell(row=20, column=14, value=POS_dr[sumr.cell(row=13, column=14).value])
    except Exception as e:
        sumr.cell(row=20, column=14, value=0)

    # POS_dr
    try:
        sumr.cell(row=19, column=3, value=Inv_dr[sumr.cell(row=13, column=3).value])
    except Exception as e:
        sumr.cell(row=19, column=3, value=0)
    try:
        sumr.cell(row=19, column=4, value=Inv_dr[sumr.cell(row=13, column=4).value])
    except Exception as e:
        sumr.cell(row=19, column=4, value=0)
    try:
        sumr.cell(row=19, column=5, value=Inv_dr[sumr.cell(row=13, column=5).value])
    except Exception as e:
        sumr.cell(row=19, column=5, value=0)
    try:
        sumr.cell(row=19, column=6, value=Inv_dr[sumr.cell(row=13, column=6).value])
    except Exception as e:
        sumr.cell(row=19, column=6, value=0)
    try:
        sumr.cell(row=19, column=7, value=Inv_dr[sumr.cell(row=13, column=7).value])
    except Exception as e:
        sumr.cell(row=19, column=7, value=0)
    try:
        sumr.cell(row=19, column=8, value=Inv_dr[sumr.cell(row=13, column=8).value])
    except Exception as e:
        sumr.cell(row=19, column=8, value=0)
    try:
        sumr.cell(row=19, column=9, value=Inv_dr[sumr.cell(row=13, column=9).value])
    except Exception as e:
        sumr.cell(row=19, column=9, value=0)
    try:
        sumr.cell(row=19, column=10, value=Inv_dr[sumr.cell(row=13, column=10).value])
    except Exception as e:
        sumr.cell(row=19, column=10, value=0)
    try:
        sumr.cell(row=19, column=11, value=Inv_dr[sumr.cell(row=13, column=11).value])
    except Exception as e:
        sumr.cell(row=19, column=11, value=0)
    try:
        sumr.cell(row=19, column=12, value=Inv_dr[sumr.cell(row=13, column=12).value])
    except Exception as e:
        sumr.cell(row=19, column=12, value=0)
    try:
        sumr.cell(row=19, column=13, value=Inv_dr[sumr.cell(row=13, column=13).value])
    except Exception as e:
        sumr.cell(row=19, column=13, value=0)
    try:
        sumr.cell(row=19, column=14, value=Inv_dr[sumr.cell(row=13, column=14).value])
    except Exception as e:
        sumr.cell(row=19, column=14, value=0)

    wb.save('Excel_Files/Dashboard/BankStatement.xlsx')

general()

def Balance():

# Opening_Bal
    wb = load_workbook('Excel_Files/Dashboard/BankStatement.xlsx')
    sumr= wb['summary']
    eod= wb['EOD Balance']
    df5 = pd.read_excel('Excel_Files/Dashboard/BankStatement.xlsx',sheet_name='EOD Balance',index_col = 'Day')
   
    

    
   
    sumr.cell(row=21,column=3).value = eod.cell(row=2,column=2).value
    sumr.cell(row=21,column=4).value= eod.cell(row=2,column=3).value
    sumr.cell(row=21,column=5).value= eod.cell(row=2,column=4).value
    sumr.cell(row=21,column=6).value= eod.cell(row=2,column=5).value
    sumr.cell(row=21,column=7).value= eod.cell(row=2,column=6).value
    sumr.cell(row=21,column=8).value= eod.cell(row=2,column=7).value
    sumr.cell(row=21,column=9).value= eod.cell(row=2,column=8).value
    sumr.cell(row=21,column=10).value= eod.cell(row=2,column=9).value
    sumr.cell(row=21,column=11).value= eod.cell(row=2,column=10).value
    sumr.cell(row=21,column=12).value= eod.cell(row=2,column=11).value
    sumr.cell(row=21,column=13).value= eod.cell(row=2,column=12).value
    sumr.cell(row=21,column=14).value= eod.cell(row=2,column=13).value

    wb.save('Excel_Files/Dashboard/BankStatement.xlsx')

Balance()

def Income():
    #print(df1)
    wb = load_workbook('Excel_Files/Dashboard/BankStatement.xlsx')
    sumr= wb['summary']
    eod= wb['EOD Balance']
    
#     Intrest Credit
    df_grouped =df1[df1["Particulars"].str.contains("Int.Pd")]
    Interest=df_grouped.groupby(df_grouped['Tran Date'].dt.strftime('%b-%Y'))['Credit'].sum().sort_values()
    df_grouped['Categories'] = 'Intrest Credit'
    df1.update(df_grouped)
    # print(Interest)

    #     Upi credits
    upi_cr = df1[df1["Particulars"].str.contains("UPI")].groupby('Credit')
    upi_cr = upi_cr.apply(lambda x: x)
    upi_cr['Categories'] = 'UPI-cr'
    upi_dr = df1[df1["Particulars"].str.contains("UPI")].groupby('Debit')
    upi_dr = upi_dr.apply(lambda x: x)
    upi_dr['Categories'] = 'UPI-dr'
    df1.update(upi_cr)
    df1.update(upi_dr)

    #     Nach receaved
    Nach = df1[df1['Particulars'].str.contains('ECS') | df1['Particulars'].str.contains('NACH-CR')]
    Nach_cr = Nach.groupby(Nach['Tran Date'].dt.strftime('%b-%Y'))['Credit'].sum().sort_values()
    Nach['Categories'] = 'NACH-CR'
    df1.update(Nach)

    IT_Nach = df1[df1['Particulars'].str.contains('ACH-CR')&df1['Particulars'].str.contains('AY2018-19-NACH',na=False) | df1['Particulars'].str.contains('ACH-CR')&df1['Particulars'].str.contains('AY2019-20-NACH',na=False)| df1['Particulars'].str.contains('ACH-CR')&df1['Particulars'].str.contains('AY2020-21-NACH',na=False)| df1['Particulars'].str.contains('ACH-CR')&df1['Particulars'].str.contains('AY2021-22-NACH',na=False)| df1['Particulars'].str.contains('ACH-CR')&df1['Particulars'].str.contains('AY2022-23-NACH',na=False)| df1['Particulars'].str.contains('ACH-CR')&df1['Particulars'].str.contains('AY2023-24-NACH',na=False)| df1['Particulars'].str.contains('ACH-CR')&df1['Particulars'].str.contains('AY2024-25-NACH',na=False)]
    IT_Nach_rec=IT_Nach.groupby(IT_Nach['Tran Date'].dt.strftime('%b-%Y'))['Credit'].sum().sort_values()
    IT_Nach['Categories']='Income Tax Refund'
    df1.update(IT_Nach)

    #     rent received
    rent_df = df1[df1['Particulars'].str.contains('RENT')]
    rent_cr = rent_df.groupby(rent_df['Tran Date'].dt.strftime('%b-%Y'))['Credit'].sum().sort_values()
    #print(rent_cr)
    rent_df['Categories']='RENT RECEIVED'
    df1.update(rent_df)
    
    #   Salary Received
    df_salary = df1[df1["Particulars"].str.contains("SALARY",na=False)]
    salary=df_salary.groupby(df_salary['Tran Date'].dt.strftime('%b-%Y'))['Credit'].sum().sort_values()
    df_salary['Categories']='SALARY RECEIVED'
    df1.update(df_salary)
  
  #   Dividend
    dividend_df = df1[df1['Particulars'].str.contains('dividend')]
    dividend = dividend_df.groupby(dividend_df['Tran Date'].dt.strftime('%b-%Y'))['Credit'].sum().sort_values()
    dividend_df['Categories']='dividend'
    df1.update(dividend_df)
    
    #Loan 
    loan_df =df1[df1['Particulars'].str.contains('loan')]
    loan = loan_df.groupby(loan_df['Tran Date'].dt.strftime('%b-%Y'))['Credit'].sum().sort_values()
    loan['Categories']='loan'
    df1.update(loan)

    try:
       sumr.cell(row=27,column=3,value=Interest[sumr.cell(row=13,column=3).value])
    except Exception as e:
      sumr.cell(row=27,column=3,value=0)
    try:
       sumr.cell(row=27,column=4,value=Interest[sumr.cell(row=13,column=4).value])
    except Exception as e:
      sumr.cell(row=27,column=4,value=0)
    try:
       sumr.cell(row=27,column=5,value=Interest[sumr.cell(row=13,column=5).value])
    except Exception as e:
      sumr.cell(row=27,column=5,value=0)
    try:
       sumr.cell(row=27,column=6,value=Interest[sumr.cell(row=13,column=6).value])
    except Exception as e:
      sumr.cell(row=27,column=6,value=0)
    try:
       sumr.cell(row=27,column=7,value=Interest[sumr.cell(row=13,column=7).value])
    except Exception as e:
      sumr.cell(row=27,column=7,value=0)
    try:
       sumr.cell(row=27,column=8,value=Interest[sumr.cell(row=13,column=8).value])
    except Exception as e:
      sumr.cell(row=27,column=8,value=0)
    try:
       sumr.cell(row=27,column=9,value=Interest[sumr.cell(row=13,column=9).value])
    except Exception as e:
      sumr.cell(row=27,column=9,value=0)
    try:
       sumr.cell(row=27,column=10,value=Interest[sumr.cell(row=13,column=10).value])
    except Exception as e:
      sumr.cell(row=27,column=10,value=0)
    try:
       sumr.cell(row=27,column=11,value=Interest[sumr.cell(row=13,column=11).value])
    except Exception as e:
      sumr.cell(row=27,column=11,value=0)
    try:
       sumr.cell(row=27,column=12,value=Interest[sumr.cell(row=13,column=12).value])
    except Exception as e:
      sumr.cell(row=27,column=12,value=0)
    try:
       sumr.cell(row=27,column=13,value=Interest[sumr.cell(row=13,column=13).value])
    except Exception as e:
      sumr.cell(row=27,column=13,value=0)
    try:
       sumr.cell(row=27,column=14,value=Interest[sumr.cell(row=13,column=14).value])
    except Exception as e:
      sumr.cell(row=27,column=14,value=0)


    try:
       sumr.cell(row=28,column=3,value=salary[sumr.cell(row=13,column=3).value])
    except Exception as e:
      sumr.cell(row=28,column=3,value=0)
    try:
       sumr.cell(row=28,column=4,value=salary[sumr.cell(row=13,column=4).value])
    except Exception as e:
      sumr.cell(row=28,column=4,value=0)
    try:
       sumr.cell(row=28,column=5,value=salary[sumr.cell(row=13,column=5).value])
    except Exception as e:
      sumr.cell(row=28,column=5,value=0)
    try:
       sumr.cell(row=28,column=6,value=salary[sumr.cell(row=13,column=6).value])
    except Exception as e:
      sumr.cell(row=28,column=6,value=0)
    try:
       sumr.cell(row=28,column=7,value=salary[sumr.cell(row=13,column=7).value])
    except Exception as e:
      sumr.cell(row=28,column=7,value=0)
    try:
       sumr.cell(row=28,column=8,value=salary[sumr.cell(row=13,column=8).value])
    except Exception as e:
      sumr.cell(row=28,column=8,value=0)
    try:
       sumr.cell(row=28,column=9,value=salary[sumr.cell(row=13,column=9).value])
    except Exception as e:
      sumr.cell(row=28,column=9,value=0)
    try:
       sumr.cell(row=28,column=10,value=salary[sumr.cell(row=13,column=10).value])
    except Exception as e:
      sumr.cell(row=28,column=10,value=0)
    try:
       sumr.cell(row=28,column=11,value=salary[sumr.cell(row=13,column=11).value])
    except Exception as e:
      sumr.cell(row=28,column=11,value=0)
    try:
       sumr.cell(row=28,column=12,value=salary[sumr.cell(row=13,column=12).value])
    except Exception as e:
      sumr.cell(row=28,column=12,value=0)
    try:
       sumr.cell(row=28,column=13,value=salary[sumr.cell(row=13,column=13).value])
    except Exception as e:
      sumr.cell(row=28,column=13,value=0)
    try:
       sumr.cell(row=28,column=14,value=salary[sumr.cell(row=13,column=14).value])
    except Exception as e:
      sumr.cell(row=28,column=14,value=0)


    
    try:
       sumr.cell(row=29,column=3,value=Nach_cr[sumr.cell(row=13,column=3).value])
    except Exception as e:
      sumr.cell(row=29,column=3,value=0)
    try:
       sumr.cell(row=29,column=4,value=Nach_cr[sumr.cell(row=13,column=4).value])
    except Exception as e:
      sumr.cell(row=29,column=4,value=0)
    try:
       sumr.cell(row=29,column=5,value=Nach_cr[sumr.cell(row=13,column=5).value])
    except Exception as e:
      sumr.cell(row=29,column=5,value=0)
    try:
       sumr.cell(row=29,column=6,value=Nach_cr[sumr.cell(row=13,column=6).value])
    except Exception as e:
      sumr.cell(row=29,column=6,value=0)
    try:
       sumr.cell(row=29,column=7,value=Nach_cr[sumr.cell(row=13,column=7).value])
    except Exception as e:
      sumr.cell(row=29,column=7,value=0)
    try:
       sumr.cell(row=29,column=8,value=Nach_cr[sumr.cell(row=13,column=8).value])
    except Exception as e:
      sumr.cell(row=29,column=8,value=0)
    try:
       sumr.cell(row=29,column=9,value=Nach_cr[sumr.cell(row=13,column=9).value])
    except Exception as e:
      sumr.cell(row=29,column=9,value=0)
    try:
       sumr.cell(row=29,column=10,value=Nach_cr[sumr.cell(row=13,column=10).value])
    except Exception as e:
      sumr.cell(row=29,column=10,value=0)
    try:
       sumr.cell(row=29,column=11,value=Nach_cr[sumr.cell(row=13,column=11).value])
    except Exception as e:
      sumr.cell(row=29,column=11,value=0)
    try:
       sumr.cell(row=29,column=12,value=Nach_cr[sumr.cell(row=13,column=12).value])
    except Exception as e:
      sumr.cell(row=29,column=12,value=0)
    try:
       sumr.cell(row=29,column=13,value=Nach_cr[sumr.cell(row=13,column=13).value])
    except Exception as e:
      sumr.cell(row=29,column=13,value=0)
    try:
       sumr.cell(row=29,column=14,value=Nach_cr[sumr.cell(row=13,column=14).value])
    except Exception as e:
      sumr.cell(row=29,column=14,value=0)

    try:
        sumr.cell(row=30, column=3, value=loan[sumr.cell(row=13, column=3).value])
    except Exception as e:
        sumr.cell(row=30, column=3, value=0)
    try:
        sumr.cell(row=30, column=4, value=loan[sumr.cell(row=13, column=4).value])
    except Exception as e:
        sumr.cell(row=30, column=4, value=0)
    try:
        sumr.cell(row=30, column=5, value=loan[sumr.cell(row=13, column=5).value])
    except Exception as e:
        sumr.cell(row=30, column=5, value=0)
    try:
        sumr.cell(row=30, column=6, value=loan[sumr.cell(row=13, column=6).value])
    except Exception as e:
        sumr.cell(row=30, column=6, value=0)
    try:
        sumr.cell(row=30, column=7, value=loan[sumr.cell(row=13, column=7).value])
    except Exception as e:
        sumr.cell(row=30, column=7, value=0)
    try:
        sumr.cell(row=30, column=8, value=loan[sumr.cell(row=13, column=8).value])
    except Exception as e:
        sumr.cell(row=30, column=8, value=0)
    try:
        sumr.cell(row=30, column=9, value=loan[sumr.cell(row=13, column=9).value])
    except Exception as e:
        sumr.cell(row=30, column=9, value=0)
    try:
        sumr.cell(row=30, column=10, value=loan[sumr.cell(row=13, column=10).value])
    except Exception as e:
        sumr.cell(row=30, column=10, value=0)
    try:
        sumr.cell(row=30, column=11, value=loan[sumr.cell(row=13, column=11).value])
    except Exception as e:
        sumr.cell(row=30, column=11, value=0)
    try:
        sumr.cell(row=30, column=12, value=loan[sumr.cell(row=13, column=12).value])
    except Exception as e:
        sumr.cell(row=30, column=12, value=0)
    try:
        sumr.cell(row=30, column=13, value=loan[sumr.cell(row=13, column=13).value])
    except Exception as e:
        sumr.cell(row=30, column=13, value=0)
    try:
        sumr.cell(row=30, column=14, value=loan[sumr.cell(row=13, column=14).value])
    except Exception as e:
        sumr.cell(row=30, column=14, value=0)


    try:
       sumr.cell(row=31,column=3,value=IT_Nach_rec[sumr.cell(row=13,column=3).value])
    except Exception as e:
      sumr.cell(row=31,column=3,value=0)
    try:
       sumr.cell(row=31,column=4,value=IT_Nach_rec[sumr.cell(row=13,column=4).value])
    except Exception as e:
      sumr.cell(row=31,column=4,value=0)
    try:
       sumr.cell(row=31,column=5,value=IT_Nach_rec[sumr.cell(row=13,column=5).value])
    except Exception as e:
      sumr.cell(row=31,column=5,value=0)
    try:
       sumr.cell(row=31,column=6,value=IT_Nach_rec[sumr.cell(row=13,column=6).value])
    except Exception as e:
      sumr.cell(row=31,column=6,value=0)
    try:
       sumr.cell(row=31,column=7,value=IT_Nach_rec[sumr.cell(row=13,column=7).value])
    except Exception as e:
      sumr.cell(row=31,column=7,value=0)
    try:
       sumr.cell(row=31,column=8,value=IT_Nach_rec[sumr.cell(row=13,column=8).value])
    except Exception as e:
      sumr.cell(row=31,column=8,value=0)
    try:
       sumr.cell(row=31,column=9,value=IT_Nach_rec[sumr.cell(row=13,column=9).value])
    except Exception as e:
      sumr.cell(row=31,column=9,value=0)
    try:
       sumr.cell(row=31,column=10,value=IT_Nach_rec[sumr.cell(row=13,column=10).value])
    except Exception as e:
      sumr.cell(row=31,column=10,value=0)
    try:
       sumr.cell(row=31,column=11,value=IT_Nach_rec[sumr.cell(row=13,column=11).value])
    except Exception as e:
      sumr.cell(row=31,column=11,value=0)
    try:
       sumr.cell(row=31,column=12,value=IT_Nach_rec[sumr.cell(row=13,column=12).value])
    except Exception as e:
      sumr.cell(row=31,column=12,value=0)
    try:
       sumr.cell(row=31,column=13,value=IT_Nach_rec[sumr.cell(row=13,column=13).value])
    except Exception as e:
      sumr.cell(row=31,column=13,value=0)
    try:
       sumr.cell(row=31,column=14,value=IT_Nach_rec[sumr.cell(row=13,column=14).value])
    except Exception as e:
      sumr.cell(row=31,column=14,value=0)


    try:
       sumr.cell(row=32,column=3,value=dividend[sumr.cell(row=13,column=3).value])
    except Exception as e:
      sumr.cell(row=32,column=3,value=0)
    try:
       sumr.cell(row=32,column=4,value=dividend[sumr.cell(row=13,column=4).value])
    except Exception as e:
      sumr.cell(row=32,column=4,value=0)
    try:
       sumr.cell(row=32,column=5,value=dividend[sumr.cell(row=13,column=5).value])
    except Exception as e:
      sumr.cell(row=32,column=5,value=0)
    try:
       sumr.cell(row=32,column=6,value=dividend[sumr.cell(row=13,column=6).value])
    except Exception as e:
      sumr.cell(row=32,column=6,value=0)
    try:
       sumr.cell(row=32,column=7,value=dividend[sumr.cell(row=13,column=7).value])
    except Exception as e:
      sumr.cell(row=32,column=7,value=0)
    try:
       sumr.cell(row=32,column=8,value=dividend[sumr.cell(row=13,column=8).value])
    except Exception as e:
      sumr.cell(row=32,column=8,value=0)
    try:
       sumr.cell(row=32,column=9,value=dividend[sumr.cell(row=13,column=9).value])
    except Exception as e:
      sumr.cell(row=32,column=9,value=0)
    try:
       sumr.cell(row=32,column=10,value=dividend[sumr.cell(row=13,column=10).value])
    except Exception as e:
      sumr.cell(row=32,column=10,value=0)
    try:
       sumr.cell(row=32,column=11,value=dividend[sumr.cell(row=13,column=11).value])
    except Exception as e:
      sumr.cell(row=32,column=11,value=0)
    try:
       sumr.cell(row=32,column=12,value=dividend[sumr.cell(row=13,column=12).value])
    except Exception as e:
      sumr.cell(row=32,column=12,value=0)
    try:
       sumr.cell(row=32,column=13,value=dividend[sumr.cell(row=13,column=13).value])
    except Exception as e:
      sumr.cell(row=32,column=13,value=0)
    try:
       sumr.cell(row=32,column=14,value=dividend[sumr.cell(row=13,column=14).value])
    except Exception as e:
      sumr.cell(row=32,column=14,value=0)




    try:
       sumr.cell(row=33,column=3,value=rent_cr[sumr.cell(row=13,column=3).value])
    except Exception as e:
      sumr.cell(row=33,column=3,value=0)
    try:
       sumr.cell(row=33,column=4,value=rent_cr[sumr.cell(row=13,column=4).value])
    except Exception as e:
      sumr.cell(row=33,column=4,value=0)
    try:
       sumr.cell(row=33,column=5,value=rent_cr[sumr.cell(row=13,column=5).value])
    except Exception as e:
      sumr.cell(row=33,column=5,value=0)
    try:
       sumr.cell(row=33,column=6,value=rent_cr[sumr.cell(row=13,column=6).value])
    except Exception as e:
      sumr.cell(row=33,column=6,value=0)
    try:
       sumr.cell(row=33,column=7,value=rent_cr[sumr.cell(row=13,column=7).value])
    except Exception as e:
      sumr.cell(row=33,column=7,value=0)
    try:
       sumr.cell(row=33,column=8,value=rent_cr[sumr.cell(row=13,column=8).value])
    except Exception as e:
      sumr.cell(row=33,column=8,value=0)
    try:
       sumr.cell(row=33,column=9,value=rent_cr[sumr.cell(row=13,column=9).value])
    except Exception as e:
      sumr.cell(row=33,column=9,value=0)
    try:
       sumr.cell(row=33,column=10,value=rent_cr[sumr.cell(row=13,column=10).value])
    except Exception as e:
      sumr.cell(row=33,column=10,value=0)
    try:
       sumr.cell(row=33,column=11,value=rent_cr[sumr.cell(row=13,column=11).value])
    except Exception as e:
      sumr.cell(row=33,column=11,value=0)
    try:
       sumr.cell(row=33,column=12,value=rent_cr[sumr.cell(row=13,column=12).value])
    except Exception as e:
      sumr.cell(row=33,column=12,value=0)
    try:
       sumr.cell(row=32,column=13,value=rent_cr[sumr.cell(row=13,column=13).value])
    except Exception as e:
      sumr.cell(row=33,column=13,value=0)
    try:
       sumr.cell(row=33,column=14,value=rent_cr[sumr.cell(row=13,column=14).value])
    except Exception as e:
      sumr.cell(row=33,column=14,value=0)
      


# Save the changes

    wb.save('Excel_Files/Dashboard/BankStatement.xlsx')

Income()

#Expenditure 
def Expenditure():


    wb = load_workbook('Excel_Files/Dashboard/BankStatement.xlsx')
    sumr= wb['summary']
    eod= wb['EOD Balance']

    Sum_of_debits = df1.groupby(df1['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
#     print(Sum_of_debits)

    # Interest Paid 
    df_interest =df1[df1["Particulars"].str.contains("Int.Pd")]
    Interest_paid=df_interest.groupby(df_interest['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
    # print(df_interest)
    Int_dr = df1[df1["Particulars"].str.contains("Int.Pd")].groupby('Debit')
    Int_dr = Int_dr.apply(lambda x: x)
    Int_dr['Categories'] = 'Intrest Debit'
    df1.update(Int_dr)
    
        # Bank charges
    df_charges =df1[df1["Particulars"].str.contains("Consolidated Charges for A/c")|df1["Particulars"].str.contains("Dr Card Charges ANNUAL")]
    Bank_charges = df_charges.groupby(df_charges['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
    df_charges['Categories'] = 'Bank charges'
    df1.update(df_charges)
    # print(df_charges)

        # Emi
    new_emi = df1[df1['Particulars'].str.contains('ECS') | df1['Particulars'].str.contains('NACH')]
    Emi_dr = new_emi.groupby(new_emi['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
    new_emi['Categories'] = 'EMI'
    df1.update(new_emi)
    # print(Emi_dr)

        # GST Paid
    df_gst = df1[df1["Particulars"].str.contains("GST")]
    gst = df_gst.groupby(df_gst['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
    df_gst['Categories'] = 'GST Paid'
    df1.update(df_gst)
    # print(df_gst)

    # Utility bills
    new_utility = df1[df1['Particulars'].str.contains('NACH-DR- BSNL') | df1['Particulars'].str.contains('NACH-DR- Vodafone')| df1['Particulars'].str.contains('NACH-DR- Reliance')| df1['Particulars'].str.contains('NACH-DR- MTNL')| df1['Particulars'].str.contains('NACH-DR- Reliance JIO')| df1['Particulars'].str.contains('NACH-DR- Airtel')| df1['Particulars'].str.contains('NTPC')| df1['Particulars'].str.contains('Adani Group')| df1['Particulars'].str.contains('Tata Power')| df1['Particulars'].str.contains('JSW Energy')| df1['Particulars'].str.contains('Torrent Power')| df1['Particulars'].str.contains('Powergrid')| df1['Particulars'].str.contains('NHPC')| df1['Particulars'].str.contains('Reliance Power Ltd')| df1['Particulars'].str.contains('SJVN')| df1['Particulars'].str.contains('CESC')| df1['Particulars'].str.contains('Oil and Natural Gas Corporation')| df1['Particulars'].str.contains('Indian Oil Corporation Limited')| df1['Particulars'].str.contains('Bharat Petroleum')| df1['Particulars'].str.contains('Gas Authority of India')| df1['Particulars'].str.contains('Reliance Petroleum Ltd')| df1['Particulars'].str.contains('Hindustan Petroleum')| df1['Particulars'].str.contains('Oil India')| df1['Particulars'].str.contains('Cairn India')| df1['Particulars'].str.contains('TATA Petrodyne Ltd')| df1['Particulars'].str.contains('Essar Oil')|df1['Particulars'].str.contains('Jio Rec')]
    Utility = new_utility.groupby(new_utility['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
    # print(new_utility)
    new_utility['Categories'] = 'Utility Bills'
    df1.update(new_utility)

    # Traveling bills
    Traveling_expence_df = df1[df1["Particulars"].str.contains("Ola")| df1['Particulars'].str.contains('Uber')| df1['Particulars'].str.contains('Irctc')| df1['Particulars'].str.contains('Indian Railways')]
    TravelingExpence = Traveling_expence_df.groupby(Traveling_expence_df['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
    # print(TravelingExpence)
    Traveling_expence_df['Categories'] = 'Traveling bills'
    df1.update(Traveling_expence_df)

    rent_df = df1[df1['Particulars'].str.contains('RENT') | df1['Particulars'].str.contains('Rent')]
    rent_dr = rent_df.groupby(rent_df['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
    RENT1 = df1[df1["Particulars"].str.contains("RENT")| df1['Particulars'].str.contains('Rent')].groupby('Debit')
    RENT1 = RENT1.apply(lambda x: x)
    RENT1['Categories'] = 'RENT Paid'
    df1.update(RENT1)
    # print(rent_dr)

    #   Salary Paid
    df_salary = df1[df1["Particulars"].str.contains("SALARY",na=False)]
    salary=df_salary.groupby(df_salary['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
    SAL = df1[df1["Particulars"].str.contains("SALARY")].groupby('Debit')
    SAL = SAL.apply(lambda x: x)
    SAL['Categories'] = 'Salary Paid'
    df1.update(SAL)

    # Tds deduction
    TDS = df1[df1["Particulars"].str.contains("TDS",na=False)]
    tds=TDS.groupby(TDS['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()

    # Total income tax paid
    TAX = df1[df1["Particulars"].str.contains("tax",na=False)]
    tax=TAX.groupby(TAX['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()


    try:
       sumr.cell(row=39,column=3,value=salary[sumr.cell(row=13,column=3).value])
    except Exception as e:
      sumr.cell(row=39,column=3,value=0)
    try:
       sumr.cell(row=39,column=4,value=salary[sumr.cell(row=13,column=4).value])
    except Exception as e:
      sumr.cell(row=39,column=4,value=0)
    try:
       sumr.cell(row=39,column=5,value=salary[sumr.cell(row=13,column=5).value])
    except Exception as e:
      sumr.cell(row=39,column=5,value=0)
    try:
       sumr.cell(row=39,column=6,value=salary[sumr.cell(row=13,column=6).value])
    except Exception as e:
      sumr.cell(row=39,column=6,value=0)
    try:
       sumr.cell(row=39,column=7,value=salary[sumr.cell(row=13,column=7).value])
    except Exception as e:
      sumr.cell(row=39,column=7,value=0)
    try:
       sumr.cell(row=39,column=8,value=salary[sumr.cell(row=13,column=8).value])
    except Exception as e:
       sumr.cell(row=39,column=8,value=0)
    try:
       sumr.cell(row=39,column=9,value=salary[sumr.cell(row=13,column=9).value])
    except Exception as e:
       sumr.cell(row=39,column=9,value=0)
    try:
       sumr.cell(row=39,column=10,value=salary[sumr.cell(row=13,column=10).value])
    except Exception as e:
       sumr.cell(row=39,column=10,value=0)
    try:
       sumr.cell(row=39,column=11,value=salary[sumr.cell(row=13,column=11).value])
    except Exception as e:
       sumr.cell(row=39,column=11,value=0)
    try:
       sumr.cell(row=39,column=12,value=salary[sumr.cell(row=13,column=12).value])
    except Exception as e:
       sumr.cell(row=39,column=12,value=0)
    try:
       sumr.cell(row=39,column=13,value=salary[sumr.cell(row=13,column=13).value])
    except Exception as e:
       sumr.cell(row=39,column=13,value=0)
    try:
       sumr.cell(row=39,column=14,value=salary[sumr.cell(row=13,column=14).value])
    except Exception as e:
       sumr.cell(row=39,column=14,value=0)


    


    try:
        sumr.cell(row=42, column=3, value=tds[sumr.cell(row=13, column=3).value])
    except Exception as e:
        sumr.cell(row=42, column=3, value=0)
    try:
        sumr.cell(row=42, column=4, value=tds[sumr.cell(row=13, column=4).value])
    except Exception as e:
        sumr.cell(row=42, column=4, value=0)
    try:
        sumr.cell(row=42, column=5, value=tds[sumr.cell(row=13, column=5).value])
    except Exception as e:
        sumr.cell(row=42, column=5, value=0)
    try:
        sumr.cell(row=42, column=6, value=tds[sumr.cell(row=13, column=6).value])
    except Exception as e:
        sumr.cell(row=42, column=6, value=0)
    try:
        sumr.cell(row=42, column=7, value=tds[sumr.cell(row=13, column=7).value])
    except Exception as e:
        sumr.cell(row=42, column=7, value=0)
    try:
        sumr.cell(row=42, column=8, value=tds[sumr.cell(row=13, column=8).value])
    except Exception as e:
        sumr.cell(row=42, column=8, value=0)
    try:
        sumr.cell(row=42, column=9, value=tds[sumr.cell(row=13, column=9).value])
    except Exception as e:
        sumr.cell(row=42, column=9, value=0)
    try:
        sumr.cell(row=42, column=10, value=tds[sumr.cell(row=13, column=10).value])
    except Exception as e:
        sumr.cell(row=42, column=10, value=0)
    try:
        sumr.cell(row=42, column=11, value=tds[sumr.cell(row=13, column=11).value])
    except Exception as e:
        sumr.cell(row=42, column=11, value=0)
    try:
        sumr.cell(row=42, column=12, value=tds[sumr.cell(row=13, column=12).value])
    except Exception as e:
        sumr.cell(row=42, column=12, value=0)
    try:
        sumr.cell(row=42, column=13, value=tds[sumr.cell(row=13, column=13).value])
    except Exception as e:
        sumr.cell(row=42, column=13, value=0)
    try:
        sumr.cell(row=42, column=14, value=tds[sumr.cell(row=13, column=14).value])
    except Exception as e:
        sumr.cell(row=42, column=14, value=0)

    try:
        sumr.cell(row=68, column=3, value=tds[sumr.cell(row=13, column=3).value])
    except Exception as e:
        sumr.cell(row=68, column=3, value=0)
    try:
        sumr.cell(row=68, column=4, value=tds[sumr.cell(row=13, column=4).value])
    except Exception as e:
        sumr.cell(row=68, column=4, value=0)
    try:
        sumr.cell(row=68, column=5, value=tds[sumr.cell(row=13, column=5).value])
    except Exception as e:
        sumr.cell(row=68, column=5, value=0)
    try:
        sumr.cell(row=68, column=6, value=tds[sumr.cell(row=13, column=6).value])
    except Exception as e:
        sumr.cell(row=68, column=6, value=0)
    try:
        sumr.cell(row=68, column=7, value=tds[sumr.cell(row=13, column=7).value])
    except Exception as e:
        sumr.cell(row=68, column=7, value=0)
    try:
        sumr.cell(row=68, column=8, value=tds[sumr.cell(row=13, column=8).value])
    except Exception as e:
        sumr.cell(row=68, column=8, value=0)
    try:
        sumr.cell(row=68, column=9, value=tds[sumr.cell(row=13, column=9).value])
    except Exception as e:
        sumr.cell(row=68, column=9, value=0)
    try:
        sumr.cell(row=68, column=10, value=tds[sumr.cell(row=13, column=10).value])
    except Exception as e:
        sumr.cell(row=68, column=10, value=0)
    try:
        sumr.cell(row=68, column=11, value=tds[sumr.cell(row=13, column=11).value])
    except Exception as e:
        sumr.cell(row=68, column=11, value=0)
    try:
        sumr.cell(row=68, column=12, value=tds[sumr.cell(row=13, column=12).value])
    except Exception as e:
        sumr.cell(row=68, column=12, value=0)
    try:
        sumr.cell(row=68, column=13, value=tds[sumr.cell(row=13, column=13).value])
    except Exception as e:
        sumr.cell(row=68, column=13, value=0)
    try:
        sumr.cell(row=68, column=14, value=tds[sumr.cell(row=13, column=14).value])
    except Exception as e:
        sumr.cell(row=68, column=14, value=0)

            
    try:
        sumr.cell(row=44, column=3, value=tax[sumr.cell(row=13, column=3).value])
    except Exception as e:
        sumr.cell(row=44, column=3, value=0)
    try:
        sumr.cell(row=44, column=4, value=tax[sumr.cell(row=13, column=4).value])
    except Exception as e:
        sumr.cell(row=44, column=4, value=0)
    try:
        sumr.cell(row=44, column=5, value=tax[sumr.cell(row=13, column=5).value])
    except Exception as e:
        sumr.cell(row=44, column=5, value=0)
    try:
        sumr.cell(row=44, column=6, value=tax[sumr.cell(row=13, column=6).value])
    except Exception as e:
        sumr.cell(row=44, column=6, value=0)
    try:
        sumr.cell(row=44, column=7, value=tax[sumr.cell(row=13, column=7).value])
    except Exception as e:
        sumr.cell(row=44, column=7, value=0)
    try:
        sumr.cell(row=44, column=8, value=tax[sumr.cell(row=13, column=8).value])
    except Exception as e:
        sumr.cell(row=44, column=8, value=0)
    try:
        sumr.cell(row=44, column=9, value=tax[sumr.cell(row=13, column=9).value])
    except Exception as e:
        sumr.cell(row=44, column=9, value=0)
    try:
        sumr.cell(row=44, column=10, value=tax[sumr.cell(row=13, column=10).value])
    except Exception as e:
        sumr.cell(row=44, column=10, value=0)
    try:
        sumr.cell(row=44, column=11, value=tax[sumr.cell(row=13, column=11).value])
    except Exception as e:
        sumr.cell(row=44, column=11, value=0)
    try:
        sumr.cell(row=44, column=12, value=tax[sumr.cell(row=13, column=12).value])
    except Exception as e:
        sumr.cell(row=44, column=12, value=0)
    try:
        sumr.cell(row=44, column=13, value=tax[sumr.cell(row=13, column=13).value])
    except Exception as e:
        sumr.cell(row=44, column=13, value=0)
    try:
        sumr.cell(row=44, column=14, value=tax[sumr.cell(row=13, column=14).value])
    except Exception as e:
        sumr.cell(row=44, column=14, value=0)


    try:
       sumr.cell(row=46,column=3,value=TravelingExpence[sumr.cell(row=13,column=3).value])
    except Exception as e:
      sumr.cell(row=46,column=3,value=0)
    try:
       sumr.cell(row=46,column=4,value=TravelingExpence[sumr.cell(row=13,column=4).value])
    except Exception as e:
      sumr.cell(row=46,column=4,value=0)
    try:
       sumr.cell(row=46,column=5,value=TravelingExpence[sumr.cell(row=13,column=5).value])
    except Exception as e:
      sumr.cell(row=46,column=5,value=0)
    try:
       sumr.cell(row=46,column=6,value=TravelingExpence[sumr.cell(row=13,column=6).value])
    except Exception as e:
      sumr.cell(row=46,column=6,value=0)
    try:
       sumr.cell(row=46,column=7,value=TravelingExpence[sumr.cell(row=13,column=7).value])
    except Exception as e:
      sumr.cell(row=46,column=7,value=0)
    try:
       sumr.cell(row=46,column=8,value=TravelingExpence[sumr.cell(row=13,column=8).value])
    except Exception as e:
      sumr.cell(row=46,column=8,value=0)
    try:
       sumr.cell(row=46,column=9,value=TravelingExpence[sumr.cell(row=13,column=9).value])
    except Exception as e:
      sumr.cell(row=46,column=9,value=0)
    try:
       sumr.cell(row=46,column=10,value=TravelingExpence[sumr.cell(row=13,column=10).value])
    except Exception as e:
      sumr.cell(row=46,column=10,value=0)
    try:
       sumr.cell(row=46,column=11,value=TravelingExpence[sumr.cell(row=13,column=11).value])
    except Exception as e:
      sumr.cell(row=46,column=11,value=0)
    try:
       sumr.cell(row=46,column=12,value=TravelingExpence[sumr.cell(row=13,column=12).value])
    except Exception as e:
      sumr.cell(row=46,column=12,value=0)
    try:
       sumr.cell(row=46,column=13,value=TravelingExpence[sumr.cell(row=13,column=13).value])
    except Exception as e:
      sumr.cell(row=46,column=13,value=0)
    try:
       sumr.cell(row=46,column=14,value=TravelingExpence[sumr.cell(row=13,column=14).value])
    except Exception as e:
      sumr.cell(row=46,column=14,value=0)

    try:
       sumr.cell(row=72,column=3,value=TravelingExpence[sumr.cell(row=13,column=3).value])
    except Exception as e:
      sumr.cell(row=72,column=3,value=0)
    try:
       sumr.cell(row=72,column=4,value=TravelingExpence[sumr.cell(row=13,column=4).value])
    except Exception as e:
      sumr.cell(row=72,column=4,value=0)
    try:
       sumr.cell(row=72,column=5,value=TravelingExpence[sumr.cell(row=13,column=5).value])
    except Exception as e:
      sumr.cell(row=72,column=5,value=0)
    try:
       sumr.cell(row=72,column=6,value=TravelingExpence[sumr.cell(row=13,column=6).value])
    except Exception as e:
      sumr.cell(row=72,column=6,value=0)
    try:
       sumr.cell(row=72,column=7,value=TravelingExpence[sumr.cell(row=13,column=7).value])
    except Exception as e:
      sumr.cell(row=72,column=7,value=0)
    try:
       sumr.cell(row=72,column=8,value=TravelingExpence[sumr.cell(row=13,column=8).value])
    except Exception as e:
      sumr.cell(row=72,column=8,value=0)
    try:
       sumr.cell(row=72,column=9,value=TravelingExpence[sumr.cell(row=13,column=9).value])
    except Exception as e:
      sumr.cell(row=72,column=9,value=0)
    try:
       sumr.cell(row=72,column=10,value=TravelingExpence[sumr.cell(row=13,column=10).value])
    except Exception as e:
      sumr.cell(row=72,column=10,value=0)
    try:
       sumr.cell(row=72,column=11,value=TravelingExpence[sumr.cell(row=13,column=11).value])
    except Exception as e:
      sumr.cell(row=72,column=11,value=0)
    try:
       sumr.cell(row=72,column=12,value=TravelingExpence[sumr.cell(row=13,column=12).value])
    except Exception as e:
      sumr.cell(row=72,column=12,value=0)
    try:
       sumr.cell(row=72,column=13,value=TravelingExpence[sumr.cell(row=13,column=13).value])
    except Exception as e:
      sumr.cell(row=72,column=13,value=0)
    try:
       sumr.cell(row=72,column=14,value=TravelingExpence[sumr.cell(row=13,column=14).value])
    except Exception as e:
      sumr.cell(row=72,column=14,value=0)


    try:
       sumr.cell(row=38,column=3,value=Interest_paid[sumr.cell(row=13,column=3).value])
    except Exception as e:
      sumr.cell(row=38,column=3,value=0)
    try:
       sumr.cell(row=38,column=4,value=Interest_paid[sumr.cell(row=13,column=4).value])
    except Exception as e:
      sumr.cell(row=38,column=4,value=0)
    try:
       sumr.cell(row=38,column=5,value=Interest_paid[sumr.cell(row=13,column=5).value])
    except Exception as e:
      sumr.cell(row=38,column=5,value=0)
    try:
       sumr.cell(row=38,column=6,value=Interest_paid[sumr.cell(row=13,column=6).value])
    except Exception as e:
      sumr.cell(row=38,column=6,value=0)
    try:
       sumr.cell(row=38,column=7,value=Interest_paid[sumr.cell(row=13,column=7).value])
    except Exception as e:
      sumr.cell(row=38,column=7,value=0)
    try:
       sumr.cell(row=38,column=8,value=Interest_paid[sumr.cell(row=13,column=8).value])
    except Exception as e:
      sumr.cell(row=38,column=8,value=0)
    try:
       sumr.cell(row=38,column=9,value=Interest_paid[sumr.cell(row=13,column=9).value])
    except Exception as e:
      sumr.cell(row=38,column=9,value=0)
    try:
       sumr.cell(row=38,column=10,value=Interest_paid[sumr.cell(row=13,column=10).value])
    except Exception as e:
      sumr.cell(row=38,column=10,value=0)
    try:
       sumr.cell(row=38,column=11,value=Interest_paid[sumr.cell(row=13,column=11).value])
    except Exception as e:
      sumr.cell(row=38,column=11,value=0)
    try:
       sumr.cell(row=38,column=12,value=Interest_paid[sumr.cell(row=13,column=12).value])
    except Exception as e:
      sumr.cell(row=38,column=12,value=0)
    try:
       sumr.cell(row=38,column=13,value=Interest_paid[sumr.cell(row=13,column=13).value])
    except Exception as e:
      sumr.cell(row=38,column=13,value=0)
    try:
       sumr.cell(row=38,column=14,value=Interest_paid[sumr.cell(row=13,column=14).value])
    except Exception as e:
      sumr.cell(row=38,column=14,value=0)


    try:
       sumr.cell(row=40,column=3,value=Bank_charges[sumr.cell(row=13,column=3).value])
    except Exception as e:
      sumr.cell(row=40,column=3,value=0)
    try:
       sumr.cell(row=40,column=4,value=Bank_charges[sumr.cell(row=13,column=4).value])
    except Exception as e:
      sumr.cell(row=40,column=4,value=0)
    try:
       sumr.cell(row=40,column=5,value=Bank_charges[sumr.cell(row=13,column=5).value])
    except Exception as e:
      sumr.cell(row=40,column=5,value=0)
    try:
       sumr.cell(row=40,column=6,value=Bank_charges[sumr.cell(row=13,column=6).value])
    except Exception as e:
      sumr.cell(row=40,column=6,value=0)
    try:
       sumr.cell(row=40,column=7,value=Bank_charges[sumr.cell(row=13,column=7).value])
    except Exception as e:
      sumr.cell(row=40,column=7,value=0)
    try:
       sumr.cell(row=40,column=8,value=Bank_charges[sumr.cell(row=13,column=8).value])
    except Exception as e:
      sumr.cell(row=40,column=8,value=0)
    try:
       sumr.cell(row=40,column=9,value=Bank_charges[sumr.cell(row=13,column=9).value])
    except Exception as e:
      sumr.cell(row=40,column=9,value=0)
    try:
       sumr.cell(row=40,column=10,value=Bank_charges[sumr.cell(row=13,column=10).value])
    except Exception as e:
      sumr.cell(row=40,column=10,value=0)
    try:
       sumr.cell(row=40,column=11,value=Bank_charges[sumr.cell(row=13,column=11).value])
    except Exception as e:
      sumr.cell(row=40,column=11,value=0)
    try:
       sumr.cell(row=40,column=12,value=Bank_charges[sumr.cell(row=13,column=12).value])
    except Exception as e:
      sumr.cell(row=40,column=12,value=0)
    try:
       sumr.cell(row=40,column=13,value=Bank_charges[sumr.cell(row=13,column=13).value])
    except Exception as e:
      sumr.cell(row=40,column=13,value=0)
    try:
       sumr.cell(row=40,column=14,value=Bank_charges[sumr.cell(row=13,column=14).value])
    except Exception as e:
      sumr.cell(row=40,column=14,value=0)




    try:
       sumr.cell(row=41,column=3,value=Emi_dr[sumr.cell(row=13,column=3).value])
    except Exception as e:
      sumr.cell(row=41,column=3,value=0)
    try:
       sumr.cell(row=41,column=4,value=Emi_dr[sumr.cell(row=13,column=4).value])
    except Exception as e:
      sumr.cell(row=41,column=4,value=0)
    try:
       sumr.cell(row=41,column=5,value=Emi_dr[sumr.cell(row=13,column=5).value])
    except Exception as e:
      sumr.cell(row=41,column=5,value=0)
    try:
       sumr.cell(row=41,column=6,value=Emi_dr[sumr.cell(row=13,column=6).value])
    except Exception as e:
      sumr.cell(row=41,column=6,value=0)
    try:
       sumr.cell(row=41,column=7,value=Emi_dr[sumr.cell(row=13,column=7).value])
    except Exception as e:
      sumr.cell(row=41,column=7,value=0)
    try:
       sumr.cell(row=41,column=8,value=Emi_dr[sumr.cell(row=13,column=8).value])
    except Exception as e:
      sumr.cell(row=41,column=8,value=0)
    try:
       sumr.cell(row=41,column=9,value=Emi_dr[sumr.cell(row=13,column=9).value])
    except Exception as e:
      sumr.cell(row=41,column=9,value=0)
    try:
       sumr.cell(row=41,column=10,value=Emi_dr[sumr.cell(row=13,column=10).value])
    except Exception as e:
      sumr.cell(row=41,column=10,value=0)
    try:
       sumr.cell(row=41,column=11,value=Emi_dr[sumr.cell(row=13,column=11).value])
    except Exception as e:
      sumr.cell(row=41,column=11,value=0)
    try:
       sumr.cell(row=41,column=12,value=Emi_dr[sumr.cell(row=13,column=12).value])
    except Exception as e:
      sumr.cell(row=41,column=12,value=0)
    try:
       sumr.cell(row=41,column=13,value=Emi_dr[sumr.cell(row=13,column=13).value])
    except Exception as e:
      sumr.cell(row=41,column=13,value=0)
    try:
       sumr.cell(row=41,column=14,value=Emi_dr[sumr.cell(row=13,column=14).value])
    except Exception as e:
      sumr.cell(row=41,column=14,value=0)



    try:
       sumr.cell(row=43,column=3,value=gst[sumr.cell(row=13,column=3).value])
    except Exception as e:
      sumr.cell(row=43,column=3,value=0)
    try:
       sumr.cell(row=43,column=4,value=gst[sumr.cell(row=13,column=4).value])
    except Exception as e:
      sumr.cell(row=43,column=4,value=0)
    try:
       sumr.cell(row=43,column=5,value=gst[sumr.cell(row=13,column=5).value])
    except Exception as e:
      sumr.cell(row=43,column=5,value=0)
    try:
       sumr.cell(row=43,column=6,value=gst[sumr.cell(row=13,column=6).value])
    except Exception as e:
      sumr.cell(row=43,column=6,value=0)
    try:
       sumr.cell(row=43,column=7,value=gst[sumr.cell(row=13,column=7).value])
    except Exception as e:
      sumr.cell(row=43,column=7,value=0)
    try:
       sumr.cell(row=43,column=8,value=gst[sumr.cell(row=13,column=8).value])
    except Exception as e:
      sumr.cell(row=43,column=8,value=0)
    try:
       sumr.cell(row=43,column=9,value=gst[sumr.cell(row=13,column=9).value])
    except Exception as e:
      sumr.cell(row=43,column=9,value=0)
    try:
       sumr.cell(row=43,column=10,value=gst[sumr.cell(row=13,column=10).value])
    except Exception as e:
      sumr.cell(row=43,column=10,value=0)
    try:
       sumr.cell(row=43,column=11,value=gst[sumr.cell(row=13,column=11).value])
    except Exception as e:
      sumr.cell(row=43,column=11,value=0)
    try:
       sumr.cell(row=43,column=12,value=gst[sumr.cell(row=13,column=12).value])
    except Exception as e:
      sumr.cell(row=43,column=12,value=0)
    try:
       sumr.cell(row=43,column=13,value=gst[sumr.cell(row=13,column=13).value])
    except Exception as e:
      sumr.cell(row=43,column=13,value=0)
    try:
       sumr.cell(row=43,column=14,value=gst[sumr.cell(row=13,column=14).value])
    except Exception as e:
      sumr.cell(row=43,column=14,value=0)


    try:
       sumr.cell(row=47,column=3,value=rent_dr[sumr.cell(row=13,column=3).value])
    except Exception as e:
      sumr.cell(row=47,column=3,value=0)
    try:
       sumr.cell(row=47,column=4,value=rent_dr[sumr.cell(row=13,column=4).value])
    except Exception as e:
      sumr.cell(row=47,column=4,value=0)
    try:
       sumr.cell(row=47,column=5,value=rent_dr[sumr.cell(row=13,column=5).value])
    except Exception as e:
      sumr.cell(row=47,column=5,value=0)
    try:
       sumr.cell(row=47,column=6,value=rent_dr[sumr.cell(row=13,column=6).value])
    except Exception as e:
      sumr.cell(row=47,column=6,value=0)
    try:
       sumr.cell(row=47,column=7,value=rent_dr[sumr.cell(row=13,column=7).value])
    except Exception as e:
      sumr.cell(row=47,column=7,value=0)
    try:
       sumr.cell(row=47,column=8,value=rent_dr[sumr.cell(row=13,column=8).value])
    except Exception as e:
      sumr.cell(row=47,column=8,value=0)
    try:
       sumr.cell(row=47,column=9,value=rent_dr[sumr.cell(row=13,column=9).value])
    except Exception as e:
      sumr.cell(row=47,column=9,value=0)
    try:
       sumr.cell(row=47,column=10,value=rent_dr[sumr.cell(row=13,column=10).value])
    except Exception as e:
      sumr.cell(row=47,column=10,value=0)
    try:
       sumr.cell(row=47,column=11,value=rent_dr[sumr.cell(row=13,column=11).value])
    except Exception as e:
      sumr.cell(row=47,column=11,value=0)
    try:
       sumr.cell(row=47,column=12,value=rent_dr[sumr.cell(row=13,column=12).value])
    except Exception as e:
      sumr.cell(row=47,column=12,value=0)
    try:
       sumr.cell(row=47,column=13,value=rent_dr[sumr.cell(row=13,column=13).value])
    except Exception as e:
      sumr.cell(row=47,column=13,value=0)
    try:
       sumr.cell(row=47,column=14,value=rent_dr[sumr.cell(row=13,column=14).value])
    except Exception as e:
      sumr.cell(row=47,column=14,value=0)




    wb.save('Excel_Files/Dashboard/BankStatement.xlsx')

Expenditure()

def PersonalExpenses():


     wb = load_workbook('Excel_Files/Dashboard/BankStatement.xlsx')
     sumr= wb['summary']
     eod= wb['EOD Balance']

     General_insurance_df=df1[df1['Particulars'].str.contains('National Insurance Co. Ltd.') |df1['Particulars'].str.contains('Go Digit General Insurance Ltd.') |df1['Particulars'].str.contains('Bajaj Allianz General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Cholamandalam MS General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Bharti AXA General Insurance Co. Ltd.') |df1['Particulars'].str.contains('HDFC ERGO General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Future Generali India Insurance Co. Ltd.') |df1['Particulars'].str.contains('The New India Assurance Co. Ltd.') |df1['Particulars'].str.contains('Iffco Tokio General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Reliance General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Royal Sundaram General Insurance Co. Ltd.') |df1['Particulars'].str.contains('The Oriental Insurance Co. Ltd.') |df1['Particulars'].str.contains('Tata AIG General Insurance Co. Ltd.') |df1['Particulars'].str.contains('SBI General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Acko General Insurance Ltd.') |df1['Particulars'].str.contains('Navi General Insurance Ltd.') |df1['Particulars'].str.contains('Edelweiss General Insurance Co. Ltd.') |df1['Particulars'].str.contains('ICICI Lombard General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Kotak Mahindra General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Liberty General Insurance Ltd.') |df1['Particulars'].str.contains('Magma HDI General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Raheja QBE General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Shriram General Insurance Co. Ltd.') |df1['Particulars'].str.contains('United India Insurance Co. Ltd.') |df1['Particulars'].str.contains('Universal Sompo General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Agriculture Insurance Company of India Ltd.') |df1['Particulars'].str.contains('Aditya Birla Health Insurance Co. Ltd.') |df1['Particulars'].str.contains('Manipal Cigna Health Insurance Company Limited') |df1['Particulars'].str.contains('ECGC Ltd.') |df1['Particulars'].str.contains('Max Bupa Health Insurance Co. Ltd') |df1['Particulars'].str.contains('Care Health Insurance Ltd') |df1['Particulars'].str.contains('Star Health & Allied Insurance Co. Ltd.')]
     GeneralInsurance = General_insurance_df.groupby(General_insurance_df['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
     General_insurance_df['Categories'] = 'General insurance'
     df1.update(General_insurance_df)

     Life_insuranc_df= df1[df1['Particulars'].str.contains('Life Insurance Corporation of India') |df1['Particulars'].str.contains('HDFC Standard Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Max Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('ICICI Prudential Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Kotak Mahindra Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Aditya Birla Sun Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('TATA AIA Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('SBI Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Exide Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Bajaj Allianz Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('PNB MetLife India Insurance Co. Ltd.') |df1['Particulars'].str.contains('Reliance Nippon Life Insurance Company') |df1['Particulars'].str.contains('Aviva Life Insurance Company India Ltd.') |df1['Particulars'].str.contains('Sahara India Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Shriram Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Bharti AXA Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Future Generali India Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('IDBI Federal Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Canara HSBC Oriental Bank of Commerce Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Aegon Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Pramerica Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Star Union Dai-Ichi Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('IndiaFirst Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Edelweiss Tokio Life Insurance Co. Ltd.')]
     LifeInsurance = Life_insuranc_df.groupby(Life_insuranc_df['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
     Life_insuranc_df['Categories'] = 'Life insurance'
     df1.update(Life_insuranc_df)
    
     Credit_card_df=df1[df1['Particulars'].str.contains('Credit Card Payment')]
     CreditCharges = Credit_card_df.groupby(Credit_card_df['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
    #  print(CreditCharges)
     Credit_card_df['Categories'] = 'Credit card'
     df1.update(Credit_card_df)

     Food_Expenxe_df = df1[df1['Particulars'].str.contains('Zomato')|df1['Particulars'].str.contains('Swiggy')|df1['Particulars'].str.contains('Pizza')|df1['Particulars'].str.contains('Dominos')]
     FoodExpence = Food_Expenxe_df.groupby(Food_Expenxe_df['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
    #  print(FoodExpence)
     Food_Expenxe_df['Categories'] = 'Food Expence'
     df1.update(Food_Expenxe_df)

     online_shopping_df = df1[df1['Particulars'].str.contains('AMAZON -Dr')|df1['Particulars'].str.contains('Flipkart')|df1['Particulars'].str.contains('FLIPKART')]
     OnlineShopping = online_shopping_df.groupby(online_shopping_df['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
  # print(OnlineShopping)
     online_shopping_df['Categories'] = 'Online shopping'
     df1.update(online_shopping_df)

     gas_payments_df =df1[df1['Particulars'].str.contains('Oil and Natural Gas Corporation')| df1['Particulars'].str.contains('Indian Oil Corporation Limited')| df1['Particulars'].str.contains('Bharat Petroleum')| df1['Particulars'].str.contains('Gas Authority of India')| df1['Particulars'].str.contains('Reliance Petroleum Ltd')| df1['Particulars'].str.contains('Hindustan Petroleum')| df1['Particulars'].str.contains('Oil India')| df1['Particulars'].str.contains('Cairn India')| df1['Particulars'].str.contains('TATA Petrodyne Ltd')| df1['Particulars'].str.contains('Essar Oil')]
     GasPayment = gas_payments_df.groupby(gas_payments_df['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
  # print(GasPayment)
     gas_payments_df['Categories'] = 'gas Payments'
     df1.update(gas_payments_df)

     Gold_loan_df = df1[df1['Particulars'].str.contains('MUTHOOTFINCORP')|df1['Particulars'].str.contains('RUPEEKCAPITAL')]
     GoldLoan=Gold_loan_df.groupby(Gold_loan_df['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
  # print(GoldLoan)
     Gold_loan_df['Categories'] = 'Gold Loan'
     df1.update(Gold_loan_df)

         # Rent Paid
     rent_df = df1[df1['Particulars'].str.contains('Rent')]
     rent_dr = rent_df.groupby(rent_df['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
     #print(rent_dr)

    # Property Tax
     Property_Tax = df1[df1['Particulars'].str.contains('Property Tax')]
     pt = Property_Tax.groupby(Property_Tax['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()

     try:
       sumr.cell(row=51,column=3,value=GeneralInsurance[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=51,column=3,value=0)
     try:
       sumr.cell(row=51,column=4,value=GeneralInsurance[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=51,column=4,value=0)
     try:
       sumr.cell(row=51,column=5,value=GeneralInsurance[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=51,column=5,value=0)
     try:
       sumr.cell(row=51,column=6,value=GeneralInsurance[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=51,column=6,value=0)
     try:
       sumr.cell(row=51,column=7,value=GeneralInsurance[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=51,column=7,value=0)
     try:
       sumr.cell(row=51,column=8,value=GeneralInsurance[sumr.cell(row=13,column=8).value])
     except Exception as e:
       sumr.cell(row=51,column=8,value=0)
     try:
       sumr.cell(row=51,column=9,value=GeneralInsurance[sumr.cell(row=13,column=9).value])
     except Exception as e:
       sumr.cell(row=51,column=9,value=0)
     try:
       sumr.cell(row=51,column=10,value=GeneralInsurance[sumr.cell(row=13,column=10).value])
     except Exception as e:
       sumr.cell(row=51,column=10,value=0)
     try:
       sumr.cell(row=51,column=11,value=GeneralInsurance[sumr.cell(row=13,column=11).value])
     except Exception as e:
       sumr.cell(row=51,column=11,value=0)
     try:
       sumr.cell(row=51,column=12,value=GeneralInsurance[sumr.cell(row=13,column=12).value])
     except Exception as e:
       sumr.cell(row=51,column=12,value=0)
     try:
       sumr.cell(row=51,column=13,value=GeneralInsurance[sumr.cell(row=13,column=13).value])
     except Exception as e:
       sumr.cell(row=51,column=13,value=0)
     try:
       sumr.cell(row=51,column=14,value=GeneralInsurance[sumr.cell(row=13,column=14).value])
     except Exception as e:
       sumr.cell(row=51,column=14,value=0)



     try:
       sumr.cell(row=76,column=3,value=GeneralInsurance[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=76,column=3,value=0)
     try:
       sumr.cell(row=76,column=4,value=GeneralInsurance[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=76,column=4,value=0)
     try:
       sumr.cell(row=76,column=5,value=GeneralInsurance[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=76,column=5,value=0)
     try:
       sumr.cell(row=76,column=6,value=GeneralInsurance[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=76,column=6,value=0)
     try:
       sumr.cell(row=76,column=7,value=GeneralInsurance[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=76,column=7,value=0)
     try:
       sumr.cell(row=76,column=8,value=GeneralInsurance[sumr.cell(row=13,column=8).value])
     except Exception as e:
       sumr.cell(row=76,column=8,value=0)
     try:
       sumr.cell(row=76,column=9,value=GeneralInsurance[sumr.cell(row=13,column=9).value])
     except Exception as e:
       sumr.cell(row=76,column=9,value=0)
     try:
       sumr.cell(row=76,column=10,value=GeneralInsurance[sumr.cell(row=13,column=10).value])
     except Exception as e:
       sumr.cell(row=76,column=10,value=0)
     try:
       sumr.cell(row=76,column=11,value=GeneralInsurance[sumr.cell(row=13,column=11).value])
     except Exception as e:
       sumr.cell(row=76,column=11,value=0)
     try:
       sumr.cell(row=76,column=12,value=GeneralInsurance[sumr.cell(row=13,column=12).value])
     except Exception as e:
       sumr.cell(row=76,column=12,value=0)
     try:
       sumr.cell(row=76,column=13,value=GeneralInsurance[sumr.cell(row=13,column=13).value])
     except Exception as e:
       sumr.cell(row=76,column=13,value=0)
     try:
       sumr.cell(row=76,column=14,value=GeneralInsurance[sumr.cell(row=13,column=14).value])
     except Exception as e:
       sumr.cell(row=76,column=14,value=0)



     try:
       sumr.cell(row=52,column=3,value=LifeInsurance[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=52,column=3,value=0)
     try:
       sumr.cell(row=52,column=4,value=LifeInsurance[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=52,column=4,value=0)
     try:
       sumr.cell(row=52,column=5,value=LifeInsurance[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=52,column=5,value=0)
     try:
       sumr.cell(row=52,column=6,value=LifeInsurance[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=52,column=6,value=0)
     try:
       sumr.cell(row=52,column=7,value=LifeInsurance[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=52,column=7,value=0)
     try:
       sumr.cell(row=52,column=8,value=LifeInsurance[sumr.cell(row=13,column=8).value])
     except Exception as e:
      sumr.cell(row=52,column=8,value=0)
     try:
       sumr.cell(row=52,column=9,value=LifeInsurance[sumr.cell(row=13,column=9).value])
     except Exception as e:
      sumr.cell(row=52,column=9,value=0)
     try:
       sumr.cell(row=52,column=10,value=LifeInsurance[sumr.cell(row=13,column=10).value])
     except Exception as e:
      sumr.cell(row=52,column=10,value=0)
     try:
       sumr.cell(row=52,column=11,value=LifeInsurance[sumr.cell(row=13,column=11).value])
     except Exception as e:
      sumr.cell(row=52,column=11,value=0)
     try:
       sumr.cell(row=52,column=12,value=LifeInsurance[sumr.cell(row=13,column=12).value])
     except Exception as e:
      sumr.cell(row=52,column=12,value=0)
     try:
       sumr.cell(row=52,column=13,value=LifeInsurance[sumr.cell(row=13,column=13).value])
     except Exception as e:
      sumr.cell(row=52,column=13,value=0)
     try:
       sumr.cell(row=52,column=14,value=LifeInsurance[sumr.cell(row=13,column=14).value])
     except Exception as e:
      sumr.cell(row=52,column=14,value=0)

      try:
          sumr.cell(row=56, column=3, value=pt[sumr.cell(row=13, column=3).value])
      except Exception as e:
          sumr.cell(row=56, column=3, value=0)
      try:
          sumr.cell(row=56, column=4, value=pt[sumr.cell(row=13, column=4).value])
      except Exception as e:
          sumr.cell(row=56, column=4, value=0)
      try:
          sumr.cell(row=56, column=5, value=pt[sumr.cell(row=13, column=5).value])
      except Exception as e:
          sumr.cell(row=56, column=5, value=0)
      try:
          sumr.cell(row=56, column=6, value=pt[sumr.cell(row=13, column=6).value])
      except Exception as e:
          sumr.cell(row=56, column=6, value=0)
      try:
          sumr.cell(row=56, column=7, value=pt[sumr.cell(row=13, column=7).value])
      except Exception as e:
          sumr.cell(row=56, column=7, value=0)
      try:
          sumr.cell(row=56, column=8, value=pt[sumr.cell(row=13, column=8).value])
      except Exception as e:
          sumr.cell(row=56, column=8, value=0)
      try:
          sumr.cell(row=56, column=9, value=pt[sumr.cell(row=13, column=9).value])
      except Exception as e:
          sumr.cell(row=56, column=9, value=0)
      try:
          sumr.cell(row=56, column=10, value=pt[sumr.cell(row=13, column=10).value])
      except Exception as e:
          sumr.cell(row=56, column=10, value=0)
      try:
          sumr.cell(row=56, column=11, value=pt[sumr.cell(row=13, column=11).value])
      except Exception as e:
          sumr.cell(row=56, column=11, value=0)
      try:
          sumr.cell(row=56, column=12, value=pt[sumr.cell(row=13, column=12).value])
      except Exception as e:
          sumr.cell(row=56, column=12, value=0)
      try:
          sumr.cell(row=56, column=13, value=pt[sumr.cell(row=13, column=13).value])
      except Exception as e:
          sumr.cell(row=56, column=13, value=0)
      try:
          sumr.cell(row=56, column=14, value=pt[sumr.cell(row=13, column=14).value])
      except Exception as e:
          sumr.cell(row=56, column=14, value=0)

      try:
          sumr.cell(row=78, column=3, value=pt[sumr.cell(row=13, column=3).value])
      except Exception as e:
          sumr.cell(row=78, column=3, value=0)
      try:
          sumr.cell(row=78, column=4, value=pt[sumr.cell(row=13, column=4).value])
      except Exception as e:
          sumr.cell(row=78, column=4, value=0)
      try:
          sumr.cell(row=78, column=5, value=pt[sumr.cell(row=13, column=5).value])
      except Exception as e:
          sumr.cell(row=78, column=5, value=0)
      try:
          sumr.cell(row=78, column=6, value=pt[sumr.cell(row=13, column=6).value])
      except Exception as e:
          sumr.cell(row=78, column=6, value=0)
      try:
          sumr.cell(row=78, column=7, value=pt[sumr.cell(row=13, column=7).value])
      except Exception as e:
          sumr.cell(row=78, column=7, value=0)
      try:
          sumr.cell(row=78, column=8, value=pt[sumr.cell(row=13, column=8).value])
      except Exception as e:
          sumr.cell(row=78, column=8, value=0)
      try:
          sumr.cell(row=78, column=9, value=pt[sumr.cell(row=13, column=9).value])
      except Exception as e:
          sumr.cell(row=78, column=9, value=0)
      try:
          sumr.cell(row=78, column=10, value=pt[sumr.cell(row=13, column=10).value])
      except Exception as e:
          sumr.cell(row=78, column=10, value=0)
      try:
          sumr.cell(row=78, column=11, value=pt[sumr.cell(row=13, column=11).value])
      except Exception as e:
          sumr.cell(row=78, column=11, value=0)
      try:
          sumr.cell(row=78, column=12, value=pt[sumr.cell(row=13, column=12).value])
      except Exception as e:
          sumr.cell(row=78, column=12, value=0)
      try:
          sumr.cell(row=78, column=13, value=pt[sumr.cell(row=13, column=13).value])
      except Exception as e:
          sumr.cell(row=78, column=13, value=0)
      try:
          sumr.cell(row=78, column=14, value=pt[sumr.cell(row=13, column=14).value])
      except Exception as e:
          sumr.cell(row=78, column=14, value=0)



     try:
       sumr.cell(row=53,column=3,value=FoodExpence[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=53,column=3,value=0)
     try:
       sumr.cell(row=53,column=4,value=FoodExpence[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=53,column=4,value=0)
     try:
       sumr.cell(row=53,column=5,value=FoodExpence[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=53,column=5,value=0)
     try:
       sumr.cell(row=53,column=6,value=FoodExpence[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=53,column=6,value=0)
     try:
       sumr.cell(row=53,column=7,value=FoodExpence[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=53,column=7,value=0)
     try:
       sumr.cell(row=53,column=8,value=FoodExpence[sumr.cell(row=13,column=8).value])
     except Exception as e:
      sumr.cell(row=53,column=8,value=0)
     try:
       sumr.cell(row=53,column=9,value=FoodExpence[sumr.cell(row=13,column=9).value])
     except Exception as e:
      sumr.cell(row=53,column=9,value=0)
     try:
       sumr.cell(row=53,column=10,value=FoodExpence[sumr.cell(row=13,column=10).value])
     except Exception as e:
      sumr.cell(row=53,column=10,value=0)
     try:
       sumr.cell(row=53,column=11,value=FoodExpence[sumr.cell(row=13,column=11).value])
     except Exception as e:
      sumr.cell(row=53,column=11,value=0)
     try:
       sumr.cell(row=53,column=12,value=FoodExpence[sumr.cell(row=13,column=12).value])
     except Exception as e:
      sumr.cell(row=53,column=12,value=0)
     try:
       sumr.cell(row=53,column=13,value=FoodExpence[sumr.cell(row=13,column=13).value])
     except Exception as e:
      sumr.cell(row=53,column=13,value=0)
     try:
       sumr.cell(row=53,column=14,value=FoodExpence[sumr.cell(row=13,column=14).value])
     except Exception as e:
      sumr.cell(row=53,column=14,value=0)




     try:
       sumr.cell(row=54,column=3,value=CreditCharges[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=54,column=3,value=0)
     try:
       sumr.cell(row=54,column=4,value=CreditCharges[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=54,column=4,value=0)
     try:
       sumr.cell(row=54,column=5,value=CreditCharges[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=54,column=5,value=0)
     try:
       sumr.cell(row=54,column=6,value=CreditCharges[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=54,column=6,value=0)
     try:
       sumr.cell(row=54,column=7,value=CreditCharges[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=54,column=7,value=0)
     try:
       sumr.cell(row=54,column=8,value=CreditCharges[sumr.cell(row=13,column=8).value])
     except Exception as e:
      sumr.cell(row=54,column=8,value=0)
     try:
       sumr.cell(row=54,column=9,value=CreditCharges[sumr.cell(row=13,column=9).value])
     except Exception as e:
      sumr.cell(row=54,column=9,value=0)
     try:
       sumr.cell(row=54,column=10,value=CreditCharges[sumr.cell(row=13,column=10).value])
     except Exception as e:
      sumr.cell(row=54,column=10,value=0)
     try:
       sumr.cell(row=54,column=11,value=CreditCharges[sumr.cell(row=13,column=11).value])
     except Exception as e:
      sumr.cell(row=54,column=11,value=0)
     try:
       sumr.cell(row=54,column=12,value=CreditCharges[sumr.cell(row=13,column=12).value])
     except Exception as e:
      sumr.cell(row=54,column=12,value=0)
     try:
       sumr.cell(row=54,column=13,value=CreditCharges[sumr.cell(row=13,column=13).value])
     except Exception as e:
      sumr.cell(row=54,column=13,value=0)
     try:
       sumr.cell(row=54,column=14,value=CreditCharges[sumr.cell(row=13,column=14).value])
     except Exception as e:
      sumr.cell(row=54,column=14,value=0)

     try:
       sumr.cell(row=55,column=3,value=OnlineShopping[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=55,column=3,value=0)
     try:
       sumr.cell(row=55,column=4,value=OnlineShopping[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=55,column=4,value=0)
     try:
       sumr.cell(row=55,column=5,value=OnlineShopping[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=55,column=5,value=0)
     try:
       sumr.cell(row=55,column=6,value=OnlineShopping[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=55,column=6,value=0)
     try:
       sumr.cell(row=55,column=7,value=OnlineShopping[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=55,column=7,value=0)
     try:
       sumr.cell(row=55,column=8,value=OnlineShopping[sumr.cell(row=13,column=8).value])
     except Exception as e:
      sumr.cell(row=55,column=8,value=0)
     try:
       sumr.cell(row=55,column=9,value=OnlineShopping[sumr.cell(row=13,column=9).value])
     except Exception as e:
      sumr.cell(row=55,column=9,value=0)
     try:
       sumr.cell(row=55,column=10,value=OnlineShopping[sumr.cell(row=13,column=10).value])
     except Exception as e:
      sumr.cell(row=55,column=10,value=0)
     try:
       sumr.cell(row=55,column=11,value=OnlineShopping[sumr.cell(row=13,column=11).value])
     except Exception as e:
      sumr.cell(row=55,column=11,value=0)
     try:
       sumr.cell(row=55,column=12,value=OnlineShopping[sumr.cell(row=13,column=12).value])
     except Exception as e:
      sumr.cell(row=55,column=12,value=0)
     try:
       sumr.cell(row=55,column=13,value=OnlineShopping[sumr.cell(row=13,column=13).value])
     except Exception as e:
      sumr.cell(row=55,column=13,value=0)
     try:
       sumr.cell(row=55,column=14,value=OnlineShopping[sumr.cell(row=13,column=14).value])
     except Exception as e:
      sumr.cell(row=55,column=14,value=0)


     try:
       sumr.cell(row=57,column=3,value=GasPayment[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=57,column=3,value=0)
     try:
       sumr.cell(row=57,column=4,value=GasPayment[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=57,column=4,value=0)
     try:
       sumr.cell(row=57,column=5,value=GasPayment[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=57,column=5,value=0)
     try:
       sumr.cell(row=57,column=6,value=GasPayment[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=57,column=6,value=0)
     try:
       sumr.cell(row=57,column=7,value=GasPayment[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=57,column=7,value=0)
     try:
       sumr.cell(row=57,column=8,value=GasPayment[sumr.cell(row=13,column=8).value])
     except Exception as e:
      sumr.cell(row=57,column=8,value=0)
     try:
       sumr.cell(row=57,column=9,value=GasPayment[sumr.cell(row=13,column=9).value])
     except Exception as e:
      sumr.cell(row=57,column=9,value=0)
     try:
       sumr.cell(row=57,column=10,value=GasPayment[sumr.cell(row=13,column=10).value])
     except Exception as e:
      sumr.cell(row=57,column=10,value=0)
     try:
       sumr.cell(row=57,column=11,value=GasPayment[sumr.cell(row=13,column=11).value])
     except Exception as e:
      sumr.cell(row=57,column=11,value=0)
     try:
       sumr.cell(row=57,column=12,value=GasPayment[sumr.cell(row=13,column=12).value])
     except Exception as e:
      sumr.cell(row=57,column=12,value=0)
     try:
       sumr.cell(row=57,column=13,value=GasPayment[sumr.cell(row=13,column=13).value])
     except Exception as e:
      sumr.cell(row=57,column=13,value=0)
     try:
       sumr.cell(row=57,column=14,value=GasPayment[sumr.cell(row=13,column=14).value])
     except Exception as e:
      sumr.cell(row=57,column=14,value=0)


     try:
       sumr.cell(row=58,column=3,value=GoldLoan[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=58,column=3,value=0)
     try:
       sumr.cell(row=58,column=4,value=GoldLoan[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=58,column=4,value=0)
     try:
       sumr.cell(row=58,column=5,value=GoldLoan[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=58,column=5,value=0)
     try:
       sumr.cell(row=58,column=6,value=GoldLoan[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=58,column=6,value=0)
     try:
       sumr.cell(row=58,column=7,value=GoldLoan[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=58,column=7,value=0)
     try:
       sumr.cell(row=58,column=8,value=GoldLoan[sumr.cell(row=13,column=8).value])
     except Exception as e:
      sumr.cell(row=58,column=8,value=0)
     try:
       sumr.cell(row=58,column=9,value=GoldLoan[sumr.cell(row=13,column=9).value])
     except Exception as e:
      sumr.cell(row=58,column=9,value=0)
     try:
       sumr.cell(row=58,column=10,value=GoldLoan[sumr.cell(row=13,column=10).value])
     except Exception as e:
      sumr.cell(row=58,column=10,value=0)
     try:
       sumr.cell(row=58,column=11,value=GoldLoan[sumr.cell(row=13,column=11).value])
     except Exception as e:
      sumr.cell(row=58,column=11,value=0)
     try:
       sumr.cell(row=58,column=12,value=GoldLoan[sumr.cell(row=13,column=12).value])
     except Exception as e:
      sumr.cell(row=58,column=12,value=0)
     try:
       sumr.cell(row=58,column=13,value=GoldLoan[sumr.cell(row=13,column=13).value])
     except Exception as e:
      sumr.cell(row=58,column=13,value=0)
     try:
       sumr.cell(row=58,column=14,value=GoldLoan[sumr.cell(row=13,column=14).value])
     except Exception as e:
      sumr.cell(row=58,column=14,value=0)


     try:
       sumr.cell(row=59,column=3,value=rent_dr[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=59,column=3,value=0)
     try:
       sumr.cell(row=59,column=4,value=rent_dr[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=59,column=4,value=0)
     try:
       sumr.cell(row=59,column=5,value=rent_dr[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=59,column=5,value=0)
     try:
       sumr.cell(row=59,column=6,value=rent_dr[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=59,column=6,value=0)
     try:
       sumr.cell(row=59,column=7,value=rent_dr[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=59,column=7,value=0)
     try:
       sumr.cell(row=59,column=8,value=rent_dr[sumr.cell(row=13,column=8).value])
     except Exception as e:
      sumr.cell(row=59,column=8,value=0)
     try:
       sumr.cell(row=59,column=9,value=rent_dr[sumr.cell(row=13,column=9).value])
     except Exception as e:
      sumr.cell(row=59,column=9,value=0)
     try:
       sumr.cell(row=59,column=10,value=rent_dr[sumr.cell(row=13,column=10).value])
     except Exception as e:
      sumr.cell(row=59,column=10,value=0)
     try:
       sumr.cell(row=59,column=11,value=rent_dr[sumr.cell(row=13,column=11).value])
     except Exception as e:
      sumr.cell(row=59,column=11,value=0)
     try:
       sumr.cell(row=59,column=12,value=rent_dr[sumr.cell(row=13,column=12).value])
     except Exception as e:
      sumr.cell(row=59,column=12,value=0)
     try:
       sumr.cell(row=59,column=13,value=rent_dr[sumr.cell(row=13,column=13).value])
     except Exception as e:
      sumr.cell(row=59,column=13,value=0)
     try:
       sumr.cell(row=59,column=14,value=rent_dr[sumr.cell(row=13,column=14).value])
     except Exception as e:
      sumr.cell(row=59,column=14,value=0)

     wb.save('Excel_Files/Dashboard/BankStatement.xlsx')

PersonalExpenses()

from pandas.core.dtypes.dtypes import dtypes
def Balance1():
    
#     row8=> Opening balance
    df5 = pd.read_excel('Excel_Files/Dashboard/BankStatement.xlsx',sheet_name='EOD Balance',index_col = 'Day')
#     print(df5)
#     print(df5.index[-1])
    df5.drop(index=df5.index[-1],axis=0,inplace=True) 
    
#     df5.fillna(inplace=True)
#     print(df5.index[-1])
    first_col=df5.iloc[:, 0:1:]
    first_col.dropna(inplace=True)
    

    
    sec_col=df5.iloc[:, 1:2:]
    sec_col.dropna(inplace=True)
    sec_2 = sec_col.tail(1).values
    sec_col.at[2] = sec_2
    

    thir_col=df5.iloc[:,2:3:]
    thir_col.dropna(inplace=True)
    third_3 = thir_col.tail(1).values
    thir_col.at[2] = third_3

    four_col=df5.iloc[:,3:4:]
    four_col.dropna(inplace=True)
    fou_4 = four_col.tail(1).values
    four_col.at[2] = fou_4

    five_col=df5.iloc[:,4:5:]
    five_col.dropna(inplace=True)
    fiv_5 = five_col.tail(1).values
    five_col.at[2] = fiv_5

    six_col=df5.iloc[:,5:6:]
    six_col.dropna(inplace=True)
    six_6 = six_col.tail(1).values
    six_col.at[2] = six_6

    seven_col=df5.iloc[:,6:7:]
    seven_col.dropna(inplace=True)
    sev_7 = seven_col.tail(1).values
    seven_col.at[2] = sev_7

    eight_col=df5.iloc[:,7:8:]
    eight_col.dropna(inplace=True)
    eig_8 = eight_col.tail(1).values
    eight_col.at[2] = eig_8

    nine_col=df5.iloc[:,8:9:]
    nine_col.dropna(inplace=True)
    nin_9 = nine_col.tail(1).values
    nine_col.at[2] = nin_9

    ten_col=df5.iloc[:,9:10:]
    ten_col.dropna(inplace=True)
    ten_10 = ten_col.tail(1).values
    ten_col.at[2] = ten_10

    eleven_col=df5.iloc[:,10:11:]
    eleven_col.dropna(inplace=True)
    ele_11 = eleven_col.tail(1).values
    eleven_col.at[2] = ele_11
    
    twelve_col=df5.iloc[:,11:12:]
    twelve_col.dropna(inplace=True)
    twe_12 = twelve_col.tail(1).values
    twelve_col.at[2] = twe_12

    v=first_col.tail(1).values
#     print(first_col_closing_Bal)
    print(v.dtype)
    
    first_col.at[2] = v

    wb = load_workbook('Excel_Files/Dashboard/BankStatement.xlsx')
    sumr= wb['summary']
    eod= wb['EOD Balance']

    month1=np.array_str(v)
    close1=month1[2:-2]
    try:
      sumr.cell(row=22,column=3,value=close1)
    except Exception as e:
       sumr.cell(row=22,column=3,value=0)

    month2=np.array_str(sec_2)
    close2=month2[2:-2]
    try:
      sumr.cell(row=22,column=4,value=close2)
    except Exception as e:
       sumr.cell(row=22,column=4,value=0)

    month3=np.array_str(third_3)
    close3=month3[2:-2]
    try:
      sumr.cell(row=22,column=5,value=close3)
    except Exception as e:
       sumr.cell(row=22,column=5,value=0)

    month4=np.array_str(fou_4)
    close4=month4[2:-2]
    try:
      sumr.cell(row=22,column=6,value=close4)
    except Exception as e:
       sumr.cell(row=22,column=6,value=0)

    month5=np.array_str(fiv_5)
    close5=month5[2:-2]
    try:
      sumr.cell(row=22,column=7,value=close5)
    except Exception as e:
       sumr.cell(row=22,column=7,value=0)

    month6=np.array_str(six_6)
    close6=month6[2:-2]
    try:
      sumr.cell(row=22,column=8,value=close6)
    except Exception as e:
       sumr.cell(row=22,column=8,value=0)

    month7=np.array_str(sev_7)
    close7=month7[2:-2]
    try:
      sumr.cell(row=22,column=9,value=close7)
    except Exception as e:
       sumr.cell(row=22,column=9,value=0)

    month8=np.array_str(eig_8)
    close8=month8[2:-2]
    try:
      sumr.cell(row=22,column=10,value=close8)
    except Exception as e:
       sumr.cell(row=22,column=10,value=0)

    month9=np.array_str(nin_9)
    close9=month9[2:-2]
    try:
      sumr.cell(row=22,column=11,value=close9)
    except Exception as e:
       sumr.cell(row=22,column=11,value=0)

    month10=np.array_str(ten_10)
    close10=month10[2:-2]
    try:
      sumr.cell(row=22,column=12,value=close10)
    except Exception as e:
       sumr.cell(row=22,column=12,value=0)

    month11=np.array_str(ele_11)
    close11=month11[2:-2]
    try:
      sumr.cell(row=22,column=13,value=close11)
    except Exception as e:
       sumr.cell(row=22,column=13,value=0)

    month12=np.array_str(twe_12)
    close12=month12[2:-2]
    try:
      sumr.cell(row=22,column=14,value=close12)
    except Exception as e:
       sumr.cell(row=22,column=14,value=0)
    

    wb.save('Excel_Files/Dashboard/BankStatement.xlsx')

Balance1()

def perExpenditure():


     wb = load_workbook('Excel_Files/Dashboard/BankStatement.xlsx')
     sumr= wb['summary']
     eod= wb['EOD Balance']

          # Bank charges
     df_charges =df1[df1["Particulars"].str.contains("Consolidated Charges for A/c",na=False)]
     Bank_charges = df_charges.groupby(df_charges['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
            # print(Bank_charges)
        
          # Emi
     new_emi = df1[df1['Particulars'].str.contains('ECS') | df1['Particulars'].str.contains('NACH-DR')]
     Emi_dr = new_emi.groupby(new_emi['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
            # print(new)
        
          # GST Paid
     df_gst = df1[df1["Particulars"].str.contains("GST",na=False)]
     gst = df_gst.groupby(df_gst['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
            # print(gst)

          # Utility bills
     new_utility = df1[df1['Particulars'].str.contains('NACH-DR- BSNL') | df1['Particulars'].str.contains('NACH-DR- Vodafone')| df1['Particulars'].str.contains('NACH-DR- Reliance')| df1['Particulars'].str.contains('NACH-DR- MTNL')| df1['Particulars'].str.contains('NACH-DR- Reliance JIO')| df1['Particulars'].str.contains('NACH-DR- Airtel')| df1['Particulars'].str.contains('NTPC')| df1['Particulars'].str.contains('Adani Group')| df1['Particulars'].str.contains('Tata Power')| df1['Particulars'].str.contains('JSW Energy')| df1['Particulars'].str.contains('Torrent Power')| df1['Particulars'].str.contains('Powergrid')| df1['Particulars'].str.contains('NHPC')| df1['Particulars'].str.contains('Reliance Power Ltd')| df1['Particulars'].str.contains('SJVN')| df1['Particulars'].str.contains('CESC')| df1['Particulars'].str.contains('Oil and Natural Gas Corporation')| df1['Particulars'].str.contains('Indian Oil Corporation Limited')| df1['Particulars'].str.contains('Bharat Petroleum')| df1['Particulars'].str.contains('Gas Authority of India')| df1['Particulars'].str.contains('Reliance Petroleum Ltd')| df1['Particulars'].str.contains('Hindustan Petroleum')| df1['Particulars'].str.contains('Oil India')| df1['Particulars'].str.contains('Cairn India')| df1['Particulars'].str.contains('TATA Petrodyne Ltd')| df1['Particulars'].str.contains('Essar Oil')]
     Utility = new_utility.groupby(new_utility['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
            # print(Utility)

          # General Insurance
     General_insurance_df=df1[df1['Particulars'].str.contains('National Insurance Co. Ltd.') |df1['Particulars'].str.contains('Go Digit General Insurance Ltd.') |df1['Particulars'].str.contains('Bajaj Allianz General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Cholamandalam MS General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Bharti AXA General Insurance Co. Ltd.') |df1['Particulars'].str.contains('HDFC ERGO General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Future Generali India Insurance Co. Ltd.') |df1['Particulars'].str.contains('The New India Assurance Co. Ltd.') |df1['Particulars'].str.contains('Iffco Tokio General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Reliance General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Royal Sundaram General Insurance Co. Ltd.') |df1['Particulars'].str.contains('The Oriental Insurance Co. Ltd.') |df1['Particulars'].str.contains('Tata AIG General Insurance Co. Ltd.') |df1['Particulars'].str.contains('SBI General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Acko General Insurance Ltd.') |df1['Particulars'].str.contains('Navi General Insurance Ltd.') |df1['Particulars'].str.contains('Edelweiss General Insurance Co. Ltd.') |df1['Particulars'].str.contains('ICICI Lombard General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Kotak Mahindra General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Liberty General Insurance Ltd.') |df1['Particulars'].str.contains('Magma HDI General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Raheja QBE General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Shriram General Insurance Co. Ltd.') |df1['Particulars'].str.contains('United India Insurance Co. Ltd.') |df1['Particulars'].str.contains('Universal Sompo General Insurance Co. Ltd.') |df1['Particulars'].str.contains('Agriculture Insurance Company of India Ltd.') |df1['Particulars'].str.contains('Aditya Birla Health Insurance Co. Ltd.') |df1['Particulars'].str.contains('Manipal Cigna Health Insurance Company Limited') |df1['Particulars'].str.contains('ECGC Ltd.') |df1['Particulars'].str.contains('Max Bupa Health Insurance Co. Ltd') |df1['Particulars'].str.contains('Care Health Insurance Ltd') |df1['Particulars'].str.contains('Star Health & Allied Insurance Co. Ltd.')]
     GeneralInsurance = General_insurance_df.groupby(General_insurance_df['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
        # print(GeneralInsurance) 

          # Life Insurance
     Life_insuranc_df= df1[df1['Particulars'].str.contains('Life Insurance Corporation of India') |df1['Particulars'].str.contains('HDFC Standard Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Max Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('ICICI Prudential Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Kotak Mahindra Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Aditya Birla Sun Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('TATA AIA Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('SBI Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Exide Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Bajaj Allianz Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('PNB MetLife India Insurance Co. Ltd.') |df1['Particulars'].str.contains('Reliance Nippon Life Insurance Company') |df1['Particulars'].str.contains('Aviva Life Insurance Company India Ltd.') |df1['Particulars'].str.contains('Sahara India Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Shriram Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Bharti AXA Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Future Generali India Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('IDBI Federal Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Canara HSBC Oriental Bank of Commerce Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Aegon Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Pramerica Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Star Union Dai-Ichi Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('IndiaFirst Life Insurance Co. Ltd.') |df1['Particulars'].str.contains('Edelweiss Tokio Life Insurance Co. Ltd.')]
     LifeInsurance = Life_insuranc_df.groupby(Life_insuranc_df['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
            # print(LifeInsurance)

          # Food Expenses
     Food_Expenxe_df = df1[df1['Particulars'].str.contains('Zomato')|df1['Particulars'].str.contains('Swiggy')]
     FoodExpence = Food_Expenxe_df.groupby(Food_Expenxe_df['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
            #print(FoodExpence)

          # Online shopping
     online_shopping_df = df1[df1['Particulars'].str.contains('AMAZON -Dr')]
     OnlineShopping = online_shopping_df.groupby(online_shopping_df['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
            #print(OnlineShopping)

          # Gas Payment
     gas_payments_df =df1[df1['Particulars'].str.contains('Oil and Natural Gas Corporation')| df1['Particulars'].str.contains('Indian Oil Corporation Limited')| df1['Particulars'].str.contains('Bharat Petroleum')| df1['Particulars'].str.contains('Gas Authority of India')| df1['Particulars'].str.contains('Reliance Petroleum Ltd')| df1['Particulars'].str.contains('Hindustan Petroleum')| df1['Particulars'].str.contains('Oil India')| df1['Particulars'].str.contains('Cairn India')| df1['Particulars'].str.contains('TATA Petrodyne Ltd')| df1['Particulars'].str.contains('Essar Oil')]
     GasPayment = gas_payments_df.groupby(gas_payments_df['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
            #print(GasPayment)

          # goal loan
     Gold_loan_df = df1[df1['Particulars'].str.contains('MUTHOOTFINCORP')|df1['Particulars'].str.contains('RUPEEKCAPITAL')]
     GoldLoan=Gold_loan_df.groupby(Gold_loan_df['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
            # print(GoldLoan)

    # Rent Paid
     rent_df = df1[df1['Particulars'].str.contains('Rent')]
     rent_dr = rent_df.groupby(rent_df['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
    #print(rent_dr)

    #   Salary Paid
     df_salary = df1[df1["Particulars"].str.contains("salary",na=False)]
     salary=df_salary.groupby(df_salary['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()

    # Total Income Tax Paid)
     TaxPaid = df1[df1["Particulars"].str.contains("Tax Paid",na=False)]
     tax_p=TaxPaid.groupby(TaxPaid['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()

    #  medical expence
     medical_df = df1[df1["Particulars"].str.contains("Pharma")|df1["Particulars"].str.contains("hospital")|df1["Particulars"].str.contains("clinic")|df1["Particulars"].str.contains("dentist")|df1["Particulars"].str.contains("PHARMA")]
     medical_charges = medical_df.groupby(medical_df['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
     try:
        sumr.cell(row=64,column=3,value=Bank_charges[sumr.cell(row=13,column=3).value])
     except Exception as e:
       sumr.cell(row=64,column=3,value=0)
     try:
        sumr.cell(row=64,column=4,value=Bank_charges[sumr.cell(row=13,column=4).value])
     except Exception as e:
       sumr.cell(row=64,column=4,value=0)
     try:
        sumr.cell(row=64,column=5,value=Bank_charges[sumr.cell(row=13,column=5).value])
     except Exception as e:
       sumr.cell(row=64,column=5,value=0)
     try:
        sumr.cell(row=64,column=6,value=Bank_charges[sumr.cell(row=13,column=6).value])
     except Exception as e:
       sumr.cell(row=64,column=6,value=0)
     try:
        sumr.cell(row=64,column=7,value=Bank_charges[sumr.cell(row=13,column=7).value])
     except Exception as e:
       sumr.cell(row=64,column=7,value=0)
     try:
        sumr.cell(row=64,column=8,value=Bank_charges[sumr.cell(row=13,column=8).value])
     except Exception as e:
       sumr.cell(row=64,column=8,value=0)
     try:
        sumr.cell(row=64,column=9,value=Bank_charges[sumr.cell(row=13,column=9).value])
     except Exception as e:
       sumr.cell(row=64,column=9,value=0)
     try:
        sumr.cell(row=64,column=10,value=Bank_charges[sumr.cell(row=13,column=10).value])
     except Exception as e:
       sumr.cell(row=64,column=10,value=0)
     try:
        sumr.cell(row=64,column=11,value=Bank_charges[sumr.cell(row=13,column=11).value])
     except Exception as e:
       sumr.cell(row=64,column=11,value=0)
     try:
        sumr.cell(row=64,column=12,value=Bank_charges[sumr.cell(row=13,column=12).value])
     except Exception as e:
       sumr.cell(row=64,column=12,value=0)
     try:
        sumr.cell(row=64,column=13,value=Bank_charges[sumr.cell(row=13,column=13).value])
     except Exception as e:
       sumr.cell(row=64,column=13,value=0)
     try:
        sumr.cell(row=64,column=14,value=Bank_charges[sumr.cell(row=13,column=14).value])
     except Exception as e:
       sumr.cell(row=64,column=14,value=0)

     try:
       sumr.cell(row=65,column=3,value=salary[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=65,column=3,value=0)
     try:
       sumr.cell(row=65,column=4,value=salary[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=65,column=4,value=0)
     try:
       sumr.cell(row=65,column=5,value=salary[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=65,column=5,value=0)
     try:
       sumr.cell(row=65,column=6,value=salary[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=65,column=6,value=0)
     try:
       sumr.cell(row=65,column=7,value=salary[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=65,column=7,value=0)
     try:
       sumr.cell(row=65,column=8,value=salary[sumr.cell(row=13,column=8).value])
     except Exception as e:
       sumr.cell(row=65,column=8,value=0)
     try:
       sumr.cell(row=65,column=9,value=salary[sumr.cell(row=13,column=9).value])
     except Exception as e:
       sumr.cell(row=65,column=9,value=0)
     try:
       sumr.cell(row=65,column=10,value=salary[sumr.cell(row=13,column=10).value])
     except Exception as e:
       sumr.cell(row=65,column=10,value=0)
     try:
       sumr.cell(row=65,column=11,value=salary[sumr.cell(row=13,column=11).value])
     except Exception as e:
       sumr.cell(row=65,column=11,value=0)
     try:
       sumr.cell(row=65,column=12,value=salary[sumr.cell(row=13,column=12).value])
     except Exception as e:
       sumr.cell(row=65,column=12,value=0)
     try:
       sumr.cell(row=65,column=13,value=salary[sumr.cell(row=13,column=13).value])
     except Exception as e:
       sumr.cell(row=65,column=13,value=0)
     try:
       sumr.cell(row=65,column=14,value=salary[sumr.cell(row=13,column=14).value])
     except Exception as e:
       sumr.cell(row=65,column=14,value=0)




     try:
       sumr.cell(row=66,column=3,value=Bank_charges[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=66,column=3,value=0)
     try:
       sumr.cell(row=66,column=4,value=Bank_charges[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=66,column=4,value=0)
     try:
       sumr.cell(row=66,column=5,value=Bank_charges[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=66,column=5,value=0)
     try:
       sumr.cell(row=66,column=6,value=Bank_charges[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=66,column=6,value=0)
     try:
       sumr.cell(row=66,column=7,value=Bank_charges[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=66,column=7,value=0)
     try:
       sumr.cell(row=66,column=8,value=Bank_charges[sumr.cell(row=13,column=8).value])
     except Exception as e:
       sumr.cell(row=66,column=8,value=0)
     try:
       sumr.cell(row=66,column=9,value=Bank_charges[sumr.cell(row=13,column=9).value])
     except Exception as e:
       sumr.cell(row=66,column=9,value=0)
     try:
       sumr.cell(row=66,column=10,value=Bank_charges[sumr.cell(row=13,column=10).value])
     except Exception as e:
       sumr.cell(row=66,column=10,value=0)
     try:
       sumr.cell(row=66,column=11,value=Bank_charges[sumr.cell(row=13,column=11).value])
     except Exception as e:
       sumr.cell(row=66,column=11,value=0)
     try:
       sumr.cell(row=66,column=12,value=Bank_charges[sumr.cell(row=13,column=12).value])
     except Exception as e:
       sumr.cell(row=66,column=12,value=0)
     try:
       sumr.cell(row=66,column=13,value=Bank_charges[sumr.cell(row=13,column=13).value])
     except Exception as e:
       sumr.cell(row=66,column=13,value=0)
     try:
       sumr.cell(row=66,column=14,value=Bank_charges[sumr.cell(row=13,column=14).value])
     except Exception as e:
       sumr.cell(row=66,column=14,value=0)



     try:
       sumr.cell(row=67,column=3,value=Emi_dr[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=67,column=3,value=0)
     try:
       sumr.cell(row=67,column=4,value=Emi_dr[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=67,column=4,value=0)
     try:
       sumr.cell(row=67,column=5,value=Emi_dr[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=67,column=5,value=0)
     try:
       sumr.cell(row=67,column=6,value=Emi_dr[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=67,column=6,value=0)
     try:
       sumr.cell(row=67,column=7,value=Emi_dr[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=67,column=7,value=0)
     try:
       sumr.cell(row=67,column=8,value=Emi_dr[sumr.cell(row=13,column=8).value])
     except Exception as e:
       sumr.cell(row=67,column=8,value=0)
     try:
       sumr.cell(row=67,column=9,value=Emi_dr[sumr.cell(row=13,column=9).value])
     except Exception as e:
       sumr.cell(row=67,column=9,value=0)
     try:
       sumr.cell(row=67,column=10,value=Emi_dr[sumr.cell(row=13,column=10).value])
     except Exception as e:
       sumr.cell(row=67,column=10,value=0)
     try:
       sumr.cell(row=67,column=11,value=Emi_dr[sumr.cell(row=13,column=11).value])
     except Exception as e:
       sumr.cell(row=67,column=11,value=0)
     try:
       sumr.cell(row=67,column=12,value=Emi_dr[sumr.cell(row=13,column=12).value])
     except Exception as e:
       sumr.cell(row=67,column=12,value=0)
     try:
       sumr.cell(row=67,column=13,value=Emi_dr[sumr.cell(row=13,column=13).value])
     except Exception as e:
       sumr.cell(row=67,column=13,value=0)
     try:
       sumr.cell(row=67,column=14,value=Emi_dr[sumr.cell(row=13,column=14).value])
     except Exception as e:
       sumr.cell(row=67,column=14,value=0)



     try:
       sumr.cell(row=69,column=3,value=gst[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=69,column=3,value=0)
     try:
       sumr.cell(row=69,column=4,value=gst[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=69,column=4,value=0)
     try:
       sumr.cell(row=69,column=5,value=gst[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=69,column=5,value=0)
     try:
       sumr.cell(row=69,column=6,value=gst[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=69,column=6,value=0)
     try:
       sumr.cell(row=69,column=7,value=gst[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=69,column=7,value=0)
     try:
       sumr.cell(row=69,column=8,value=gst[sumr.cell(row=13,column=8).value])
     except Exception as e:
       sumr.cell(row=69,column=8,value=0)
     try:
       sumr.cell(row=69,column=9,value=gst[sumr.cell(row=13,column=9).value])
     except Exception as e:
       sumr.cell(row=69,column=9,value=0)
     try:
       sumr.cell(row=69,column=10,value=gst[sumr.cell(row=13,column=10).value])
     except Exception as e:
       sumr.cell(row=69,column=10,value=0)
     try:
       sumr.cell(row=69,column=11,value=gst[sumr.cell(row=13,column=11).value])
     except Exception as e:
       sumr.cell(row=69,column=11,value=0)
     try:
       sumr.cell(row=69,column=12,value=gst[sumr.cell(row=13,column=12).value])
     except Exception as e:
       sumr.cell(row=69,column=12,value=0)
     try:
       sumr.cell(row=69,column=13,value=gst[sumr.cell(row=13,column=13).value])
     except Exception as e:
       sumr.cell(row=69,column=13,value=0)
     try:
       sumr.cell(row=69,column=14,value=gst[sumr.cell(row=13,column=14).value])
     except Exception as e:
       sumr.cell(row=69,column=14,value=0)

     try:
         sumr.cell(row=70, column=3, value=tax_p[sumr.cell(row=13, column=3).value])
     except Exception as e:
         sumr.cell(row=70, column=3, value=0)
     try:
         sumr.cell(row=70, column=4, value=tax_p[sumr.cell(row=13, column=4).value])
     except Exception as e:
         sumr.cell(row=70, column=4, value=0)
     try:
         sumr.cell(row=70, column=5, value=tax_p[sumr.cell(row=13, column=5).value])
     except Exception as e:
         sumr.cell(row=70, column=5, value=0)
     try:
         sumr.cell(row=70, column=6, value=tax_p[sumr.cell(row=13, column=6).value])
     except Exception as e:
         sumr.cell(row=70, column=6, value=0)
     try:
         sumr.cell(row=70, column=7, value=tax_p[sumr.cell(row=13, column=7).value])
     except Exception as e:
         sumr.cell(row=70, column=7, value=0)
     try:
         sumr.cell(row=70, column=8, value=tax_p[sumr.cell(row=13, column=8).value])
     except Exception as e:
         sumr.cell(row=70, column=8, value=0)
     try:
         sumr.cell(row=70, column=9, value=tax_p[sumr.cell(row=13, column=9).value])
     except Exception as e:
         sumr.cell(row=70, column=9, value=0)
     try:
         sumr.cell(row=70, column=10, value=tax_p[sumr.cell(row=13, column=10).value])
     except Exception as e:
         sumr.cell(row=70, column=10, value=0)
     try:
         sumr.cell(row=70, column=11, value=tax_p[sumr.cell(row=13, column=11).value])
     except Exception as e:
         sumr.cell(row=70, column=11, value=0)
     try:
         sumr.cell(row=70, column=12, value=tax_p[sumr.cell(row=13, column=12).value])
     except Exception as e:
         sumr.cell(row=70, column=12, value=0)
     try:
         sumr.cell(row=70, column=13, value=tax_p[sumr.cell(row=13, column=13).value])
     except Exception as e:
         sumr.cell(row=70, column=13, value=0)
     try:
         sumr.cell(row=70, column=14, value=tax_p[sumr.cell(row=13, column=14).value])
     except Exception as e:
         sumr.cell(row=70, column=14, value=0)



     try:
       sumr.cell(row=71,column=3,value=Utility[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=71,column=3,value=0)
     try:
       sumr.cell(row=71,column=4,value=Utility[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=71,column=4,value=0)
     try:
       sumr.cell(row=71,column=5,value=Utility[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=71,column=5,value=0)
     try:
       sumr.cell(row=71,column=6,value=Utility[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=71,column=6,value=0)
     try:
       sumr.cell(row=71,column=7,value=Utility[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=71,column=7,value=0)
     try:
       sumr.cell(row=71,column=8,value=Utility[sumr.cell(row=13,column=8).value])
     except Exception as e:
       sumr.cell(row=71,column=8,value=0)
     try:
       sumr.cell(row=71,column=9,value=Utility[sumr.cell(row=13,column=9).value])
     except Exception as e:
       sumr.cell(row=71,column=9,value=0)
     try:
       sumr.cell(row=71,column=10,value=Utility[sumr.cell(row=13,column=10).value])
     except Exception as e:
       sumr.cell(row=71,column=10,value=0)
     try:
       sumr.cell(row=71,column=11,value=Utility[sumr.cell(row=13,column=11).value])
     except Exception as e:
       sumr.cell(row=71,column=11,value=0)
     try:
       sumr.cell(row=71,column=12,value=Utility[sumr.cell(row=13,column=12).value])
     except Exception as e:
       sumr.cell(row=71,column=12,value=0)
     try:
       sumr.cell(row=71,column=13,value=Utility[sumr.cell(row=13,column=13).value])
     except Exception as e:
       sumr.cell(row=71,column=13,value=0)
     try:
       sumr.cell(row=71,column=14,value=Utility[sumr.cell(row=13,column=14).value])
     except Exception as e:
       sumr.cell(row=71,column=14,value=0)


     try:
       sumr.cell(row=45,column=3,value=Utility[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=45,column=3,value=0)
     try:
       sumr.cell(row=45,column=4,value=Utility[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=45,column=4,value=0)
     try:
       sumr.cell(row=45,column=5,value=Utility[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=45,column=5,value=0)
     try:
       sumr.cell(row=45,column=6,value=Utility[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=45,column=6,value=0)
     try:
       sumr.cell(row=45,column=7,value=Utility[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=45,column=7,value=0)
     try:
       sumr.cell(row=45,column=8,value=Utility[sumr.cell(row=13,column=8).value])
     except Exception as e:
       sumr.cell(row=45,column=8,value=0)
     try:
       sumr.cell(row=45,column=9,value=Utility[sumr.cell(row=13,column=9).value])
     except Exception as e:
       sumr.cell(row=45,column=9,value=0)
     try:
       sumr.cell(row=45,column=10,value=Utility[sumr.cell(row=13,column=10).value])
     except Exception as e:
       sumr.cell(row=45,column=10,value=0)
     try:
       sumr.cell(row=45,column=11,value=Utility[sumr.cell(row=13,column=11).value])
     except Exception as e:
       sumr.cell(row=45,column=11,value=0)
     try:
       sumr.cell(row=45,column=12,value=Utility[sumr.cell(row=13,column=12).value])
     except Exception as e:
       sumr.cell(row=45,column=12,value=0)
     try:
       sumr.cell(row=45,column=13,value=Utility[sumr.cell(row=13,column=13).value])
     except Exception as e:
       sumr.cell(row=45,column=13,value=0)
     try:
       sumr.cell(row=45,column=14,value=Utility[sumr.cell(row=13,column=14).value])
     except Exception as e:
       sumr.cell(row=45,column=14,value=0)



     try:
       sumr.cell(row=73,column=3,value=GeneralInsurance[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=73,column=3,value=0)
     try:
       sumr.cell(row=73,column=4,value=GeneralInsurance[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=73,column=4,value=0)
     try:
       sumr.cell(row=73,column=5,value=GeneralInsurance[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=73,column=5,value=0)
     try:
       sumr.cell(row=73,column=6,value=GeneralInsurance[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=73,column=6,value=0)
     try:
       sumr.cell(row=73,column=7,value=GeneralInsurance[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=73,column=7,value=0)
     try:
       sumr.cell(row=73,column=8,value=GeneralInsurance[sumr.cell(row=13,column=8).value])
     except Exception as e:
       sumr.cell(row=73,column=8,value=0)
     try:
       sumr.cell(row=73,column=9,value=GeneralInsurance[sumr.cell(row=13,column=9).value])
     except Exception as e:
       sumr.cell(row=73,column=9,value=0)
     try:
       sumr.cell(row=73,column=10,value=GeneralInsurance[sumr.cell(row=13,column=10).value])
     except Exception as e:
       sumr.cell(row=73,column=10,value=0)
     try:
       sumr.cell(row=73,column=11,value=GeneralInsurance[sumr.cell(row=13,column=11).value])
     except Exception as e:
       sumr.cell(row=73,column=11,value=0)
     try:
       sumr.cell(row=73,column=12,value=GeneralInsurance[sumr.cell(row=13,column=12).value])
     except Exception as e:
       sumr.cell(row=73,column=12,value=0)
     try:
       sumr.cell(row=73,column=13,value=GeneralInsurance[sumr.cell(row=13,column=13).value])
     except Exception as e:
       sumr.cell(row=73,column=13,value=0)
     try:
       sumr.cell(row=73,column=14,value=GeneralInsurance[sumr.cell(row=13,column=14).value])
     except Exception as e:
       sumr.cell(row=73,column=14,value=0)


     try:
       sumr.cell(row=74,column=3,value=LifeInsurance[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=74,column=3,value=0)
     try:
       sumr.cell(row=74,column=4,value=LifeInsurance[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=74,column=4,value=0)
     try:
       sumr.cell(row=74,column=5,value=LifeInsurance[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=74,column=5,value=0)
     try:
       sumr.cell(row=74,column=6,value=LifeInsurance[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=74,column=6,value=0)
     try:
       sumr.cell(row=74,column=7,value=LifeInsurance[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=74,column=7,value=0)
     try:
       sumr.cell(row=74,column=8,value=LifeInsurance[sumr.cell(row=13,column=8).value])
     except Exception as e:
       sumr.cell(row=74,column=8,value=0)
     try:
       sumr.cell(row=74,column=9,value=LifeInsurance[sumr.cell(row=13,column=9).value])
     except Exception as e:
       sumr.cell(row=74,column=9,value=0)
     try:
       sumr.cell(row=74,column=10,value=LifeInsurance[sumr.cell(row=13,column=10).value])
     except Exception as e:
       sumr.cell(row=74,column=10,value=0)
     try:
       sumr.cell(row=74,column=11,value=LifeInsurance[sumr.cell(row=13,column=11).value])
     except Exception as e:
       sumr.cell(row=74,column=11,value=0)
     try:
       sumr.cell(row=74,column=12,value=LifeInsurance[sumr.cell(row=13,column=12).value])
     except Exception as e:
       sumr.cell(row=74,column=12,value=0)
     try:
       sumr.cell(row=74,column=13,value=LifeInsurance[sumr.cell(row=13,column=13).value])
     except Exception as e:
       sumr.cell(row=74,column=13,value=0)
     try:
       sumr.cell(row=74,column=14,value=LifeInsurance[sumr.cell(row=13,column=14).value])
     except Exception as e:
       sumr.cell(row=74,column=14,value=0)



     try:
       sumr.cell(row=75,column=3,value=FoodExpence[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=75,column=3,value=0)
     try:
       sumr.cell(row=75,column=4,value=FoodExpence[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=75,column=4,value=0)
     try:
       sumr.cell(row=75,column=5,value=FoodExpence[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=75,column=5,value=0)
     try:
       sumr.cell(row=75,column=6,value=FoodExpence[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=75,column=6,value=0)
     try:
       sumr.cell(row=75,column=7,value=FoodExpence[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=75,column=7,value=0)
     try:
       sumr.cell(row=75,column=8,value=FoodExpence[sumr.cell(row=13,column=8).value])
     except Exception as e:
       sumr.cell(row=75,column=8,value=0)
     try:
       sumr.cell(row=75,column=9,value=FoodExpence[sumr.cell(row=13,column=9).value])
     except Exception as e:
       sumr.cell(row=75,column=9,value=0)
     try:
       sumr.cell(row=75,column=10,value=FoodExpence[sumr.cell(row=13,column=10).value])
     except Exception as e:
       sumr.cell(row=75,column=10,value=0)
     try:
       sumr.cell(row=75,column=11,value=FoodExpence[sumr.cell(row=13,column=11).value])
     except Exception as e:
       sumr.cell(row=75,column=11,value=0)
     try:
       sumr.cell(row=75,column=12,value=FoodExpence[sumr.cell(row=13,column=12).value])
     except Exception as e:
       sumr.cell(row=75,column=12,value=0)
     try:
       sumr.cell(row=75,column=13,value=FoodExpence[sumr.cell(row=13,column=13).value])
     except Exception as e:
       sumr.cell(row=75,column=13,value=0)
     try:
       sumr.cell(row=75,column=14,value=FoodExpence[sumr.cell(row=13,column=14).value])
     except Exception as e:
       sumr.cell(row=75,column=14,value=0)



     try:
       sumr.cell(row=77,column=3,value=OnlineShopping[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=77,column=3,value=0)
     try:
       sumr.cell(row=77,column=4,value=OnlineShopping[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=77,column=4,value=0)
     try:
       sumr.cell(row=77,column=5,value=OnlineShopping[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=77,column=5,value=0)
     try:
       sumr.cell(row=77,column=6,value=OnlineShopping[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=77,column=6,value=0)
     try:
       sumr.cell(row=77,column=7,value=OnlineShopping[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=77,column=7,value=0)
     try:
       sumr.cell(row=77,column=8,value=OnlineShopping[sumr.cell(row=13,column=8).value])
     except Exception as e:
       sumr.cell(row=77,column=8,value=0)
     try:
       sumr.cell(row=77,column=9,value=OnlineShopping[sumr.cell(row=13,column=9).value])
     except Exception as e:
       sumr.cell(row=77,column=9,value=0)
     try:
       sumr.cell(row=77,column=10,value=OnlineShopping[sumr.cell(row=13,column=10).value])
     except Exception as e:
       sumr.cell(row=77,column=10,value=0)
     try:
       sumr.cell(row=77,column=11,value=OnlineShopping[sumr.cell(row=13,column=11).value])
     except Exception as e:
       sumr.cell(row=77,column=11,value=0)
     try:
       sumr.cell(row=77,column=12,value=OnlineShopping[sumr.cell(row=13,column=12).value])
     except Exception as e:
       sumr.cell(row=77,column=12,value=0)
     try:
       sumr.cell(row=77,column=13,value=OnlineShopping[sumr.cell(row=13,column=13).value])
     except Exception as e:
       sumr.cell(row=77,column=13,value=0)
     try:
       sumr.cell(row=77,column=14,value=OnlineShopping[sumr.cell(row=13,column=14).value])
     except Exception as e:
       sumr.cell(row=77,column=14,value=0)



     try:
       sumr.cell(row=79,column=3,value=GasPayment[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=79,column=3,value=0)
     try:
       sumr.cell(row=79,column=4,value=GasPayment[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=79,column=4,value=0)
     try:
       sumr.cell(row=79,column=5,value=GasPayment[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=79,column=5,value=0)
     try:
       sumr.cell(row=79,column=6,value=GasPayment[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=79,column=6,value=0)
     try:
       sumr.cell(row=79,column=7,value=GasPayment[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=79,column=7,value=0)
     try:
       sumr.cell(row=79,column=8,value=GasPayment[sumr.cell(row=13,column=8).value])
     except Exception as e:
       sumr.cell(row=79,column=8,value=0)
     try:
       sumr.cell(row=79,column=9,value=GasPayment[sumr.cell(row=13,column=9).value])
     except Exception as e:
       sumr.cell(row=79,column=9,value=0)
     try:
       sumr.cell(row=79,column=10,value=GasPayment[sumr.cell(row=13,column=10).value])
     except Exception as e:
       sumr.cell(row=79,column=10,value=0)
     try:
       sumr.cell(row=79,column=11,value=GasPayment[sumr.cell(row=13,column=11).value])
     except Exception as e:
       sumr.cell(row=79,column=11,value=0)
     try:
       sumr.cell(row=79,column=12,value=GasPayment[sumr.cell(row=13,column=12).value])
     except Exception as e:
       sumr.cell(row=79,column=12,value=0)
     try:
       sumr.cell(row=79,column=13,value=GasPayment[sumr.cell(row=13,column=13).value])
     except Exception as e:
       sumr.cell(row=79,column=13,value=0)
     try:
       sumr.cell(row=79,column=14,value=GasPayment[sumr.cell(row=13,column=14).value])
     except Exception as e:
       sumr.cell(row=79,column=14,value=0)




     try:
       sumr.cell(row=80,column=3,value=GoldLoan[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=80,column=3,value=0)
     try:
       sumr.cell(row=80,column=4,value=GoldLoan[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=80,column=4,value=0)
     try:
       sumr.cell(row=80,column=5,value=GoldLoan[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=80,column=5,value=0)
     try:
       sumr.cell(row=80,column=6,value=GoldLoan[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=80,column=6,value=0)
     try:
       sumr.cell(row=80,column=7,value=GoldLoan[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=80,column=7,value=0)
     try:
       sumr.cell(row=80,column=8,value=GoldLoan[sumr.cell(row=13,column=8).value])
     except Exception as e:
       sumr.cell(row=80,column=8,value=0)
     try:
       sumr.cell(row=80,column=9,value=GoldLoan[sumr.cell(row=13,column=9).value])
     except Exception as e:
       sumr.cell(row=80,column=9,value=0)
     try:
       sumr.cell(row=80,column=10,value=GoldLoan[sumr.cell(row=13,column=10).value])
     except Exception as e:
       sumr.cell(row=80,column=10,value=0)
     try:
       sumr.cell(row=80,column=11,value=GoldLoan[sumr.cell(row=13,column=11).value])
     except Exception as e:
       sumr.cell(row=80,column=11,value=0)
     try:
       sumr.cell(row=80,column=12,value=GoldLoan[sumr.cell(row=13,column=12).value])
     except Exception as e:
       sumr.cell(row=80,column=12,value=0)
     try:
       sumr.cell(row=80,column=13,value=GoldLoan[sumr.cell(row=13,column=13).value])
     except Exception as e:
       sumr.cell(row=80,column=13,value=0)
     try:
       sumr.cell(row=80,column=14,value=GoldLoan[sumr.cell(row=13,column=14).value])
     except Exception as e:
       sumr.cell(row=80,column=14,value=0)



     try:
       sumr.cell(row=81,column=3,value=rent_dr[sumr.cell(row=13,column=3).value])
     except Exception as e:
      sumr.cell(row=81,column=3,value=0)
     try:
       sumr.cell(row=81,column=4,value=rent_dr[sumr.cell(row=13,column=4).value])
     except Exception as e:
      sumr.cell(row=81,column=4,value=0)
     try:
       sumr.cell(row=81,column=5,value=rent_dr[sumr.cell(row=13,column=5).value])
     except Exception as e:
      sumr.cell(row=81,column=5,value=0)
     try:
       sumr.cell(row=81,column=6,value=rent_dr[sumr.cell(row=13,column=6).value])
     except Exception as e:
      sumr.cell(row=81,column=6,value=0)
     try:
       sumr.cell(row=81,column=7,value=rent_dr[sumr.cell(row=13,column=7).value])
     except Exception as e:
      sumr.cell(row=81,column=7,value=0)
     try:
       sumr.cell(row=81,column=8,value=rent_dr[sumr.cell(row=13,column=8).value])
     except Exception as e:
       sumr.cell(row=81,column=8,value=0)
     try:
       sumr.cell(row=81,column=9,value=rent_dr[sumr.cell(row=13,column=9).value])
     except Exception as e:
       sumr.cell(row=81,column=9,value=0)
     try:
       sumr.cell(row=81,column=10,value=rent_dr[sumr.cell(row=13,column=10).value])
     except Exception as e:
       sumr.cell(row=81,column=10,value=0)
     try:
       sumr.cell(row=81,column=11,value=rent_dr[sumr.cell(row=13,column=11).value])
     except Exception as e:
       sumr.cell(row=81,column=11,value=0)
     try:
       sumr.cell(row=81,column=12,value=rent_dr[sumr.cell(row=13,column=12).value])
     except Exception as e:
       sumr.cell(row=81,column=12,value=0)
     try:
       sumr.cell(row=81,column=13,value=rent_dr[sumr.cell(row=13,column=13).value])
     except Exception as e:
       sumr.cell(row=81,column=13,value=0)
     try:
       sumr.cell(row=81,column=14,value=rent_dr[sumr.cell(row=13,column=14).value])
     except Exception as e:
       sumr.cell(row=81,column=14,value=0)

     try:
         sumr.cell(row=82,column=3,value=medical_charges[sumr.cell(row=13,column=3).value])
     except Exception as e:
       sumr.cell(row=82,column=3,value=0)
     try:
        sumr.cell(row=82,column=4,value=medical_charges[sumr.cell(row=13,column=4).value])
     except Exception as e:
       sumr.cell(row=82,column=4,value=0)
     try:
        sumr.cell(row=82,column=5,value=medical_charges[sumr.cell(row=13,column=5).value])
     except Exception as e:
       sumr.cell(row=82,column=5,value=0)
     try:
        sumr.cell(row=82,column=6,value=medical_charges[sumr.cell(row=13,column=6).value])
     except Exception as e:
       sumr.cell(row=82,column=6,value=0)
     try:
        sumr.cell(row=82,column=7,value=medical_charges[sumr.cell(row=13,column=7).value])
     except Exception as e:
       sumr.cell(row=82,column=7,value=0)
     try:
        sumr.cell(row=82,column=8,value=medical_charges[sumr.cell(row=13,column=8).value])
     except Exception as e:
       sumr.cell(row=82,column=8,value=0)
     try:
        sumr.cell(row=82,column=9,value=medical_charges[sumr.cell(row=13,column=9).value])
     except Exception as e:
       sumr.cell(row=82,column=9,value=0)
     try:
        sumr.cell(row=82,column=10,value=medical_charges[sumr.cell(row=13,column=10).value])
     except Exception as e:
       sumr.cell(row=82,column=10,value=0)
     try:
        sumr.cell(row=82,column=11,value=medical_charges[sumr.cell(row=13,column=11).value])
     except Exception as e:
       sumr.cell(row=82,column=11,value=0)
     try:
        sumr.cell(row=82,column=12,value=medical_charges[sumr.cell(row=13,column=12).value])
     except Exception as e:
       sumr.cell(row=82,column=12,value=0)
     try:
        sumr.cell(row=82,column=13,value=medical_charges[sumr.cell(row=13,column=13).value])
     except Exception as e:
       sumr.cell(row=82,column=13,value=0)
     try:
        sumr.cell(row=82,column=14,value=medical_charges[sumr.cell(row=13,column=14).value])
     except Exception as e:
       sumr.cell(row=82,column=14,value=0)
   
      

     wb.save('Excel_Files/Dashboard/BankStatement.xlsx')

perExpenditure()

def lastrows():
  # Sum_of_credits = df1.groupby(df1['Tran Date'].dt.strftime('%b-%Y'))['Credit'].sum().sort_values()
  Sum_of_credits =df1['Credit'].sum()

  Nach = df1[df1['Particulars'].str.contains('ECS') | df1['Particulars'].str.contains('NACH-CR')]
  # Nach_cr = Nach.groupby(Nach['Tran Date'].dt.strftime('%b-%Y'))['Credit'].sum().sort_values()
  Nach_cr = Nach['Credit'].sum()
  
  loan_df =df1[df1['Particulars'].str.contains('loan')]
  # loan = loan_df.groupby(loan_df['Tran Date'].dt.strftime('%b-%Y'))['Credit'].sum().sort_values()
  loan = loan_df['Credit'].sum()

  df_grouped =df1[df1["Particulars"].str.contains("Int.Pd")]
  # Interest=df_grouped.groupby(df_grouped['Tran Date'].dt.strftime('%b-%Y'))['Credit'].sum().sort_values()
  Interest = df_grouped['Credit'].sum()

  IT_Nach = df1[df1['Particulars'].str.contains('ACH-CR')&df1['Particulars'].str.contains('AY2018-19-NACH',na=False) | df1['Particulars'].str.contains('ACH-CR')&df1['Particulars'].str.contains('AY2019-20-NACH',na=False)| df1['Particulars'].str.contains('ACH-CR')&df1['Particulars'].str.contains('AY2020-21-NACH',na=False)| df1['Particulars'].str.contains('ACH-CR')&df1['Particulars'].str.contains('AY2021-22-NACH',na=False)| df1['Particulars'].str.contains('ACH-CR')&df1['Particulars'].str.contains('AY2022-23-NACH',na=False)| df1['Particulars'].str.contains('ACH-CR')&df1['Particulars'].str.contains('AY2023-24-NACH',na=False)| df1['Particulars'].str.contains('ACH-CR')&df1['Particulars'].str.contains('AY2024-25-NACH',na=False)]
  # IT_Nach_rec=IT_Nach.groupby(IT_Nach['Tran Date'].dt.strftime('%b-%Y'))['Credit'].sum().sort_values() 
  IT_Nach_rec = IT_Nach['Credit'].sum() 

  dividend_df = df1[df1['Particulars'].str.contains('dividend')]
  # dividend = dividend_df.groupby(dividend_df['Tran Date'].dt.strftime('%b-%Y'))['Credit'].sum().sort_values()
  dividend = dividend_df['Credit'].sum() 

  df_salary = df1[df1["Particulars"].str.contains("SALARY",na=False)]
  # salary=df_salary.groupby(df_salary['Tran Date'].dt.strftime('%b-%Y'))['Credit'].sum().sort_values()
  salary = df_salary['Credit'].sum()

  new=(Nach_cr-loan-Interest-IT_Nach_rec-dividend-salary)
  total_sale = (Sum_of_credits+new)
  # print(loan)
  # print(Sum_of_credits)
  print(total_sale)

  # print(Nach_cr.isna().any())
  # print(loan.isna().any())
  # print(Interest.isna().any())
  # print(IT_Nach_rec.isna().any())
  # print(dividend.isna().any())
  # print(salary.isna().any())

lastrows()

def lastrows1():


    wb = load_workbook('Excel_Files/Dashboard/BankStatement.xlsx')
    sumr= wb['summary']
    
    # Total Sales (receipts) -->

    sumr['C87'] = ((sumr['C14'].value) - (sumr['C29'].value) - (sumr['C30'].value) - (sumr['C27'].value) - (sumr['C31'].value) - (sumr['C32'].value) - (sumr['C28'].value))
    sumr['D87'] = ((sumr['D14'].value) - (sumr['D29'].value) - (sumr['D30'].value) - (sumr['D27'].value) - (sumr['D31'].value) - (sumr['D32'].value) - (sumr['D28'].value))
    sumr['E87'] = ((sumr['E14'].value) - (sumr['E29'].value) - (sumr['E30'].value) - (sumr['E27'].value) - (sumr['E31'].value) - (sumr['E32'].value) - (sumr['E28'].value))
    sumr['F87'] = ((sumr['F14'].value) - (sumr['F29'].value) - (sumr['F30'].value) - (sumr['F27'].value) - (sumr['F31'].value) - (sumr['F32'].value) - (sumr['F28'].value))
    sumr['G87'] = ((sumr['G14'].value) - (sumr['G29'].value) - (sumr['G30'].value) - (sumr['G27'].value) - (sumr['G31'].value) - (sumr['G32'].value) - (sumr['G28'].value))
    sumr['H87'] = ((sumr['H14'].value) - (sumr['H29'].value) - (sumr['H30'].value) - (sumr['H27'].value) - (sumr['H31'].value) - (sumr['H32'].value) - (sumr['H28'].value))
    sumr['I87'] = ((sumr['I14'].value) - (sumr['I29'].value) - (sumr['I30'].value) - (sumr['I27'].value) - (sumr['I31'].value) - (sumr['I32'].value) - (sumr['I28'].value))
    sumr['J87'] = ((sumr['J14'].value) - (sumr['J29'].value) - (sumr['J30'].value) - (sumr['J27'].value) - (sumr['J31'].value) - (sumr['J32'].value) - (sumr['J28'].value))
    sumr['K87'] = ((sumr['K14'].value) - (sumr['K29'].value) - (sumr['K30'].value) - (sumr['K27'].value) - (sumr['K31'].value) - (sumr['K32'].value) - (sumr['K28'].value))
    sumr['L87'] = ((sumr['L14'].value) - (sumr['L29'].value) - (sumr['L30'].value) - (sumr['L27'].value) - (sumr['L31'].value) - (sumr['L32'].value) - (sumr['L28'].value))
    sumr['M87'] = ((sumr['M14'].value) - (sumr['M29'].value) - (sumr['M30'].value) - (sumr['M27'].value) - (sumr['M31'].value) - (sumr['M32'].value) - (sumr['M28'].value))
    sumr['N87'] = ((sumr['N14'].value) - (sumr['N29'].value) - (sumr['N30'].value) - (sumr['N27'].value) - (sumr['N31'].value) - (sumr['N32'].value) - (sumr['N28'].value))
    

    # Total Purchase (payments) -->

    sumr['C88'] = ((sumr['C63'].value) - (sumr['C16'].value) - (sumr['C41'].value) - (sumr['C73'].value) - (sumr['C74'].value) - (sumr['C43'].value) - (sumr['C69'].value) - (sumr['C19'].value) - (sumr['C40'].value) - (sumr['C39'].value) - (sumr['C44'].value) - (sumr['C45'].value) - (sumr['C72'].value) - (sumr['C73'].value) - (sumr['C74'].value) - (sumr['C75'].value) - (sumr['C76'].value) - (sumr['C77'].value) - (sumr['C78'].value))
    sumr['D88'] = ((sumr['D63'].value) - (sumr['D16'].value) - (sumr['D41'].value) - (sumr['D73'].value) - (sumr['D74'].value) - (sumr['D43'].value) - (sumr['D69'].value) - (sumr['D19'].value) - (sumr['D40'].value) - (sumr['D39'].value) - (sumr['D44'].value) - (sumr['D45'].value) - (sumr['D72'].value) - (sumr['D73'].value) - (sumr['D74'].value) - (sumr['D75'].value) - (sumr['D76'].value) - (sumr['D77'].value) - (sumr['D78'].value))
    sumr['E88'] = ((sumr['E63'].value) - (sumr['E16'].value) - (sumr['E41'].value) - (sumr['E73'].value) - (sumr['E74'].value) - (sumr['E43'].value) - (sumr['E69'].value) - (sumr['E19'].value) - (sumr['E40'].value) - (sumr['E39'].value) - (sumr['E44'].value) - (sumr['E45'].value) - (sumr['E72'].value) - (sumr['E73'].value) - (sumr['E74'].value) - (sumr['E75'].value) - (sumr['E76'].value) - (sumr['E77'].value) - (sumr['E78'].value))
    sumr['F88'] = ((sumr['F63'].value) - (sumr['F16'].value) - (sumr['F41'].value) - (sumr['F73'].value) - (sumr['F74'].value) - (sumr['F43'].value) - (sumr['F69'].value) - (sumr['F19'].value) - (sumr['F40'].value) - (sumr['F39'].value) - (sumr['F44'].value) - (sumr['F45'].value) - (sumr['F72'].value) - (sumr['F73'].value) - (sumr['F74'].value) - (sumr['F75'].value) - (sumr['F76'].value) - (sumr['F77'].value) - (sumr['F78'].value))
    sumr['G88'] = ((sumr['G63'].value) - (sumr['G16'].value) - (sumr['G41'].value) - (sumr['G73'].value) - (sumr['G74'].value) - (sumr['G43'].value) - (sumr['G69'].value) - (sumr['G19'].value) - (sumr['G40'].value) - (sumr['G39'].value) - (sumr['G44'].value) - (sumr['G45'].value) - (sumr['G72'].value) - (sumr['G73'].value) - (sumr['G74'].value) - (sumr['G75'].value) - (sumr['G76'].value) - (sumr['G77'].value) - (sumr['G78'].value))
    sumr['H88'] = ((sumr['H63'].value) - (sumr['H16'].value) - (sumr['H41'].value) - (sumr['H73'].value) - (sumr['H74'].value) - (sumr['H43'].value) - (sumr['H69'].value) - (sumr['H19'].value) - (sumr['H40'].value) - (sumr['H39'].value) - (sumr['H44'].value) - (sumr['H45'].value) - (sumr['H72'].value) - (sumr['H73'].value) - (sumr['H74'].value) - (sumr['H75'].value) - (sumr['H76'].value) - (sumr['H77'].value) - (sumr['H78'].value))
    sumr['I88'] = ((sumr['I63'].value) - (sumr['I16'].value) - (sumr['I41'].value) - (sumr['I73'].value) - (sumr['I74'].value) - (sumr['I43'].value) - (sumr['I69'].value) - (sumr['I19'].value) - (sumr['I40'].value) - (sumr['I39'].value) - (sumr['I44'].value) - (sumr['I45'].value) - (sumr['I72'].value) - (sumr['I73'].value) - (sumr['I74'].value) - (sumr['I75'].value) - (sumr['I76'].value) - (sumr['I77'].value) - (sumr['I78'].value))
    sumr['J88'] = ((sumr['J63'].value) - (sumr['J16'].value) - (sumr['J41'].value) - (sumr['J73'].value) - (sumr['J74'].value) - (sumr['J43'].value) - (sumr['J69'].value) - (sumr['J19'].value) - (sumr['J40'].value) - (sumr['J39'].value) - (sumr['J44'].value) - (sumr['J45'].value) - (sumr['J72'].value) - (sumr['J73'].value) - (sumr['J74'].value) - (sumr['J75'].value) - (sumr['J76'].value) - (sumr['J77'].value) - (sumr['J78'].value))
    sumr['K88'] = ((sumr['K63'].value) - (sumr['K16'].value) - (sumr['K41'].value) - (sumr['K73'].value) - (sumr['K74'].value) - (sumr['K43'].value) - (sumr['K69'].value) - (sumr['K19'].value) - (sumr['K40'].value) - (sumr['K39'].value) - (sumr['K44'].value) - (sumr['K45'].value) - (sumr['K72'].value) - (sumr['K73'].value) - (sumr['K74'].value) - (sumr['K75'].value) - (sumr['K76'].value) - (sumr['K77'].value) - (sumr['K78'].value))
    sumr['L88'] = ((sumr['L63'].value) - (sumr['L16'].value) - (sumr['L41'].value) - (sumr['L73'].value) - (sumr['L74'].value) - (sumr['L43'].value) - (sumr['L69'].value) - (sumr['L19'].value) - (sumr['L40'].value) - (sumr['L39'].value) - (sumr['L44'].value) - (sumr['L45'].value) - (sumr['L72'].value) - (sumr['L73'].value) - (sumr['L74'].value) - (sumr['L75'].value) - (sumr['L76'].value) - (sumr['L77'].value) - (sumr['L78'].value))
    sumr['M88'] = ((sumr['M63'].value) - (sumr['M16'].value) - (sumr['M41'].value) - (sumr['M73'].value) - (sumr['M74'].value) - (sumr['M43'].value) - (sumr['M69'].value) - (sumr['M19'].value) - (sumr['M40'].value) - (sumr['M39'].value) - (sumr['M44'].value) - (sumr['M45'].value) - (sumr['M72'].value) - (sumr['M73'].value) - (sumr['M74'].value) - (sumr['M75'].value) - (sumr['M76'].value) - (sumr['M77'].value) - (sumr['M78'].value))
    sumr['N88'] = ((sumr['N63'].value) - (sumr['N16'].value) - (sumr['N41'].value) - (sumr['N73'].value) - (sumr['N74'].value) - (sumr['N43'].value) - (sumr['N69'].value) - (sumr['N19'].value) - (sumr['N40'].value) - (sumr['N39'].value) - (sumr['N44'].value) - (sumr['N45'].value) - (sumr['N72'].value) - (sumr['N73'].value) - (sumr['N74'].value) - (sumr['N75'].value) - (sumr['N76'].value) - (sumr['N77'].value) - (sumr['N78'].value))
    
    # Total Profit as per Bank A/c -->

    sumr['C89'] = ((sumr['C87'].value) - ((sumr['C88'].value) + (sumr['C43'].value) + (sumr['C42'].value) + (sumr['C44'].value))) 
    sumr['D89'] = ((sumr['D87'].value) - ((sumr['D88'].value) + (sumr['D43'].value) + (sumr['D42'].value) + (sumr['D44'].value))) 
    sumr['E89'] = ((sumr['E87'].value) - ((sumr['E88'].value) + (sumr['E43'].value) + (sumr['E42'].value) + (sumr['E44'].value)))
    sumr['F89'] = ((sumr['F87'].value) - ((sumr['F88'].value) + (sumr['F43'].value) + (sumr['F42'].value) + (sumr['F44'].value))) 
    sumr['G89'] = ((sumr['G87'].value) - ((sumr['G88'].value) + (sumr['G43'].value) + (sumr['G42'].value) + (sumr['G44'].value))) 
    sumr['H89'] = ((sumr['H87'].value) - ((sumr['H88'].value) + (sumr['H43'].value) + (sumr['H42'].value) + (sumr['H44'].value))) 
    sumr['I89'] = ((sumr['I87'].value) - ((sumr['I88'].value) + (sumr['I43'].value) + (sumr['I42'].value) + (sumr['I44'].value))) 
    sumr['J89'] = ((sumr['J87'].value) - ((sumr['J88'].value) + (sumr['J43'].value) + (sumr['J42'].value) + (sumr['J44'].value))) 
    sumr['K89'] = ((sumr['K87'].value) - ((sumr['K88'].value) + (sumr['K43'].value) + (sumr['K42'].value) + (sumr['K44'].value))) 
    sumr['L89'] = ((sumr['L87'].value) - ((sumr['L88'].value) + (sumr['L43'].value) + (sumr['L42'].value) + (sumr['L44'].value))) 
    sumr['M89'] = ((sumr['M87'].value) - ((sumr['M88'].value) + (sumr['M43'].value) + (sumr['M42'].value) + (sumr['M44'].value))) 
    sumr['N89'] = ((sumr['N87'].value) - ((sumr['N88'].value) + (sumr['N43'].value) + (sumr['N42'].value) + (sumr['N44'].value)))
    
    





    sumr['C92'] = ((sumr['C14'].value) - (sumr['C29'].value) - (sumr['C30'].value) - (sumr['C27'].value) - (sumr['C31'].value) - (sumr['C32'].value) - (sumr['C28'].value))
    sumr['D92'] = ((sumr['D14'].value) - (sumr['D29'].value) - (sumr['D30'].value) - (sumr['D27'].value) - (sumr['D31'].value) - (sumr['D32'].value) - (sumr['D28'].value))
    sumr['E92'] = ((sumr['E14'].value) - (sumr['E29'].value) - (sumr['E30'].value) - (sumr['E27'].value) - (sumr['E31'].value) - (sumr['E32'].value) - (sumr['E28'].value))
    sumr['F92'] = ((sumr['F14'].value) - (sumr['F29'].value) - (sumr['F30'].value) - (sumr['F27'].value) - (sumr['F31'].value) - (sumr['F32'].value) - (sumr['F28'].value))
    sumr['G92'] = ((sumr['G14'].value) - (sumr['G29'].value) - (sumr['G30'].value) - (sumr['G27'].value) - (sumr['G31'].value) - (sumr['G32'].value) - (sumr['G28'].value))
    sumr['H92'] = ((sumr['H14'].value) - (sumr['H29'].value) - (sumr['H30'].value) - (sumr['H27'].value) - (sumr['H31'].value) - (sumr['H32'].value) - (sumr['H28'].value))
    sumr['I92'] = ((sumr['I14'].value) - (sumr['I29'].value) - (sumr['I30'].value) - (sumr['I27'].value) - (sumr['I31'].value) - (sumr['I32'].value) - (sumr['I28'].value))
    sumr['J92'] = ((sumr['J14'].value) - (sumr['J29'].value) - (sumr['J30'].value) - (sumr['J27'].value) - (sumr['J31'].value) - (sumr['J32'].value) - (sumr['J28'].value))
    sumr['K92'] = ((sumr['K14'].value) - (sumr['K29'].value) - (sumr['K30'].value) - (sumr['K27'].value) - (sumr['K31'].value) - (sumr['K32'].value) - (sumr['K28'].value))
    sumr['L92'] = ((sumr['L14'].value) - (sumr['L29'].value) - (sumr['L30'].value) - (sumr['L27'].value) - (sumr['L31'].value) - (sumr['L32'].value) - (sumr['L28'].value))
    sumr['M92'] = ((sumr['M14'].value) - (sumr['M29'].value) - (sumr['M30'].value) - (sumr['M27'].value) - (sumr['M31'].value) - (sumr['M32'].value) - (sumr['M28'].value))
    sumr['N92'] = ((sumr['N14'].value) - (sumr['N29'].value) - (sumr['N30'].value) - (sumr['N27'].value) - (sumr['N31'].value) - (sumr['N32'].value) - (sumr['N28'].value))
    

    # Total Purchase (payments) -->

    sumr['C93'] = ((sumr['C63'].value) - (sumr['C16'].value) - (sumr['C41'].value) - (sumr['C73'].value) - (sumr['C74'].value) - (sumr['C43'].value) - (sumr['C69'].value) - (sumr['C19'].value) - (sumr['C40'].value) - (sumr['C39'].value) - (sumr['C44'].value) - (sumr['C45'].value) - (sumr['C72'].value))
    sumr['D93'] = ((sumr['D63'].value) - (sumr['D16'].value) - (sumr['D41'].value) - (sumr['D73'].value) - (sumr['D74'].value) - (sumr['D43'].value) - (sumr['D69'].value) - (sumr['D19'].value) - (sumr['D40'].value) - (sumr['D39'].value) - (sumr['D44'].value) - (sumr['D45'].value) - (sumr['D72'].value))
    sumr['E93'] = ((sumr['E63'].value) - (sumr['E16'].value) - (sumr['E41'].value) - (sumr['E73'].value) - (sumr['E74'].value) - (sumr['E43'].value) - (sumr['E69'].value) - (sumr['E19'].value) - (sumr['E40'].value) - (sumr['E39'].value) - (sumr['E44'].value) - (sumr['E45'].value) - (sumr['E72'].value))
    sumr['F93'] = ((sumr['F63'].value) - (sumr['F16'].value) - (sumr['F41'].value) - (sumr['F73'].value) - (sumr['F74'].value) - (sumr['F43'].value) - (sumr['F69'].value) - (sumr['F19'].value) - (sumr['F40'].value) - (sumr['F39'].value) - (sumr['F44'].value) - (sumr['F45'].value) - (sumr['F72'].value))
    sumr['G93'] = ((sumr['G63'].value) - (sumr['G16'].value) - (sumr['G41'].value) - (sumr['G73'].value) - (sumr['G74'].value) - (sumr['G43'].value) - (sumr['G69'].value) - (sumr['G19'].value) - (sumr['G40'].value) - (sumr['G39'].value) - (sumr['G44'].value) - (sumr['G45'].value) - (sumr['G72'].value))
    sumr['H93'] = ((sumr['H63'].value) - (sumr['H16'].value) - (sumr['H41'].value) - (sumr['H73'].value) - (sumr['H74'].value) - (sumr['H43'].value) - (sumr['H69'].value) - (sumr['H19'].value) - (sumr['H40'].value) - (sumr['H39'].value) - (sumr['H44'].value) - (sumr['H45'].value) - (sumr['H72'].value))
    sumr['I93'] = ((sumr['I63'].value) - (sumr['I16'].value) - (sumr['I41'].value) - (sumr['I73'].value) - (sumr['I74'].value) - (sumr['I43'].value) - (sumr['I69'].value) - (sumr['I19'].value) - (sumr['I40'].value) - (sumr['I39'].value) - (sumr['I44'].value) - (sumr['I45'].value) - (sumr['I72'].value))
    sumr['J93'] = ((sumr['J63'].value) - (sumr['J16'].value) - (sumr['J41'].value) - (sumr['J73'].value) - (sumr['J74'].value) - (sumr['J43'].value) - (sumr['J69'].value) - (sumr['J19'].value) - (sumr['J40'].value) - (sumr['J39'].value) - (sumr['J44'].value) - (sumr['J45'].value) - (sumr['J72'].value))
    sumr['K93'] = ((sumr['K63'].value) - (sumr['K16'].value) - (sumr['K41'].value) - (sumr['K73'].value) - (sumr['K74'].value) - (sumr['K43'].value) - (sumr['K69'].value) - (sumr['K19'].value) - (sumr['K40'].value) - (sumr['K39'].value) - (sumr['K44'].value) - (sumr['K45'].value) - (sumr['K72'].value))
    sumr['L93'] = ((sumr['L63'].value) - (sumr['L16'].value) - (sumr['L41'].value) - (sumr['L73'].value) - (sumr['L74'].value) - (sumr['L43'].value) - (sumr['L69'].value) - (sumr['L19'].value) - (sumr['L40'].value) - (sumr['L39'].value) - (sumr['L44'].value) - (sumr['L45'].value) - (sumr['L72'].value))
    sumr['M93'] = ((sumr['M63'].value) - (sumr['M16'].value) - (sumr['M41'].value) - (sumr['M73'].value) - (sumr['M74'].value) - (sumr['M43'].value) - (sumr['M69'].value) - (sumr['M19'].value) - (sumr['M40'].value) - (sumr['M39'].value) - (sumr['M44'].value) - (sumr['M45'].value) - (sumr['M72'].value))
    sumr['N93'] = ((sumr['N63'].value) - (sumr['N16'].value) - (sumr['N41'].value) - (sumr['N73'].value) - (sumr['N74'].value) - (sumr['N43'].value) - (sumr['N69'].value) - (sumr['N19'].value) - (sumr['N40'].value) - (sumr['N39'].value) - (sumr['N44'].value) - (sumr['N45'].value) - (sumr['N72'].value))
    
    # Total Profit as per Bank A/c -->

    sumr['C94'] = ((sumr['C87'].value) - ((sumr['C88'].value) + (sumr['C43'].value) + (sumr['C42'].value) + (sumr['C44'].value))) 
    sumr['D94'] = ((sumr['D87'].value) - ((sumr['D88'].value) + (sumr['D43'].value) + (sumr['D42'].value) + (sumr['D44'].value))) 
    sumr['E94'] = ((sumr['E87'].value) - ((sumr['E88'].value) + (sumr['E43'].value) + (sumr['E42'].value) + (sumr['E44'].value)))
    sumr['F94'] = ((sumr['F87'].value) - ((sumr['F88'].value) + (sumr['F43'].value) + (sumr['F42'].value) + (sumr['F44'].value))) 
    sumr['G94'] = ((sumr['G87'].value) - ((sumr['G88'].value) + (sumr['G43'].value) + (sumr['G42'].value) + (sumr['G44'].value))) 
    sumr['H94'] = ((sumr['H87'].value) - ((sumr['H88'].value) + (sumr['H43'].value) + (sumr['H42'].value) + (sumr['H44'].value))) 
    sumr['I94'] = ((sumr['I87'].value) - ((sumr['I88'].value) + (sumr['I43'].value) + (sumr['I42'].value) + (sumr['I44'].value))) 
    sumr['J94'] = ((sumr['J87'].value) - ((sumr['J88'].value) + (sumr['J43'].value) + (sumr['J42'].value) + (sumr['J44'].value))) 
    sumr['K94'] = ((sumr['K87'].value) - ((sumr['K88'].value) + (sumr['K43'].value) + (sumr['K42'].value) + (sumr['K44'].value))) 
    sumr['L94'] = ((sumr['L87'].value) - ((sumr['L88'].value) + (sumr['L43'].value) + (sumr['L42'].value) + (sumr['L44'].value))) 
    sumr['M94'] = ((sumr['M87'].value) - ((sumr['M88'].value) + (sumr['M43'].value) + (sumr['M42'].value) + (sumr['M44'].value))) 
    sumr['N94'] = ((sumr['N87'].value) - ((sumr['N88'].value) + (sumr['N43'].value) + (sumr['N42'].value) + (sumr['N44'].value)))
    
    
    wb.save('Excel_Files/Dashboard/BankStatement.xlsx')

lastrows1()

def sum():

    wb = load_workbook('Excel_Files/Dashboard/BankStatement.xlsx')
    sumr= wb['summary']
    


    sumr['C34'] = '= SUM(C26:C33)'
    sumr['D34'] = '= SUM(D26:D33)'
    sumr['E34'] = '= SUM(E26:E33)'
    sumr['F34'] = '= SUM(F26:F33)'
    sumr['G34'] = '= SUM(G26:G33)'
    sumr['H34'] = '= SUM(H26:H33)'
    sumr['I34'] = '= SUM(I26:I33)'
    sumr['J34'] = '= SUM(J26:J33)'
    sumr['K34'] = '= SUM(K26:K33)'
    sumr['L34'] = '= SUM(L26:L33)'
    sumr['M34'] = '= SUM(M26:M33)'
    sumr['N34'] = '= SUM(N26:N33)'


    sumr['C48'] = '= SUM(C37:C47)'
    sumr['D48'] = '= SUM(D37:D47)'
    sumr['E48'] = '= SUM(E37:E47)'
    sumr['F48'] = '= SUM(F37:F47)'
    sumr['G48'] = '= SUM(G37:G47)'
    sumr['H48'] = '= SUM(H37:H47)'
    sumr['I48'] = '= SUM(I37:I47)'
    sumr['J48'] = '= SUM(J37:J47)'
    sumr['K48'] = '= SUM(K37:K47)'
    sumr['L48'] = '= SUM(L37:L47)'
    sumr['M48'] = '= SUM(M37:M47)'
    sumr['N48'] = '= SUM(N37:N47)'


    sumr['C60'] = '= SUM(C51:C59)'
    sumr['D60'] = '= SUM(D51:D59)'
    sumr['E60'] = '= SUM(E51:E59)'
    sumr['F60'] = '= SUM(F51:F59)'
    sumr['G60'] = '= SUM(G51:G59)'
    sumr['H60'] = '= SUM(H51:H59)'
    sumr['I60'] = '= SUM(I51:I59)'
    sumr['J60'] = '= SUM(J51:J59)'
    sumr['K60'] = '= SUM(K51:K59)'
    sumr['L60'] = '= SUM(L51:L59)'
    sumr['M60'] = '= SUM(M51:M59)'
    sumr['N60'] = '= SUM(N51:N59)'


    sumr['C82'] = '= SUM(C63:C81)'
    sumr['D82'] = '= SUM(D63:D81)'
    sumr['E82'] = '= SUM(E63:E81)'
    sumr['F82'] = '= SUM(F63:F81)'
    sumr['G82'] = '= SUM(G63:G81)'
    sumr['H82'] = '= SUM(H63:H81)'
    sumr['I82'] = '= SUM(I63:I81)'
    sumr['J82'] = '= SUM(J63:J81)'
    sumr['K82'] = '= SUM(K63:K81)'
    sumr['L82'] = '= SUM(L63:L81)'
    sumr['M82'] = '= SUM(M63:M81)'
    sumr['N82'] = '= SUM(N63:N81)'
    



    wb.save('Excel_Files/Dashboard/BankStatement.xlsx')

sum()

df1

with pd.ExcelWriter('Excel_Files/Dashboard/BankStatement.xlsx',mode='a') as writer:  
   df1.to_excel(writer, sheet_name='Transaction1',index=False)

#wb = load_workbook('Excel_Files/Dashboard/BankStatement.xlsx')
#tran=wb['Transaction']
#wb.remove(tran)

# wb = load_workbook('Excel_Files/Dashboard/BankStatement.xlsx')


# sheet_names = wb.sheetnames

# # Print the list of sheet names
# for name in sheet_names:
#     print(name)

import openpyxl
import openpyxl

# load the workbook
wb = openpyxl.load_workbook('Excel_Files/Dashboard/BankStatement.xlsx')

# specify the name of the worksheet to move
sheet_name = 'Transaction1'

# specify the new index for the worksheet
new_index = 1

# get the worksheet object
ws = wb[sheet_name]

# remove the worksheet from its current position
wb.remove(ws)

# insert the worksheet at the new index
wb.create_sheet(sheet_name, new_index)

# copy data from old sheet to new sheet
for row in ws.iter_rows():
    for cell in row:
        wb[sheet_name][cell.coordinate].value = cell.value

# save the changes
wb.save('Excel_Files/Dashboard/BankStatement.xlsx')

import openpyxl

# Load the Excel file
workbook = openpyxl.load_workbook('Excel_Files/Dashboard/BankStatement.xlsx')

# Get the sheet to be removed
sheet_to_remove = workbook['Transaction']

# Remove the sheet
workbook.remove(sheet_to_remove)

# Save the changes
workbook.save('Excel_Files/Dashboard/BankStatement.xlsx')

import pandas as pd

# Load the Excel file
df = pd.read_excel('Excel_Files/Dashboard/BankStatement.xlsx', sheet_name='Transaction1')

# Select column 'A'
column_Tran_Date = df['Tran Date']

# Convert the Series to a DataFrame
column_Tran_Date_df = pd.DataFrame(column_Tran_Date)

column_Tran_Date_df

# rent_df = df1[df1['Particulars'].str.contains('RENT') | df1['Particulars'].str.contains('Rent')]
# rent_dr = rent_df.groupby(rent_df['Tran Date'].dt.strftime('%b-%Y'))['Debit'].sum().sort_values()
# RENT1 = df1[df1["Particulars"].str.contains("RENT")| df1['Particulars'].str.contains('Rent')].groupby('Debit')
# RENT1 = RENT1.apply(lambda x: x)
# RENT1['Categories'] = 'RENT Paid'
# df1.update(RENT1)

# rent_df = df1[df1['Particulars'].str.contains('RENT') | df1['Particulars'].str.contains('Rent')]

# rent_df

# df1.update(RENT1)

# df1.head(197)

cr = df1['Credit'].sum()